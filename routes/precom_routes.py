from flask import Blueprint, request, render_template, jsonify, redirect, url_for, send_file, current_app
from flask_babel import gettext as _
from database import create_connection, ensure_precom_tables
from models.system import SystemModel
from models.subsystem import SubsystemModel
from werkzeug.utils import secure_filename
from datetime import datetime
from io import BytesIO
import os
import re
import pandas as pd


precom_bp = Blueprint('precom', __name__)

PRECOM_UPLOAD_FOLDER = 'uploads/precom_tasks'
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'xlsx', 'xls', 'doc', 'docx', 'zip'}

from importlib import import_module

try:
    # 通过动态导入避免静态检查器对可选依赖报错
    magic = import_module("magic")  # type: ignore[assignment]
except ImportError:
    magic = None

# 使用gettext进行翻译
def get_task_type_label(task_type):
    """获取任务类型的翻译标签"""
    labels = {
        'Manhole': _('人孔检查'),
        'MotorSolo': _('电机单试'),
        'SkidInstall': _('仪表台件安装'),
        'LoopTest': _('回路测试'),
        'Alignment': _('动设备最终对中'),
        'MRT': _('MRT机械联动测试'),
        'FunctionTest': 'Function Test',
    }
    return labels.get(task_type, _('预试车任务管理'))


def _allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _validate_precom_upload_file(file) -> tuple[bool, str | None]:
    """验证预试车附件上传文件的扩展名、MIME 和大小。"""
    if not file or not file.filename:
        return False, "未选择文件"

    if not _allowed_file(file.filename):
        return False, "不允许的文件类型"

    max_size_bytes = 50 * 1024 * 1024
    try:
        pos = file.stream.tell()
        file.stream.seek(0, os.SEEK_END)
        size = file.stream.tell()
        file.stream.seek(pos)
    except Exception:
        size = None
    if size is not None and size > max_size_bytes:
        return False, "单个文件大小不能超过 50MB"

    if magic is not None:
        try:
            pos = file.stream.tell()
            sample = file.stream.read(2048)
            file.stream.seek(pos)
            mime_type = magic.from_buffer(sample, mime=True)
        except Exception:
            mime_type = None

        if mime_type:
            allowed_mimes = {
                'application/pdf',
                'image/png',
                'image/jpeg',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'application/vnd.ms-excel',
                'application/msword',
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'application/zip',
            }
            if mime_type not in allowed_mimes:
                return False, "文件内容与扩展名不匹配或类型不受支持"

    return True, None


def _ensure_upload_folder(task_id: int) -> str:
    folder = os.path.join(PRECOM_UPLOAD_FOLDER, str(task_id))
    os.makedirs(folder, exist_ok=True)
    return folder


def _task_status(row) -> str:
    total = int(row.get('QuantityTotal') or 0)
    done = int(row.get('QuantityDone') or 0)
    if total <= 0:
        return '未开始'
    if done <= 0:
        return '未开始'
    if done < total:
        return '进行中'
    return '已完成'


@precom_bp.route('/precom/tasks')
def precom_task_list():
    task_type = (request.args.get('task_type') or '').strip() or None
    system_code = (request.args.get('system_code') or '').strip() or None
    subsystem_code = (request.args.get('subsystem_code') or '').strip() or None
    block = (request.args.get('block') or '').strip() or None
    status_filter = (request.args.get('status') or '').strip() or None

    systems = SystemModel.get_all_systems()
    subsystems = SubsystemModel.get_all_subsystems() if not system_code else SubsystemModel.get_subsystems_by_system(system_code)

    page_title = get_task_type_label(task_type)

    conn = create_connection()
    tasks = []
    if conn:
        try:
            cur = conn.cursor(dictionary=True)
            clauses = []
            params = []
            if task_type:
                clauses.append("TaskType = %s")
                params.append(task_type)
            if system_code:
                clauses.append("SystemCode = %s")
                params.append(system_code)
            if subsystem_code:
                clauses.append("SubSystemCode = %s")
                params.append(subsystem_code)
            if block:
                clauses.append("PositionBlock = %s")
                params.append(block)
            where_sql = " AND ".join(clauses) if clauses else "1=1"
            cur.execute(
                f"""
                SELECT TaskID, TaskType, SystemCode, SubSystemCode, DrawingNumber,
                       TagNumber, PointTag, Description, PositionBlock,
                       QuantityTotal, QuantityDone, PlannedDate, ActualDate, PerformedBy, TestType
                FROM PrecomTask
                WHERE {where_sql}
                ORDER BY SubSystemCode, TaskType, TagNumber, PointTag
                """,
                tuple(params),
            )
            rows = cur.fetchall()
            for row in rows:
                row['Status'] = _task_status(row)
                if status_filter and row['Status'] != status_filter:
                    continue
                tasks.append(row)
        finally:
            conn.close()

    return render_template(
        'precom_task_list.html',
        tasks=tasks,
        systems=systems,
        subsystems=subsystems,
        filter_task_type=task_type or '',
        filter_system=system_code or '',
        filter_subsystem=subsystem_code or '',
        filter_block=block or '',
        filter_status=status_filter or '',
        page_title=page_title,
        new_task_url=url_for('precom.precom_task_new', task_type=task_type) if task_type else url_for('precom.precom_task_new'),
    )


@precom_bp.route('/precom/manhole')
def precom_manhole_entry():
    return redirect(url_for('precom.precom_task_list', task_type='Manhole'))


@precom_bp.route('/precom/motor_solo')
def precom_motor_solo_entry():
    return redirect(url_for('precom.precom_task_list', task_type='MotorSolo'))


@precom_bp.route('/precom/skid_install')
def precom_skid_install_entry():
    return redirect(url_for('precom.precom_task_list', task_type='SkidInstall'))


@precom_bp.route('/precom/loop_test')
def precom_loop_test_entry():
    return redirect(url_for('precom.precom_task_list', task_type='LoopTest'))


@precom_bp.route('/precom/alignment')
def precom_alignment_entry():
    return redirect(url_for('precom.precom_task_list', task_type='Alignment'))


@precom_bp.route('/precom/mrt')
def precom_mrt_entry():
    return redirect(url_for('precom.precom_task_list', task_type='MRT'))


@precom_bp.route('/precom/function_test')
def precom_function_test_entry():
    return redirect(url_for('precom.precom_task_list', task_type='FunctionTest'))


@precom_bp.route('/precom/tasks/new', methods=['GET', 'POST'])
def precom_task_new():
    task_type = (request.args.get('task_type') or '').strip() or 'Manhole'
    systems = SystemModel.get_all_systems()
    subsystems = []
    if request.method == 'POST':
        system_code = request.form.get('SystemCode') or None
        subsystem_code = request.form.get('SubSystemCode') or None
        if system_code:
            subsystems = SubsystemModel.get_subsystems_by_system(system_code)
        data = {
            'TaskType': task_type,
            'SystemCode': system_code,
            'SubSystemCode': subsystem_code,
            'DrawingNumber': _join_drawing_numbers_from_request(),
            'TagNumber': request.form.get('TagNumber') or None,
            'PointTag': request.form.get('PointTag') or None,
            'Description': request.form.get('Description') or None,
            'PositionBlock': _join_position_blocks_from_request(),
            'QuantityTotal': int(request.form.get('QuantityTotal') or 1),
            'QuantityDone': int(request.form.get('QuantityDone') or 0),
            'PlannedDate': request.form.get('PlannedDate') or None,
            'ActualDate': request.form.get('ActualDate') or None,
            'PerformedBy': request.form.get('PerformedBy') or None,
            'TestType': request.form.get('TestType') or None,
            # 施工进度汇总字段（预留，可选）
            'ProgressID': request.form.get('ProgressID') or None,
            'Discipline': request.form.get('Discipline') or None,
            'WorkPackage': request.form.get('WorkPackage') or None,
            'KeyQuantityTotal': int(request.form.get('KeyQuantityTotal') or 0),
            'KeyQuantityDone': int(request.form.get('KeyQuantityDone') or 0),
            'KeyProgressPercent': (
                float(request.form.get('KeyProgressPercent') or 0)
                if request.form.get('KeyProgressPercent')
                else None
            ),
        }
        activities = _parse_activity_rows_from_request()
        conn = create_connection()
        if not conn:
            return "数据库连接失败", 500
        try:
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO PrecomTask (
                    TaskType, SystemCode, SubSystemCode, DrawingNumber,
                    TagNumber, PointTag, Description, PositionBlock,
                    QuantityTotal, QuantityDone, PlannedDate, ActualDate,
                    PerformedBy, TestType,
                    ProgressID, Discipline, WorkPackage,
                    KeyQuantityTotal, KeyQuantityDone, KeyProgressPercent
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    data['TaskType'],
                    data['SystemCode'],
                    data['SubSystemCode'],
                    data['DrawingNumber'],
                    data['TagNumber'],
                    data['PointTag'],
                    data['Description'],
                    data['PositionBlock'],
                    data['QuantityTotal'],
                    data['QuantityDone'],
                    data['PlannedDate'],
                    data['ActualDate'],
                    data['PerformedBy'],
                    data['TestType'],
                    data['ProgressID'],
                    data['Discipline'],
                    data['WorkPackage'],
                    data['KeyQuantityTotal'],
                    data['KeyQuantityDone'],
                    data['KeyProgressPercent'],
                ),
            )
            task_id = cur.lastrowid
            _save_task_activities(conn, task_id, activities)
            conn.commit()
            return redirect(url_for('precom.precom_task_edit', task_id=task_id))
        finally:
            conn.close()
    else:
        system_code = request.args.get('system_code') or None
        if system_code:
            subsystems = SubsystemModel.get_subsystems_by_system(system_code)

    return render_template(
        'precom_task_edit.html',
        task=None,
        task_type=task_type,
        systems=systems,
        subsystems=subsystems,
        block_rows=_split_position_blocks(None),
        drawing_rows=_split_drawing_numbers(None),
        activities=[],
    )


@precom_bp.route('/precom/tasks/<int:task_id>/edit', methods=['GET', 'POST'])
def precom_task_edit(task_id: int):
    conn = create_connection()
    if not conn:
        return "数据库连接失败", 500
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM PrecomTask WHERE TaskID = %s", (task_id,))
        task = cur.fetchone()
        if not task:
            return "未找到任务", 404

        systems = SystemModel.get_all_systems()
        subsystems = SubsystemModel.get_subsystems_by_system(task.get('SystemCode')) if task.get('SystemCode') else []

        if request.method == 'POST':
            data = {
                'SystemCode': request.form.get('SystemCode') or None,
                'SubSystemCode': request.form.get('SubSystemCode') or None,
                'DrawingNumber': _join_drawing_numbers_from_request(),
                'TagNumber': request.form.get('TagNumber') or None,
                'PointTag': request.form.get('PointTag') or None,
                'Description': request.form.get('Description') or None,
                'PositionBlock': _join_position_blocks_from_request(),
                'QuantityTotal': int(request.form.get('QuantityTotal') or 1),
                'QuantityDone': int(request.form.get('QuantityDone') or 0),
                'PlannedDate': request.form.get('PlannedDate') or None,
                'ActualDate': request.form.get('ActualDate') or None,
                'PerformedBy': request.form.get('PerformedBy') or None,
                'TestType': request.form.get('TestType') or None,
                # 施工进度汇总字段（预留，可选）
                'ProgressID': request.form.get('ProgressID') or None,
                'Discipline': request.form.get('Discipline') or None,
                'WorkPackage': request.form.get('WorkPackage') or None,
                'KeyQuantityTotal': int(request.form.get('KeyQuantityTotal') or 0),
                'KeyQuantityDone': int(request.form.get('KeyQuantityDone') or 0),
                'KeyProgressPercent': (
                    float(request.form.get('KeyProgressPercent') or 0)
                    if request.form.get('KeyProgressPercent')
                    else None
                ),
            }
            activities = _parse_activity_rows_from_request()
            cur_update = conn.cursor()
            cur_update.execute(
                """
                UPDATE PrecomTask
                SET SystemCode=%s, SubSystemCode=%s, DrawingNumber=%s,
                    TagNumber=%s, PointTag=%s, Description=%s, PositionBlock=%s,
                    QuantityTotal=%s, QuantityDone=%s, PlannedDate=%s, ActualDate=%s,
                    PerformedBy=%s, TestType=%s,
                    ProgressID=%s, Discipline=%s, WorkPackage=%s,
                    KeyQuantityTotal=%s, KeyQuantityDone=%s, KeyProgressPercent=%s
                WHERE TaskID=%s
                """,
                (
                    data['SystemCode'],
                    data['SubSystemCode'],
                    data['DrawingNumber'],
                    data['TagNumber'],
                    data['PointTag'],
                    data['Description'],
                    data['PositionBlock'],
                    data['QuantityTotal'],
                    data['QuantityDone'],
                    data['PlannedDate'],
                    data['ActualDate'],
                    data['PerformedBy'],
                    data['TestType'],
                    data['ProgressID'],
                    data['Discipline'],
                    data['WorkPackage'],
                    data['KeyQuantityTotal'],
                    data['KeyQuantityDone'],
                    data['KeyProgressPercent'],
                    task_id,
                ),
            )
            _save_task_activities(conn, task_id, activities)
            conn.commit()
            return redirect(url_for('precom.precom_task_edit', task_id=task_id))

        # 加载附件和Punch简要信息用于页面展示
        cur.execute(
            """
            SELECT AttachmentID, FileName, FileSize, UploadedBy, UploadedAt, ModuleName
            FROM PrecomTaskAttachment
            WHERE TaskID = %s
            ORDER BY UploadedAt DESC
            """,
            (task_id,),
        )
        attachments = cur.fetchall()

        cur.execute(
            """
            SELECT ID, PunchNo, RefNo, SheetNo, RevNo, Description, Category, Cause,
                   IssuedBy, Rectified, RectifiedDate, Verified, VerifiedDate, Deleted
            FROM PrecomTaskPunch
            WHERE TaskID = %s
            ORDER BY ID
            """,
            (task_id,),
        )
        punches = cur.fetchall()

        # 加载施工活动明细
        cur.execute(
            """
            SELECT ID, ActID, Block, ActDescription, Scope, Discipline,
                   WorkPackage, TotalQuantity, CompletedQuantity, CompletedPercent,
                   WeightFactor, ManHours, Subproject
            FROM PrecomTaskActivity
            WHERE TaskID = %s
            ORDER BY ID
            """,
            (task_id,),
        )
        activities = cur.fetchall()

        return render_template(
            'precom_task_edit.html',
            task=task,
            task_type=task.get('TaskType'),
            systems=systems,
            subsystems=subsystems,
            attachments=attachments,
            punches=punches,
            activities=activities,
            block_rows=_split_position_blocks(task.get('PositionBlock')),
            drawing_rows=_split_drawing_numbers(task.get('DrawingNumber')),
        )
    finally:
        conn.close()


@precom_bp.route('/precom/tasks/<int:task_id>/delete', methods=['POST'])
def precom_task_delete(task_id: int):
    conn = create_connection()
    if not conn:
        return jsonify({'error': '数据库连接失败'}), 500
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM PrecomTask WHERE TaskID = %s", (task_id,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


@precom_bp.route('/api/precom/tasks/<int:task_id>/attachments', methods=['GET'])
def precom_attachments(task_id: int):
    conn = create_connection()
    if not conn:
        return jsonify([])
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT AttachmentID, FileName, FileSize, UploadedBy, UploadedAt, ModuleName
            FROM PrecomTaskAttachment
            WHERE TaskID = %s
            ORDER BY UploadedAt DESC
            """,
            (task_id,),
        )
        rows = cur.fetchall()
        for row in rows:
            if row.get('UploadedAt'):
                row['UploadedAt'] = row['UploadedAt'].strftime('%Y-%m-%d %H:%M:%S')
        return jsonify(rows)
    finally:
        conn.close()


@precom_bp.route('/api/faclist/blocks')
def api_faclist_blocks():
    """返回 Faclist 中所有唯一 Block，用于预试车任务 Block 下拉多选。"""
    conn = create_connection()
    if not conn:
        return jsonify([])
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT DISTINCT Block
            FROM Faclist
            WHERE Block IS NOT NULL AND Block <> ''
            ORDER BY Block
            """
        )
        rows = cur.fetchall()
        blocks = [row[0] for row in rows if row and row[0]]
        return jsonify(blocks)
    finally:
        conn.close()


@precom_bp.route('/api/precom/tasks/<int:task_id>/attachments', methods=['POST'])
def precom_upload_attachment(task_id: int):
    if 'files' not in request.files:
        return jsonify({'error': '未选择文件'}), 400
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': '未选择文件'}), 400

    conn = create_connection()
    if not conn:
        return jsonify({'error': '数据库连接失败'}), 500
    try:
        folder = _ensure_upload_folder(task_id)
        cur = conn.cursor()
        uploaded = []
        for file in files:
            ok, err = _validate_precom_upload_file(file)
            if not ok:
                current_app.logger.warning(f"预试车附件校验失败: {err} (filename={file.filename})")
                continue
            original = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{original}"
            file_path = os.path.join(folder, filename)
            file.save(file_path)
            file_size = os.path.getsize(file_path)
            cur.execute(
                """
                INSERT INTO PrecomTaskAttachment (TaskID, FileName, FilePath, FileSize, UploadedBy)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (task_id, original, file_path, file_size, 'web'),
            )
            uploaded.append({'id': cur.lastrowid, 'filename': original, 'size': file_size})
        conn.commit()
        return jsonify({'success': True, 'files': uploaded})
    except Exception as exc:
        conn.rollback()
        current_app.logger.error(f"上传预试车附件失败: {exc}", exc_info=True)
        return jsonify({'error': '服务器内部错误，请稍后重试'}), 500
    finally:
        conn.close()


@precom_bp.route('/api/precom/tasks/<int:task_id>/attachments/<int:attachment_id>', methods=['DELETE'])
def precom_delete_attachment(task_id: int, attachment_id: int):
    conn = create_connection()
    if not conn:
        return jsonify({'error': '数据库连接失败'}), 500
    try:
        cur = conn.cursor(dictionary=True)
        # 先查询附件信息，包括文件路径
        cur.execute(
            """
            SELECT AttachmentID, FilePath
            FROM PrecomTaskAttachment
            WHERE AttachmentID = %s AND TaskID = %s
            """,
            (attachment_id, task_id)
        )
        attachment = cur.fetchone()
        if not attachment:
            return jsonify({'error': '附件不存在'}), 404
        
        # 删除数据库记录
        cur.execute(
            "DELETE FROM PrecomTaskAttachment WHERE AttachmentID = %s",
            (attachment_id,)
        )
        conn.commit()
        
        # 尝试删除文件（如果文件存在）
        file_path = attachment.get('FilePath')
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as e:
                current_app.logger.warning(f"无法删除附件文件: {file_path}, 错误: {e}")
        
        return jsonify({'success': True})
    except Exception as exc:
        conn.rollback()
        current_app.logger.error(f"删除附件失败: {exc}", exc_info=True)
        return jsonify({'error': '删除失败'}), 500
    finally:
        conn.close()


@precom_bp.route('/api/precom/tasks/<int:task_id>/punch_list', methods=['GET'])
def precom_punch_list(task_id: int):
    conn = create_connection()
    if not conn:
        return jsonify([])
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT ID, PunchNo, RefNo, SheetNo, RevNo, Description, Category, Cause,
                   IssuedBy, Rectified, RectifiedDate, Verified, VerifiedDate, Deleted, CreatedAt
            FROM PrecomTaskPunch
            WHERE TaskID = %s
            ORDER BY ID
            """,
            (task_id,),
        )
        rows = cur.fetchall()
        for row in rows:
            for key in ('RectifiedDate', 'VerifiedDate', 'CreatedAt'):
                if row.get(key):
                    row[key] = row[key].strftime('%Y-%m-%d %H:%M:%S')
        return jsonify(rows)
    finally:
        conn.close()


@precom_bp.route('/api/precom/tasks/<int:task_id>/punch_list', methods=['POST'])
def precom_add_punch(task_id: int):
    data = request.json or {}
    description = (data.get('Description') or '').strip()
    if not description:
        return jsonify({'error': 'Description 为必填项。'}), 400

    conn = create_connection()
    if not conn:
        return jsonify({'error': '数据库连接失败'}), 500
    try:
        cur = conn.cursor()
        rectified_flag = 'Y' if str(data.get('Rectified') or '').upper() == 'Y' else 'N'
        rectified_date = datetime.now() if rectified_flag == 'Y' else None
        verified_flag = 'Y' if str(data.get('Verified') or '').upper() == 'Y' else 'N'
        verified_date = datetime.now() if verified_flag == 'Y' else None

        cur.execute(
            """
            INSERT INTO PrecomTaskPunch (
                TaskID, PunchNo, RefNo, SheetNo, RevNo, Description, Category, Cause,
                IssuedBy, Rectified, RectifiedDate, Verified, VerifiedDate
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                task_id,
                data.get('PunchNo'),
                data.get('RefNo'),
                data.get('SheetNo'),
                data.get('RevNo'),
                description,
                data.get('Category'),
                data.get('Cause'),
                data.get('IssuedBy'),
                rectified_flag,
                rectified_date,
                verified_flag,
                verified_date,
            ),
        )
        conn.commit()
        return jsonify({'success': True, 'id': cur.lastrowid})
    except Exception as exc:
        conn.rollback()
        current_app.logger.error(f"新增预试车 Punch 失败: {exc}", exc_info=True)
        return jsonify({'error': '服务器内部错误，请稍后重试'}), 500
    finally:
        conn.close()


@precom_bp.route('/api/precom/tasks/<int:task_id>/punch_list/<int:punch_id>', methods=['DELETE'])
def precom_delete_punch(task_id: int, punch_id: int):
    conn = create_connection()
    if not conn:
        return jsonify({'error': '数据库连接失败'}), 500
    try:
        cur = conn.cursor()
        cur.execute(
            "DELETE FROM PrecomTaskPunch WHERE ID = %s AND TaskID = %s",
            (punch_id, task_id),
        )
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


@precom_bp.route('/precom/tasks/<int:task_id>/punch/import/template')
def precom_punch_template(task_id: int):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
    from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
    
    # 获取任务信息和已有punch数据
    system_code = ''
    subsystem_code = ''
    task_type = ''
    punch_data = []
    conn = create_connection()
    if conn:
        try:
            cur = conn.cursor(dictionary=True)
            cur.execute(
                """
                SELECT SystemCode, SubSystemCode, TaskType
                FROM PrecomTask
                WHERE TaskID = %s
                """,
                (task_id,)
            )
            row = cur.fetchone()
            if row:
                system_code = row.get('SystemCode') or ''
                subsystem_code = row.get('SubSystemCode') or ''
                task_type = row.get('TaskType') or ''
            
            # 查询已有的punch记录
            cur.execute(
                """
                SELECT PunchNo, RefNo, SheetNo, RevNo, Description,
                       Category, Cause, IssuedBy, CreatedAt as IssuedDate,
                       Rectified, RectifiedDate, Verified, VerifiedDate, Deleted
                FROM PrecomTaskPunch
                WHERE TaskID = %s
                ORDER BY ID
                """,
                (task_id,)
            )
            punch_data = cur.fetchall()
            # 格式化日期字段
            for punch in punch_data:
                if punch.get('IssuedDate'):
                    punch['IssuedDate'] = punch['IssuedDate'].strftime('%Y-%m-%d %H:%M:%S')
                if punch.get('RectifiedDate'):
                    punch['RectifiedDate'] = punch['RectifiedDate'].strftime('%Y-%m-%d %H:%M:%S')
                if punch.get('VerifiedDate'):
                    punch['VerifiedDate'] = punch['VerifiedDate'].strftime('%Y-%m-%d %H:%M:%S')
        finally:
            conn.close()

    info_df = pd.DataFrame([{
        'TaskID': task_id,
        'TaskType': task_type,
        'SystemCode': system_code,
        'SubSystemCode': subsystem_code,
        'GeneratedAt': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }])
    
    # 与试压包保持一致的列结构
    template_columns = [
        'PunchNo', 'RefNo', 'SheetNo', 'RevNo', 'Description',
        'Category', 'Cause', 'IssuedBy', 'IssuedDate',
        'Rectified', 'RectifiedDate', 'Verified', 'VerifiedDate', 'Deleted'
    ]
    # 如果有已有数据，使用已有数据；否则创建空模板
    if punch_data:
        template_df = pd.DataFrame(punch_data, columns=template_columns)
    else:
        template_df = pd.DataFrame(columns=template_columns)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        info_df.to_excel(writer, sheet_name='Info', index=False)
        template_df.to_excel(writer, sheet_name='PunchList', index=False)
        
        # 获取工作表对象以应用样式
        workbook = writer.book
        ws_punch = workbook['PunchList']
        
        # 定义样式
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 蓝色底纹
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # 设置列宽
        column_widths = {
            'A': 15,   # PunchNo（只读，用于更新）
            'B': 18,   # RefNo（ISO / Tag No.）
            'C': 10,   # SheetNo
            'D': 10,   # RevNo
            'E': 40,   # Description
            'F': 15,   # Category
            'G': 15,   # Cause
            'H': 12,   # IssuedBy
            'I': 18,   # IssuedDate
            'J': 10,   # Rectified
            'K': 15,   # RectifiedDate
            'L': 10,   # Verified
            'M': 15,   # VerifiedDate
            'N': 10    # Deleted
        }
        for col, width in column_widths.items():
            ws_punch.column_dimensions[col].width = width
        
        # 设置行高
        ws_punch.row_dimensions[1].height = 25  # 表头行高
        
        # 应用表头样式（第1行）
        for col_idx, col_name in enumerate(template_columns, start=1):
            cell = ws_punch.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 创建数据验证（下拉列表）
        # 找到相关列的索引
        punch_no_col = template_columns.index('PunchNo') + 1
        category_col = template_columns.index('Category') + 1
        cause_col = template_columns.index('Cause') + 1
        rectified_col = template_columns.index('Rectified') + 1
        verified_col = template_columns.index('Verified') + 1
        deleted_col = template_columns.index('Deleted') + 1
        
        # Category列：A/B/C/D
        category_validation = DataValidation(
            type="list",
            formula1='"A,B,C,D"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="输入错误",
            error="只能选择 A, B, C 或 D"
        )
        
        # Cause列：N/F/E
        cause_validation = DataValidation(
            type="list",
            formula1='"N,F,E"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="输入错误",
            error="只能选择 N(Non-Conformity), F(Field Revion) 或 E(Re-Engineering)"
        )
        
        # Rectified列：Y/N
        rectified_validation = DataValidation(
            type="list",
            formula1='"Y,N"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="输入错误",
            error="只能输入 Y 或 N"
        )
        
        # Verified列：Y/N
        verified_validation = DataValidation(
            type="list",
            formula1='"Y,N"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="输入错误",
            error="只能输入 Y 或 N"
        )
        
        # Deleted列：Y/N
        deleted_validation = DataValidation(
            type="list",
            formula1='"Y,N"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="输入错误",
            error="只能输入 Y 或 N（Y表示已删除，不计算在计数中）"
        )
        
        # 为PunchNo列添加注释说明
        from openpyxl.comments import Comment  # type: ignore
        punch_no_comment = Comment(
            "PunchNo由系统自动生成，新增记录时留空。\n如需更新或删除已有记录，请填写对应的PunchNo。",
            "系统提示"
        )
        
        for row in range(2, 101):
            for col in range(1, len(template_columns) + 1):
                cell = ws_punch.cell(row=row, column=col)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                # 为PunchNo列添加注释（仅第一行数据作为示例）
                if col == punch_no_col and row == 2:
                    cell.comment = punch_no_comment
                
                # 为Category列添加数据验证
                if col == category_col:
                    category_validation.add(cell)
                # 为Cause列添加数据验证
                if col == cause_col:
                    cause_validation.add(cell)
                # 为Rectified列添加数据验证
                if col == rectified_col:
                    rectified_validation.add(cell)
                # 为Verified列添加数据验证
                if col == verified_col:
                    verified_validation.add(cell)
                # 为Deleted列添加数据验证
                if col == deleted_col:
                    deleted_validation.add(cell)
        
        # 将数据验证添加到工作表
        ws_punch.add_data_validation(category_validation)
        ws_punch.add_data_validation(cause_validation)
        ws_punch.add_data_validation(rectified_validation)
        ws_punch.add_data_validation(verified_validation)
        ws_punch.add_data_validation(deleted_validation)
        
        # 冻结首行
        ws_punch.freeze_panes = 'A2'
        
    output.seek(0)
    filename = f"PrecomTask_{task_id}_PunchTemplate.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@precom_bp.route('/api/precom/tasks/<int:task_id>/punch_list/import', methods=['POST'])
def precom_import_punch(task_id: int):
    def parse_flag(value):
        """解析Y/N标志"""
        if value is None:
            return 'N'
        v = str(value).strip().upper()
        if v in {'Y', 'YES', '1'}:
            return 'Y'
        if v in {'N', 'NO', '0', ''}:
            return 'N'
        return None

    def parse_datetime(value):
        """解析日期时间"""
        if value in (None, ''):
            return None
        dt = pd.to_datetime(value, errors='coerce')
        if pd.isna(dt):
            return None
        return dt.to_pydatetime()

    upload = request.files.get('file')
    if not upload or upload.filename == '':
        return jsonify({'success': False, 'message': '请上传需要导入的 Excel 文件'}), 400

    file_bytes = upload.read()
    try:
        excel = pd.ExcelFile(BytesIO(file_bytes))
    except Exception as exc:
        return jsonify({'success': False, 'message': f'读取文件失败: {exc}'}), 400

    sheet_name = next((name for name in excel.sheet_names if 'punch' in name.lower()), excel.sheet_names[0])
    df = excel.parse(sheet_name)
    if df.empty:
        return jsonify({'success': False, 'message': 'Excel 内容为空'}), 400

    required_columns = ['Description']
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        return jsonify({'success': False, 'message': f"缺少必要字段: {', '.join(missing)}"}), 400

    records = []
    errors = []

    for idx, row in df.iterrows():
        line_no = idx + 2

        def cell(name):
            value = row.get(name)
            if pd.isna(value):
                return ''
            return str(value).strip()

        punch_no = cell('PunchNo')
        ref_no = cell('RefNo')
        description = cell('Description')
        
        # 如果有PunchNo，说明是更新已有记录；如果没有，说明是新增记录
        if punch_no:
            # 更新记录：只需要PunchNo即可，其他字段可选
            record = {
                'PunchNo': punch_no,
                'RefNo': ref_no or None,
                'SheetNo': cell('SheetNo') or None,
                'RevNo': cell('RevNo') or None,
                'Description': description or None,
                'Category': cell('Category') or None,
                'Cause': cell('Cause') or None,
                'IssuedBy': cell('IssuedBy') or None,
                'Rectified': parse_flag(cell('Rectified')) or None,
                'Verified': parse_flag(cell('Verified')) or None,
                'RectifiedDate': parse_datetime(cell('RectifiedDate')),
                'VerifiedDate': parse_datetime(cell('VerifiedDate')),
                'Deleted': parse_flag(cell('Deleted')) or 'N',
                'is_update': True
            }
        else:
            # 新增记录：必须要有Description
            if not description:
                errors.append(f"第 {line_no} 行：新增记录时，请填写 Description。")
                continue
            record = {
                'PunchNo': None,  # 由系统自动生成
                'RefNo': ref_no or None,
                'SheetNo': cell('SheetNo') or None,
                'RevNo': cell('RevNo') or None,
                'Description': description,
                'Category': cell('Category') or None,
                'Cause': cell('Cause') or None,
                'IssuedBy': cell('IssuedBy') or None,
                'Rectified': parse_flag(cell('Rectified')) or 'N',
                'Verified': parse_flag(cell('Verified')) or 'N',
                'RectifiedDate': parse_datetime(cell('RectifiedDate')),
                'VerifiedDate': parse_datetime(cell('VerifiedDate')),
                'Deleted': parse_flag(cell('Deleted')) or 'N',
                'is_update': False
            }
        records.append(record)

    if errors:
        return jsonify({'success': False, 'message': '数据校验失败', 'errors': errors}), 400
    if not records:
        return jsonify({'success': False, 'message': '没有可以导入的数据'}), 400

    conn = create_connection()
    if not conn:
        return jsonify({'success': False, 'message': '数据库连接失败'}), 500

    inserted = 0
    updated = 0
    try:
        cur = conn.cursor(dictionary=True)
        
        # 分离新增和更新的记录
        update_records = [r for r in records if r.get('is_update')]
        insert_records = [r for r in records if not r.get('is_update')]
        
        # 处理更新记录：通过PunchNo匹配
        if update_records:
            punch_numbers = [r['PunchNo'] for r in update_records]
            placeholders = ','.join(['%s'] * len(punch_numbers))
            query = f"""
                SELECT ID, PunchNo, RefNo, SheetNo, RevNo, Description,
                       Category, Cause, IssuedBy, Rectified, RectifiedDate, 
                       Verified, VerifiedDate, Deleted
                FROM PrecomTaskPunch
                WHERE TaskID = %s AND PunchNo IN ({placeholders})
            """
            cur.execute(query, tuple([task_id] + punch_numbers))
            existing_map = {row['PunchNo']: row for row in cur.fetchall()}
            
            for record in update_records:
                punch_no = record['PunchNo']
                if punch_no not in existing_map:
                    errors.append(f"PunchNo {punch_no} 不存在，无法更新。")
                    continue
                
                existing = existing_map[punch_no]
                # 只更新提供的字段，未提供的字段保持原值
                update_fields = []
                update_values = []
                
                if record.get('RefNo') is not None:
                    update_fields.append('RefNo=%s')
                    update_values.append(record['RefNo'])
                if record.get('SheetNo') is not None:
                    update_fields.append('SheetNo=%s')
                    update_values.append(record['SheetNo'])
                if record.get('RevNo') is not None:
                    update_fields.append('RevNo=%s')
                    update_values.append(record['RevNo'])
                if record.get('Description'):
                    update_fields.append('Description=%s')
                    update_values.append(record['Description'])
                if record.get('Category') is not None:
                    update_fields.append('Category=%s')
                    update_values.append(record['Category'])
                if record.get('Cause') is not None:
                    update_fields.append('Cause=%s')
                    update_values.append(record['Cause'])
                if record.get('IssuedBy') is not None:
                    update_fields.append('IssuedBy=%s')
                    update_values.append(record['IssuedBy'])
                if record.get('Rectified') is not None:
                    update_fields.append('Rectified=%s')
                    update_values.append(record['Rectified'])
                if record.get('Verified') is not None:
                    update_fields.append('Verified=%s')
                    update_values.append(record['Verified'])
                if record.get('RectifiedDate') is not None:
                    update_fields.append('RectifiedDate=%s')
                    update_values.append(record['RectifiedDate'])
                if record.get('VerifiedDate') is not None:
                    update_fields.append('VerifiedDate=%s')
                    update_values.append(record['VerifiedDate'])
                if record.get('Deleted') is not None:
                    update_fields.append('Deleted=%s')
                    update_values.append(record['Deleted'])
                
                if update_fields:
                    update_sql = f"UPDATE PrecomTaskPunch SET {', '.join(update_fields)} WHERE ID = %s"
                    update_values.append(existing['ID'])
                    cur.execute(update_sql, tuple(update_values))
                    updated += cur.rowcount
        
        # 处理新增记录
        if insert_records:
            insert_values = []
            for record in insert_records:
                insert_values.append((
                    task_id,
                    record.get('RefNo'),
                    record.get('SheetNo'),
                    record.get('RevNo'),
                    record['Description'],
                    record.get('Category'),
                    record.get('Cause'),
                    record.get('IssuedBy'),
                    record.get('Rectified', 'N'),
                    record.get('RectifiedDate'),
                    record.get('Verified', 'N'),
                    record.get('VerifiedDate'),
                    record.get('Deleted', 'N')
                ))
            
            cur.executemany(
                """
                INSERT INTO PrecomTaskPunch (
                    TaskID, RefNo, SheetNo, RevNo, Description, Category, Cause,
                    IssuedBy, Rectified, RectifiedDate, Verified, VerifiedDate, Deleted
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                insert_values
            )
            inserted = cur.rowcount
        
        conn.commit()
        
        message = []
        if inserted > 0:
            message.append(f"新增 {inserted} 条")
        if updated > 0:
            message.append(f"更新 {updated} 条")
        
        return jsonify({
            'success': True,
            'message': '、'.join(message) if message else '没有变化',
            'inserted': inserted,
            'updated': updated
        })
    except Exception as exc:
        conn.rollback()
        return jsonify({'success': False, 'message': f'导入失败: {exc}'}), 500
    finally:
        conn.close()


def _parse_activity_rows_from_request():
    """从表单中解析施工活动行数据，返回字典列表"""
    act_ids = request.form.getlist('ActivityActID')
    blocks = request.form.getlist('ActivityBlock')
    descriptions = request.form.getlist('ActivityDescription')
    scopes = request.form.getlist('ActivityScope')
    disciplines = request.form.getlist('ActivityDiscipline')
    work_packages = request.form.getlist('ActivityWorkPackage')
    total_quantities = request.form.getlist('ActivityTotalQuantity')
    completed_quantities = request.form.getlist('ActivityCompletedQuantity')
    completed_percents = request.form.getlist('ActivityCompletedPercent')
    weight_factors = request.form.getlist('ActivityWeightFactor')
    man_hours_list = request.form.getlist('ActivityManHours')
    subprojects = request.form.getlist('ActivitySubproject')

    activities = []
    row_count = max(
        len(act_ids),
        len(blocks),
        len(descriptions),
        len(scopes),
        len(disciplines),
        len(work_packages),
        len(total_quantities),
        len(completed_quantities),
        len(completed_percents),
        len(weight_factors),
        len(man_hours_list),
        len(subprojects),
    )
    for idx in range(row_count):
        act_id = (act_ids[idx].strip() if idx < len(act_ids) and act_ids[idx] else '')
        block = (blocks[idx].strip() if idx < len(blocks) and blocks[idx] else '')
        desc = (descriptions[idx].strip() if idx < len(descriptions) and descriptions[idx] else '')
        scope = (scopes[idx].strip() if idx < len(scopes) and scopes[idx] else '')
        disc = (disciplines[idx].strip() if idx < len(disciplines) and disciplines[idx] else '')
        wp = (work_packages[idx].strip() if idx < len(work_packages) and work_packages[idx] else '')
        total_raw = total_quantities[idx].strip() if idx < len(total_quantities) and total_quantities[idx] else ''
        completed_raw = completed_quantities[idx].strip() if idx < len(completed_quantities) and completed_quantities[idx] else ''
        percent_raw = completed_percents[idx].strip() if idx < len(completed_percents) and completed_percents[idx] else ''
        wf_raw = weight_factors[idx].strip() if idx < len(weight_factors) and weight_factors[idx] else ''
        mh_raw = man_hours_list[idx].strip() if idx < len(man_hours_list) and man_hours_list[idx] else ''
        subprj = (subprojects[idx].strip() if idx < len(subprojects) and subprojects[idx] else '')

        # 如果整行都是空，则跳过
        if not any([act_id, block, desc, scope, disc, wp, total_raw, completed_raw, percent_raw, wf_raw, mh_raw, subprj]):
            continue

        try:
            total_qty = float(total_raw) if total_raw else None
        except ValueError:
            total_qty = None
        try:
            completed_qty = float(completed_raw) if completed_raw else None
        except ValueError:
            completed_qty = None
        try:
            completed_percent = float(percent_raw) if percent_raw else None
        except ValueError:
            completed_percent = None
        try:
            wf = float(wf_raw) if wf_raw else None
        except ValueError:
            wf = None
        try:
            mh = float(mh_raw) if mh_raw else None
        except ValueError:
            mh = None

        activities.append(
            {
                'ActID': act_id or None,
                'Block': block or None,
                'ActDescription': desc or None,
                'Scope': scope or None,
                'Discipline': disc or None,
                'WorkPackage': wp or None,
                'TotalQuantity': total_qty,
                'CompletedQuantity': completed_qty,
                'CompletedPercent': completed_percent,
                'WeightFactor': wf,
                'ManHours': mh,
                'Subproject': subprj or None,
            }
        )
    return activities


def _save_task_activities(conn, task_id: int, activities):
    """将施工活动明细保存到 PrecomTaskActivity 表，先清空后重新插入"""
    cur = conn.cursor()
    # 先删除原有记录
    cur.execute("DELETE FROM PrecomTaskActivity WHERE TaskID = %s", (task_id,))
    if not activities:
        return
    insert_sql = """
        INSERT INTO PrecomTaskActivity (
            TaskID, ActID, Block, ActDescription, Scope,
            Discipline, WorkPackage, TotalQuantity, CompletedQuantity, CompletedPercent,
            WeightFactor, ManHours, Subproject
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    values = [
        (
            task_id,
            act['ActID'],
            act['Block'],
            act['ActDescription'],
            act['Scope'],
            act['Discipline'],
            act['WorkPackage'],
            act['TotalQuantity'],
            act['CompletedQuantity'],
            act['CompletedPercent'],
            act['WeightFactor'],
            act['ManHours'],
            act['Subproject'],
        )
        for act in activities
    ]
    cur.executemany(insert_sql, values)


def _sanitize_list(values):
    return [v.strip() for v in values if v and v.strip()]


def _join_position_blocks_from_request():
    blocks = _sanitize_list(request.form.getlist('PositionBlocks'))
    return ';'.join(blocks) if blocks else None


def _join_drawing_numbers_from_request():
    numbers = _sanitize_list(request.form.getlist('DrawingNumbers'))
    if numbers:
        return '\n'.join(numbers)
    fallback = request.form.get('DrawingNumber')
    if fallback and fallback.strip():
        return fallback.strip()
    return None


def _split_position_blocks(value):
    if not value:
        return ['']
    parts = [p.strip() for p in re.split(r'[;\n,]', value) if p.strip()]
    return parts or ['']


def _split_drawing_numbers(value):
    if not value:
        return ['']
    parts = [p.strip() for p in re.split(r'[\n;]+', value) if p.strip()]
    return parts or ['']


@precom_bp.route('/precom/tasks/export', methods=['GET'])
def precom_export_tasks():
    task_type = (request.args.get('task_type') or '').strip() or None
    conn = create_connection()
    if not conn:
        return "数据库连接失败", 500
    try:
        cur = conn.cursor(dictionary=True)
        clauses = []
        params = []
        if task_type:
            clauses.append("TaskType = %s")
            params.append(task_type)
        where_sql = " AND ".join(clauses) if clauses else "1=1"
        cur.execute(
            f"""
            SELECT TaskID, TaskType, SystemCode, SubSystemCode, DrawingNumber,
                   TagNumber, PointTag, Description, PositionBlock,
                   QuantityTotal, QuantityDone, PlannedDate, ActualDate,
                   PerformedBy, TestType
            FROM PrecomTask
            WHERE {where_sql}
            ORDER BY SubSystemCode, TaskType, TagNumber, PointTag
            """,
            tuple(params),
        )
        rows = cur.fetchall()
    finally:
        conn.close()

    if not rows:
        rows = []
    df = pd.DataFrame(rows)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='PrecomTasks', index=False)
    output.seek(0)
    type_part = task_type or 'ALL'
    filename = f"PrecomTasks_{type_part}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@precom_bp.route('/precom/tasks/import', methods=['POST'])
def precom_import_tasks():
    task_type = (request.args.get('task_type') or '').strip() or None
    upload = request.files.get('file')
    if not upload or upload.filename == '':
        return jsonify({'success': False, 'message': '请上传需要导入的 Excel 文件'}), 400

    file_bytes = upload.read()
    try:
        excel = pd.ExcelFile(BytesIO(file_bytes))
    except Exception as exc:
        return jsonify({'success': False, 'message': f'读取文件失败: {exc}'}), 400

    sheet_name = excel.sheet_names[0]
    df = excel.parse(sheet_name)
    if df.empty:
        return jsonify({'success': False, 'message': 'Excel 内容为空'}), 400

    required_columns = ['SubSystemCode', 'Description']
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        return jsonify({'success': False, 'message': f"缺少必要字段: {', '.join(missing)}"}), 400

    records = []
    for _, row in df.iterrows():
        subsystem = str(row.get('SubSystemCode') or '').strip()
        description = str(row.get('Description') or '').strip()
        if not subsystem or not description:
            continue
        records.append(
            (
                task_type or str(row.get('TaskType') or '').strip() or 'Manhole',
                str(row.get('SystemCode') or '').strip() or None,
                subsystem,
                str(row.get('DrawingNumber') or '').strip() or None,
                str(row.get('TagNumber') or '').strip() or None,
                str(row.get('PointTag') or '').strip() or None,
                description,
                str(row.get('PositionBlock') or '').strip() or None,
                int(row.get('QuantityTotal') or 1),
                int(row.get('QuantityDone') or 0),
                row.get('PlannedDate') if pd.notna(row.get('PlannedDate')) else None,
                row.get('ActualDate') if pd.notna(row.get('ActualDate')) else None,
                str(row.get('PerformedBy') or '').strip() or None,
                str(row.get('TestType') or '').strip() or None,
            )
        )

    if not records:
        return jsonify({'success': False, 'message': '没有可以导入的记录'}), 400

    conn = create_connection()
    if not conn:
        return jsonify({'success': False, 'message': '数据库连接失败'}), 500

    try:
        cur = conn.cursor()
        cur.executemany(
            """
            INSERT INTO PrecomTask (
                TaskType, SystemCode, SubSystemCode, DrawingNumber,
                TagNumber, PointTag, Description, PositionBlock,
                QuantityTotal, QuantityDone, PlannedDate, ActualDate,
                PerformedBy, TestType
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            records,
        )
        conn.commit()
        return jsonify({'success': True, 'inserted': cur.rowcount})
    except Exception as exc:
        conn.rollback()
        return jsonify({'success': False, 'message': f'导入失败: {exc}'}), 500
    finally:
        conn.close()


@precom_bp.route('/precom/tasks/<int:task_id>/export', methods=['GET'])
def precom_export_single_task(task_id: int):
    conn = create_connection()
    if not conn:
        return "数据库连接失败", 500
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM PrecomTask WHERE TaskID = %s", (task_id,))
        task = cur.fetchone()
        if not task:
            return "未找到任务", 404

        cur.execute(
            """
            SELECT FileName, FilePath
            FROM PrecomTaskAttachment
            WHERE TaskID = %s
            """,
            (task_id,),
        )
        attachments = cur.fetchall()
    finally:
        conn.close()

    # 构建任务 Excel
    df = pd.DataFrame([task])
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Task', index=False)
    excel_buffer.seek(0)

    # 打包为 ZIP（Excel + 附件）
    import zipfile

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"PrecomTask_{task_id}.xlsx", excel_buffer.read())
        for att in attachments or []:
            path = att.get('FilePath')
            name = att.get('FileName') or os.path.basename(path or '')
            if path and os.path.exists(path):
                zf.write(path, arcname=f"attachments/{name}")
    zip_buffer.seek(0)

    filename = f"PrecomTask_{task_id}_with_attachments.zip"
    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/zip',
    )


