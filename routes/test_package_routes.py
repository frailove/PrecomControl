from flask import Blueprint, request, redirect, jsonify, send_file, render_template, current_app  # type: ignore
from models.test_package import TestPackageModel
from models.system import SystemModel
from models.subsystem import SubsystemModel
from database import create_connection, ensure_hydro_columns
from utils.exporters import export_test_packages_to_excel
from utils.test_package_exporter import export_test_package_from_template
from utils.test_package_status import TestPackageStatus
from utils.pipeline_alerts import get_pipeline_alerts, update_pipeline_alert
from utils.refresh_aggregated_data import PIPELINE_SYSTEM_SHARE_THRESHOLD, refresh_nde_pwht_status
from urllib.parse import urlencode, unquote
from werkzeug.utils import secure_filename  # type: ignore
from math import ceil
from datetime import datetime
from io import BytesIO
import pandas as pd  # type: ignore
import os
import re
import html


test_package_bp = Blueprint('test_package', __name__)
PER_PAGE = 50

ATTACHMENT_MODULES = [
    {"code": "3.0", "module": "PID_Drawings", "title": "P&ID Drawings", "description": "Piping & Instrument Diagram"},
    {"code": "4.0", "module": "ISO_Drawings", "title": "Hydro Test ISO", "description": "Hydrostatic test isometrics"},
    {"code": "5.0", "module": "Symbols_Legend", "title": "Symbols Legend", "description": "Legend and abbreviations"},
    {"code": "8.0", "module": "Test_Flow_Chart", "title": "Test Flow Chart", "description": "Test procedure"},
    {"code": "9.0", "module": "Test_Check_List", "title": "Test Check List", "description": "Checklist"},
    {"code": "10.0", "module": "Calibration_Certificates", "title": "Calibration Certificates", "description": "Instrument calibration"},
    {"code": "11.0", "module": "Test_Certificate", "title": "Test Certificate", "description": "Pipeline certificate"},
    {"code": "12.0", "module": "Flushing_Certificate", "title": "Flushing Certificate", "description": "Flushing / cleaning"},
    {"code": "13.0", "module": "Reinstatement_Check_List", "title": "Reinstatement Check List", "description": "Reinstatement"},
    {"code": "14.0", "module": "Others", "title": "Other Attachments", "description": "Additional documents"}
]

REQUIRED_ATTACHMENT_MODULES = [
    "PID_Drawings",
    "ISO_Drawings",
    "Symbols_Legend",
    "Test_Check_List",
    "Calibration_Certificates",
    "Test_Certificate"
]

UPLOAD_FOLDER = 'uploads/test_packages'
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'dwg', 'dxf', 'xlsx', 'xls', 'doc', 'docx', 'zip'}

from importlib import import_module

try:
    # 通过动态导入避免静态检查器对可选依赖报错
    magic = import_module("magic")  # type: ignore[assignment]
except ImportError:  # 安全但不强依赖，缺失时仅根据扩展名校验
    magic = None

PUNCH_HEADER_MAP = {
    'punchno': 'PunchNo',
    'itemno': 'PunchNo',
    'punch no': 'PunchNo',
    'iso': 'ISODrawingNo',
    'iso/tag': 'ISODrawingNo',
    'sheetno': 'SheetNo',
    'revno': 'RevNo',
    'description': 'Description',
    'description of punch': 'Description',
    'category': 'Category',
    'cause': 'Cause',
    'issuedby': 'IssuedBy',
    'rectified': 'Rectified',
    'rectifieddate': 'RectifiedDate',
    'verified': 'Verified',
    'verifieddate': 'VerifiedDate',
    'deleted': 'Deleted',
    'remarks': 'Remarks',
    'testpackageid': 'TestPackageID'
}


def ensure_punch_list_schema():
    conn = create_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS PunchListImportLog (
                ImportID INT AUTO_INCREMENT PRIMARY KEY,
                TestPackageID VARCHAR(50) NOT NULL,
                FileName VARCHAR(255),
                TotalCount INT DEFAULT 0,
                InsertedCount INT DEFAULT 0,
                UpdatedCount INT DEFAULT 0,
                ErrorCount INT DEFAULT 0,
                Message TEXT,
                CreatedAt DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_punch_import_tp (TestPackageID)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )
        cur.execute(
            """
            SELECT COUNT(*)
            FROM information_schema.columns
            WHERE table_schema = DATABASE()
              AND table_name = 'PunchList'
              AND column_name = 'PunchNo'
            """
        )
        if cur.fetchone()[0] == 0:
            cur.execute("ALTER TABLE PunchList ADD COLUMN PunchNo VARCHAR(100) NULL AFTER ID")
        cur.execute("SHOW INDEX FROM PunchList WHERE Key_name = 'idx_punch_no'")
        if cur.fetchone() is None:
            cur.execute("ALTER TABLE PunchList ADD INDEX idx_punch_no (PunchNo)")
        
        # 添加Deleted字段，用于标记已删除的记录（不计算在计数中）
        cur.execute(
            """
            SELECT COUNT(*)
            FROM information_schema.columns
            WHERE table_schema = DATABASE()
              AND table_name = 'PunchList'
              AND column_name = 'Deleted'
            """
        )
        if cur.fetchone()[0] == 0:
            cur.execute("ALTER TABLE PunchList ADD COLUMN Deleted CHAR(1) DEFAULT 'N' AFTER VerifiedDate")
            cur.execute("ALTER TABLE PunchList ADD INDEX idx_deleted (Deleted)")
    except Exception as exc:
        print(f"ensure_punch_list_schema failed: {exc}")
    finally:
        conn.close()


ensure_punch_list_schema()


def normalize_header(value):
    if value is None:
        return ''
    if not isinstance(value, str):
        value = str(value)
    return value.splitlines()[0].strip().lower()


def parse_flag(value):
    if value is None:
        return 'N'
    v = str(value).strip().upper()
    if v in {'Y', 'YES', '1'}:
        return 'Y'
    if v in {'N', 'NO', '0', ''}:
        return 'N'
    return None


def parse_datetime(value):
    if value in (None, ''):
        return None
    dt = pd.to_datetime(value, errors='coerce')
    if pd.isna(dt):
        return None
    return dt.to_pydatetime()


def sanitize_date_columns(df, columns):
    for col in columns:
        if col in df.columns:
            series = pd.to_datetime(df[col], errors='coerce')
            df[col] = series.dt.strftime('%Y-%m-%d')
            df[col] = df[col].where(series.notna(), None)


def extract_drawing_pattern(drawing_number):
    if not drawing_number:
        return None
    parts = re.findall(r'\d+', drawing_number)
    if len(parts) >= 3:
        return '-'.join(parts[:3])
    if len(parts) == 2:
        return '-'.join(parts)
    if len(parts) == 1:
        return parts[0]
    return None


def normalize_block_for_matching(block):
    if not block:
        return None
    parts = [p.strip() for p in str(block).split('-') if p.strip()]
    if len(parts) == 3:
        return '-'.join([parts[1], parts[2], parts[0]])
    if len(parts) == 2:
        return '-'.join([parts[1], parts[0]])
    if len(parts) == 1:
        return parts[0]
    if len(parts) > 3:
        return '-'.join(parts[1:] + [parts[0]])
    return None


def fetch_drawings_by_block_patterns(cursor, block_patterns, chunk_size=25):
    patterns = [p for p in block_patterns if p]
    if not patterns:
        return set()
    matched = set()
    for i in range(0, len(patterns), chunk_size):
        chunk = patterns[i:i + chunk_size]
        like_clause = " OR ".join(["DrawingNumber LIKE %s"] * len(chunk))
        params = tuple(f"%{pattern}%" for pattern in chunk)
        cursor.execute(
            f"""
            SELECT DISTINCT DrawingNumber
            FROM WeldingList
            WHERE DrawingNumber IS NOT NULL
              AND DrawingNumber <> ''
              AND ({like_clause})
            """,
            params
        )
        for row in cursor.fetchall():
            drawing = row.get('DrawingNumber')
            if drawing:
                matched.add(drawing)
    return matched


def get_faclist_filter_options(filter_subproject=None, filter_train=None, filter_unit=None,
                               filter_simpleblk=None, filter_mainblock=None, filter_block=None,
                               filter_bccquarter=None):
    conn = create_connection()
    if not conn:
        return {}

    clauses = []
    params = []
    if filter_subproject:
        clauses.append("SubProjectCode = %s")
        params.append(filter_subproject)
    if filter_train:
        clauses.append("Train = %s")
        params.append(filter_train)
    if filter_unit:
        clauses.append("Unit = %s")
        params.append(filter_unit)
    if filter_simpleblk:
        clauses.append("SimpleBLK = %s")
        params.append(filter_simpleblk)
    if filter_mainblock:
        clauses.append("MainBlock = %s")
        params.append(filter_mainblock)
    if filter_block:
        clauses.append("Block = %s")
        params.append(filter_block)
    if filter_bccquarter:
        clauses.append("BCCQuarter = %s")
        params.append(filter_bccquarter)

    where_sql = " AND ".join(clauses) if clauses else "1=1"

    options = {
        'subproject_codes': [],
        'trains': [],
        'units': [],
        'simpleblks': [],
        'mainblocks': {},
        'blocks': {},
        'bccquarters': []
    }

    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            f"""
            SELECT DISTINCT SubProjectCode, Train, Unit, SimpleBLK, MainBlock, Block, BCCQuarter
            FROM Faclist
            WHERE ({where_sql})
              AND (SubProjectCode IS NOT NULL OR Train IS NOT NULL OR Unit IS NOT NULL
                   OR SimpleBLK IS NOT NULL OR MainBlock IS NOT NULL OR Block IS NOT NULL OR BCCQuarter IS NOT NULL)
            ORDER BY SubProjectCode, Train, Unit, SimpleBLK, MainBlock, Block, BCCQuarter
            """,
            tuple(params)
        )
        for row in cur.fetchall():
            if row.get('SubProjectCode') and row['SubProjectCode'] not in options['subproject_codes']:
                options['subproject_codes'].append(str(row['SubProjectCode']))
            if row.get('Train') and row['Train'] not in options['trains']:
                options['trains'].append(str(row['Train']))
            if row.get('Unit') and row['Unit'] not in options['units']:
                options['units'].append(str(row['Unit']))
            if row.get('SimpleBLK') and row['SimpleBLK'] not in options['simpleblks']:
                options['simpleblks'].append(str(row['SimpleBLK']))
            if row.get('BCCQuarter') and row['BCCQuarter'] not in options['bccquarters']:
                options['bccquarters'].append(str(row['BCCQuarter']))

            if row.get('SimpleBLK'):
                simpleblk = str(row['SimpleBLK'])
                options['mainblocks'].setdefault(simpleblk, [])
                if row.get('MainBlock') and str(row['MainBlock']) not in options['mainblocks'][simpleblk]:
                    options['mainblocks'][simpleblk].append(str(row['MainBlock']))

            if row.get('MainBlock'):
                mainblock = str(row['MainBlock'])
                options['blocks'].setdefault(mainblock, [])
                if row.get('Block') and str(row['Block']) not in options['blocks'][mainblock]:
                    options['blocks'][mainblock].append(str(row['Block']))

        options['subproject_codes'].sort()
        options['trains'].sort()
        options['units'].sort()
        options['simpleblks'].sort()
        options['bccquarters'].sort()
        for mainblock in options['blocks']:
            options['blocks'][mainblock].sort(
                key=lambda x: tuple(int(p) if p.isdigit() else 999999 for p in x.split('-')) if '-' in x else (int(x) if x.isdigit() else 999999)
            )
    finally:
        conn.close()

    return options


def get_bootstrap_css():
    return '<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">'


def get_navbar():
    return '''
    <nav class="navbar navbar-expand-lg navbar-dark bg-dark">
        <div class="container-fluid">
            <a class="navbar-brand" href="/">预试车管理系统</a>
            <div class="navbar-nav">
                <a class="nav-link" href="/">首页</a>
                <a class="nav-link" href="/systems">系统管理</a>
                <a class="nav-link" href="/subsystems">子系统管理</a>
                <a class="nav-link" href="/test_packages">试压包管理</a>
            </div>
        </div>
    </nav>
    '''


def build_pagination_base_path(args, path='/test_packages'):
    params = args.to_dict(flat=False)
    params.pop('page', None)
    query_pairs = []
    for key, values in params.items():
        for value in values:
            if value not in (None, ''):
                query_pairs.append((key, value))
    encoded = urlencode(query_pairs)
    if encoded:
        return f"{path}?{encoded}&page="
    return f"{path}?page="


def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _validate_upload_file(file) -> tuple[bool, str | None]:
    """验证上传文件的扩展名、MIME 类型和单个文件大小，返回 (是否通过, 错误消息)。"""
    if not file or not file.filename:
        return False, "未选择文件"

    if not allowed_file(file.filename):
        return False, "不允许的文件类型"

    # 单个文件大小限制：50MB
    max_size_bytes = 50 * 1024 * 1024
    try:
        pos = file.stream.tell()
        file.stream.seek(0, os.SEEK_END)
        size = file.stream.tell()
        file.stream.seek(pos)
    except Exception:
        size = None
    if size is not None and size > max_size_bytes:
        return False, "单个文件大小不能超过 50MB"

    # MIME 类型校验（在安装了 python-magic 的环境下启用）
    if magic is not None:
        try:
            pos = file.stream.tell()
            sample = file.stream.read(2048)
            file.stream.seek(pos)
            mime_type = magic.from_buffer(sample, mime=True)
        except Exception:
            mime_type = None

        if mime_type:
            allowed_mimes = {
                'application/pdf',
                'image/png',
                'image/jpeg',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'application/vnd.ms-excel',
                'application/msword',
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'application/zip',
                'image/vnd.dwg',
            }
            if mime_type not in allowed_mimes:
                return False, "文件内容与扩展名不匹配或类型不受支持"

    return True, None


def ensure_upload_folder(test_package_id: str, module_name: str) -> str:
    folder = os.path.join(UPLOAD_FOLDER, secure_filename(test_package_id), module_name)
    os.makedirs(folder, exist_ok=True)
    return folder


def match_faclist_drawings(filters) -> set | None:
    if not any(filters.values()):
        return None
    conn = create_connection()
    if not conn:
        return set()
    try:
        cur = conn.cursor(dictionary=True)
        clauses = []
        params = []
        mapping = {
            'subproject_code': 'SubProjectCode',
            'train': 'Train',
            'unit': 'Unit',
            'simpleblk': 'SimpleBLK',
            'mainblock': 'MainBlock',
            'block': 'Block',
            'bccquarter': 'BCCQuarter'
        }
        for key, column in mapping.items():
            value = filters.get(key)
            if value:
                clauses.append(f"{column} = %s")
                params.append(value)
        if not clauses:
            return None
        cur.execute(
            f"SELECT DISTINCT Block FROM Faclist WHERE {' AND '.join(clauses)} AND Block IS NOT NULL AND Block <> ''",
            tuple(params)
        )
        blocks = [row['Block'] for row in cur.fetchall() if row.get('Block')]
        # Block 格式已与 Faclist 一致，直接使用
        patterns = {b.strip() for b in blocks if b and b.strip()}
        return fetch_drawings_by_block_patterns(cur, list(patterns))
    finally:
        conn.close()


def load_package_extras(package_ids):
    extras = {
        'pid_counts': {},
        'iso_counts': {},
        'attachments': {},
        'punch_open': {},
        'nde_status': {}
    }
    if not package_ids:
        return extras
    conn = create_connection()
    if not conn:
        return extras
    try:
        cur = conn.cursor(dictionary=True)
        placeholders = ','.join(['%s'] * len(package_ids))
        params = tuple(package_ids)

        cur.execute(
            f"SELECT TestPackageID, COUNT(*) AS cnt FROM PIDList WHERE TestPackageID IN ({placeholders}) GROUP BY TestPackageID",
            params
        )
        for row in cur.fetchall():
            extras['pid_counts'][row['TestPackageID']] = row['cnt']

        cur.execute(
            f"SELECT TestPackageID, COUNT(*) AS cnt FROM ISODrawingList WHERE TestPackageID IN ({placeholders}) GROUP BY TestPackageID",
            params
        )
        for row in cur.fetchall():
            extras['iso_counts'][row['TestPackageID']] = row['cnt']

        cur.execute(
            f"""
            SELECT TestPackageID, ModuleName, COUNT(*) AS cnt
            FROM TestPackageAttachments
            WHERE TestPackageID IN ({placeholders})
            GROUP BY TestPackageID, ModuleName
            """,
            params
        )
        for row in cur.fetchall():
            extras['attachments'].setdefault(row['TestPackageID'], {})[row['ModuleName']] = row['cnt']

        cur.execute(
            f"""
            SELECT TestPackageID, COUNT(*) AS cnt
            FROM PunchList
            WHERE TestPackageID IN ({placeholders}) AND Rectified = 'N'
            GROUP BY TestPackageID
            """,
            params
        )
        for row in cur.fetchall():
            extras['punch_open'][row['TestPackageID']] = row['cnt']

        cur.execute(
            f"""
            SELECT TestPackageID,
                   VT_Total, VT_Completed,
                   RT_Total, RT_Completed,
                   PT_Total, PT_Completed,
                   UT_Total, UT_Completed,
                   MT_Total, MT_Completed,
                   PMI_Total, PMI_Completed,
                   FT_Total, FT_Completed,
                   HT_Total, HT_Completed,
                   PWHT_Total, PWHT_Completed
            FROM NDEPWHTStatus
            WHERE TestPackageID IN ({placeholders})
            """,
            params
        )
        for row in cur.fetchall():
            extras['nde_status'][row['TestPackageID']] = row
    finally:
        conn.close()
    return extras


def compute_package_status(package, extras):
    tp_id = package['TestPackageID']
    total_din = float(package.get('total_din') or 0)
    completed_din = float(package.get('completed_din') or 0)
    joints_total = int(package.get('total_welds') or 0)
    joints_completed = int(package.get('completed_welds') or 0)
    joints_complete = joints_total > 0 and joints_completed >= joints_total

    nde_row = extras['nde_status'].get(tp_id)
    tests_complete = True
    if nde_row:
        for key in ('VT', 'RT', 'PT', 'UT', 'MT', 'PMI', 'FT', 'HT', 'PWHT'):
            total = nde_row.get(f'{key}_Total') or 0
            completed = nde_row.get(f'{key}_Completed') or 0
            if total and completed < total:
                tests_complete = False
                break
    else:
        tests_complete = False

    pid_ready = extras['pid_counts'].get(tp_id, 0) > 0
    iso_ready = extras['iso_counts'].get(tp_id, 0) > 0
    attachments_ready = all(extras['attachments'].get(tp_id, {}).get(module, 0) > 0 for module in REQUIRED_ATTACHMENT_MODULES)
    punch_ready = extras['punch_open'].get(tp_id, 0) == 0

    basic_info = all([
        package.get('SystemCode'),
        package.get('SubSystemCode'),
        package.get('Description'),
        package.get('TestType'),
        package.get('DesignPressure'),
        package.get('TestPressure')
    ])

    document_ready = all([
        basic_info,
        pid_ready or iso_ready,
        attachments_ready,
        punch_ready,
        joints_complete,
        tests_complete
    ])

    hp_status = package.get('HPStatus') or 'Pending'
    actual_date = package.get('ActualDate')

    if not (joints_complete and tests_complete):
        status = TestPackageStatus.CONSTRUCTION_INCOMPLETE
    elif not document_ready:
        status = TestPackageStatus.CONSTRUCTION_COMPLETE
    elif hp_status == 'Completed' and actual_date:
        status = TestPackageStatus.TEST_COMPLETE
    else:
        status = TestPackageStatus.DOCUMENT_READY

    package['status_info'] = {
        'status': status,
        'status_name': TestPackageStatus.STATUS_NAMES[status],
        'status_color': TestPackageStatus.STATUS_COLORS[status],
        'status_description': TestPackageStatus.STATUS_DESCRIPTIONS[status]
    }
    package['welding_status'] = {
        'completed_din': completed_din,
        'total_din': total_din,
        'progress_pct': int(round((completed_din / total_din) * 100)) if total_din else 0,
        'joints_complete': joints_complete
    }
    package['info_status'] = {
        'basic': basic_info,
        'attachments': attachments_ready
    }
    package['drawings_ready'] = pid_ready or iso_ready
    package['document_ready'] = document_ready
    package['punch_unresolved'] = extras['punch_open'].get(tp_id, 0)
    package['tests_complete'] = tests_complete


def build_pagination(page: int, total_count: int, base_path: str):
    total_pages = max(1, ceil(total_count / PER_PAGE))
    page = max(1, min(page, total_pages))
    start_index = (page - 1) * PER_PAGE + 1 if total_count else 0
    end_index = min(page * PER_PAGE, total_count)
    window = list(range(max(1, page - 2), min(total_pages, page + 2) + 1))
    return {
        'current_page': page,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages,
        'base_url': base_path,  # 添加 base_url 字段，供模板使用
        'prev_url': f"{base_path}{page - 1}" if page > 1 else None,
        'next_url': f"{base_path}{page + 1}" if page < total_pages else None,
        'start_index': start_index,
        'end_index': end_index,
        'window': window
    }


@test_package_bp.route('/test_packages')
def test_packages():
    query = (request.args.get('q') or '').strip()
    filter_system = (request.args.get('system_code') or '').strip()
    filter_subsystem = (request.args.get('subsystem_code') or '').strip()
    filter_status = (request.args.get('status') or '').strip()
    fac_filters = {
        'subproject_code': (request.args.get('subproject_code') or '').strip(),
        'train': (request.args.get('train') or '').strip(),
        'unit': (request.args.get('unit') or '').strip(),
        'simpleblk': (request.args.get('simpleblk') or '').strip(),
        'mainblock': (request.args.get('mainblock') or '').strip(),
        'block': (request.args.get('block') or '').strip(),
        'bccquarter': (request.args.get('bccquarter') or '').strip()
    }
    sort_order = (request.args.get('sort') or '').strip()
    page = max(int(request.args.get('page', '1') or 1), 1)

    systems = SystemModel.get_all_systems()
    subsystems = SubsystemModel.get_subsystems_by_system(filter_system) if filter_system else []
    faclist_options = get_faclist_filter_options(
        filter_subproject=fac_filters['subproject_code'] or None,
        filter_train=fac_filters['train'] or None,
        filter_unit=fac_filters['unit'] or None,
        filter_simpleblk=fac_filters['simpleblk'] or None,
        filter_mainblock=fac_filters['mainblock'] or None,
        filter_block=fac_filters['block'] or None,
        filter_bccquarter=fac_filters['bccquarter'] or None
    )

    matched_drawing_numbers = match_faclist_drawings(fac_filters)
    if matched_drawing_numbers == set():
        packages = []
        total_count = completed_count = in_progress_count = pending_count = 0
    else:
        packages, total_count, completed_count, in_progress_count, pending_count = TestPackageModel.list_test_packages(
            search=query or None,
            system_code=filter_system or None,
            subsystem_code=filter_subsystem or None,
            status=filter_status or None,
            allowed_drawing_numbers=list(matched_drawing_numbers) if matched_drawing_numbers is not None else None,
            page=page,
            per_page=PER_PAGE,
            sort_order=sort_order or None
        )

    package_ids = [p['TestPackageID'] for p in packages]
    extras = load_package_extras(package_ids)
    for pkg in packages:
        pkg['Status'] = pkg.get('HPStatus') or 'Pending'
        compute_package_status(pkg, extras)

    pagination_base = build_pagination_base_path(request.args, '/test_packages')
    pagination = build_pagination(page, total_count, pagination_base)

    return render_template(
        'test_package_list_industrial.html',
        packages=packages,
        systems=systems,
        subsystems=subsystems,
        faclist_options=faclist_options,
        search_query=query,
        filter_system=filter_system,
        filter_subsystem=filter_subsystem,
        filter_status=filter_status,
        total_count=total_count,
        completed_count=completed_count,
        in_progress_count=in_progress_count,
        pending_count=pending_count,
        pagination=pagination,
        active_page='test_packages'
    )


def _faclist_options_response():
    return get_faclist_filter_options(
        filter_subproject=(request.args.get('subproject_code') or '').strip() or None,
        filter_train=(request.args.get('train') or '').strip() or None,
        filter_unit=(request.args.get('unit') or '').strip() or None,
        filter_simpleblk=(request.args.get('simpleblk') or '').strip() or None,
        filter_mainblock=(request.args.get('mainblock') or '').strip() or None,
        filter_block=(request.args.get('block') or '').strip() or None,
        filter_bccquarter=(request.args.get('bccquarter') or '').strip() or None
    )


test_package_bp.route('/test_packages/filter_options')(lambda: jsonify(_faclist_options_response()))
@test_package_bp.route('/api/faclist_options')
def api_faclist_options():
    return jsonify(_faclist_options_response())


@test_package_bp.route('/test_packages/alerts')
def test_package_alerts_page():
    status = (request.args.get('status') or 'PENDING').upper()
    alerts = get_pipeline_alerts(status)
    return render_template(
        'test_package_alerts.html',
        alerts=alerts,
        pipeline_alert_threshold=PIPELINE_SYSTEM_SHARE_THRESHOLD,
        current_alert_status=status,
        active_page='test_packages'
    )


@test_package_bp.route('/test_packages/alerts/export')
def export_test_package_alerts():
    status = (request.args.get('status') or 'PENDING').upper()
    alerts = get_pipeline_alerts(status)
    columns = ['系统编码', '管线号', '管线完成率', '系统完成占比', 'DIN总量', 'DIN完成量']
    rows = []
    for alert in alerts or []:
        completion_rate = round(float(alert.get('CompletionRate') or 0) * 100, 2)
        system_share = round(float(alert.get('SystemDINShare') or 0) * 100, 2)
        rows.append({
            '系统编码': alert.get('SystemCode') or '',
            '管线号': alert.get('PipelineNumber') or '',
            '管线完成率': completion_rate,
            '系统完成占比': system_share,
            'DIN总量': float(alert.get('TotalDIN') or 0),
            'DIN完成量': float(alert.get('CompletedDIN') or 0)
        })
    df = pd.DataFrame(rows, columns=columns)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='编制提醒', index=False)
    output.seek(0)
    filename = f"TestPackagePreparationAlerts_{status}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@test_package_bp.route('/test_packages/alerts/<int:alert_id>', methods=['POST'])
def handle_test_package_alert(alert_id):
    data = request.get_json(silent=True) or {}
    action = (data.get('action') or '').upper()
    if action not in {'ACKED', 'IGNORED'}:
        return jsonify({'success': False, 'message': '无效操作'}), 400
    if update_pipeline_alert(alert_id, action):
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': '更新失败'}), 500


@test_package_bp.route('/api/test_packages/alerts')
def api_test_package_alerts():
    status = (request.args.get('status') or 'PENDING').upper()
    alerts = get_pipeline_alerts(status)
    return jsonify({'success': True, 'alerts': alerts, 'threshold': PIPELINE_SYSTEM_SHARE_THRESHOLD})


@test_package_bp.route('/test_packages/add', methods=['GET', 'POST'])
def add_test_package():
    message = "试压包由焊口同步生成，无法手工新增。"
    return f"""
    <!doctype html>
    <html>
        <head><title>提示</title>{get_bootstrap_css()}</head>
        <body>
            {get_navbar()}
            <div class='container mt-4'>
                <div class='alert alert-warning'>{message}</div>
                <a href='/test_packages' class='btn btn-secondary'>返回试压包列表</a>
            </div>
        </body>
    </html>
    """


@test_package_bp.route('/test_packages/edit/<test_package_id>', methods=['GET', 'POST'])
def edit_test_package(test_package_id):
    normalized_id = unquote(test_package_id)
    ensure_hydro_columns()
    conn = create_connection()
    if not conn:
        return "数据库连接失败", 500
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM HydroTestPackageList WHERE TestPackageID = %s", (normalized_id,))
        test_package = cur.fetchone()
        if not test_package:
            cur.execute(
                """
                SELECT COALESCE(NULLIF(TRIM(SystemCode), ''), NULL) AS SystemCode,
                       COALESCE(NULLIF(TRIM(SubSystemCode), ''), NULL) AS SubSystemCode
                FROM WeldingList
                WHERE TestPackageID = %s
                GROUP BY SystemCode, SubSystemCode
                LIMIT 1
                """,
                (normalized_id,)
            )
            codes = cur.fetchone() or {}
            cur.execute(
                """
                INSERT INTO HydroTestPackageList (TestPackageID, SystemCode, SubSystemCode, Description, Status, created_by, last_updated_by)
                VALUES (%s, %s, %s, %s, 'Pending', 'system', 'system')
                """,
                (normalized_id, codes.get('SystemCode'), codes.get('SubSystemCode'), normalized_id)
            )
            conn.commit()
            cur.execute("SELECT * FROM HydroTestPackageList WHERE TestPackageID = %s", (normalized_id,))
            test_package = cur.fetchone()

        systems = SystemModel.get_all_systems()
        subsystems = SubsystemModel.get_subsystems_by_system(test_package.get('SystemCode')) if test_package.get('SystemCode') else []
        success_flag = False
        error_message = None

        if request.method == 'POST':
            update_data = {
                'SystemCode': request.form.get('SystemCode') or None,
                'SubSystemCode': request.form.get('SubSystemCode') or None,
                'Description': request.form.get('Description') or normalized_id,
                # 新增：管道材料
                'PipeMaterial': request.form.get('PipeMaterial') or None,
                'TestType': request.form.get('TestType') or None,
                'TestMedium': request.form.get('TestMedium') or None,
                'DesignPressure': request.form.get('DesignPressure') or None,
                'TestPressure': request.form.get('TestPressure') or None,
                'PlannedDate': request.form.get('PlannedDate') or None,
                'ActualDate': request.form.get('ActualDate') or None,
                'Status': request.form.get('Status') or 'Pending',
                'Remarks': request.form.get('Remarks') or '',
                'last_updated_by': 'web'
            }
            import logging
            logger = logging.getLogger('routes.test_package_routes')
            client_ip = request.remote_addr
            logger.info(f'[API] 更新试压包请求: test_package_id={normalized_id}, 客户端IP: {client_ip}')
            
            if TestPackageModel.update_test_package(normalized_id, update_data):
                logger.info(f'[API] 试压包更新成功: test_package_id={normalized_id}, 客户端IP: {client_ip}')
                # 确保重定向响应正确发送
                redirect_response = redirect(f"/test_packages/edit/{normalized_id}?success=1")
                logger.info(f'[API] 准备返回重定向响应，客户端IP: {client_ip}')
                return redirect_response
            logger.error(f'[API] 试压包更新失败: test_package_id={normalized_id}, 客户端IP: {client_ip}')
            error_message = '保存失败，请检查输入。'
            for key, value in update_data.items():
                if value is not None:
                    test_package[key] = value
        else:
            success_flag = request.args.get('success') == '1'

        # 查询每个附件模块的附件数量
        attachment_status = {}
        for module in ATTACHMENT_MODULES:
            cur.execute("""
                SELECT COUNT(*) as count 
                FROM TestPackageAttachments 
                WHERE TestPackageID = %s AND ModuleName = %s
            """, (normalized_id, module['module']))
            result = cur.fetchone()
            attachment_status[module['module']] = result['count'] if result else 0

        # 为每个模块添加附件数量信息
        attachment_modules_with_status = []
        for module in ATTACHMENT_MODULES:
            module_copy = module.copy()
            module_copy['attachment_count'] = attachment_status.get(module['module'], 0)
            module_copy['has_attachments'] = module_copy['attachment_count'] > 0
            attachment_modules_with_status.append(module_copy)

        return render_template(
            'test_package_edit_industrial.html',
            test_package=test_package,
            systems=systems,
            subsystems=subsystems,
            success=success_flag,
            error_message=error_message,
            attachment_modules=attachment_modules_with_status,
            active_page='test_packages'
        )
    finally:
        conn.close()


def _render_punch_form(title, action_url, test_package_id, punch=None, errors=None):
    punch = punch or {}

    def _val(key):
        value = punch.get(key)
        if value is None:
            return ''
        if isinstance(value, datetime):
            return html.escape(value.strftime('%Y-%m-%d %H:%M:%S'))
        return html.escape(str(value))

    def _datetime_val(key):
        value = punch.get(key)
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%dT%H:%M')
        if isinstance(value, str) and value:
            return html.escape(value.replace(' ', 'T'))
        return ''

    rectified_flag = (punch.get('Rectified') or 'N').upper()
    verified_flag = (punch.get('Verified') or 'N').upper()
    errors = errors or []
    error_html = ''.join(f"<div class='alert alert-danger'>{html.escape(msg)}</div>" for msg in errors)
    back_url = f"/test_packages/edit/{html.escape(test_package_id)}#punch"
    return f"""
    <!doctype html>
    <html>
        <head>
            <title>{html.escape(title)}</title>
            {get_bootstrap_css()}
        </head>
        <body>
            {get_navbar()}
            <div class="container mt-4">
                <h1 class="mb-3">{html.escape(title)}</h1>
                <p class="text-muted">试压包：{html.escape(test_package_id)}</p>
                <a href="{back_url}" class="btn btn-secondary mb-3">返回试压包页面</a>
                {error_html}
                <form method="post" action="{html.escape(action_url)}">
                    <div class="row">
                        <div class="col-md-4 mb-3">
                            <label class="form-label">Punch No.</label>
                            <input type="text" class="form-control" name="PunchNo" value="{_val('PunchNo')}">
                        </div>
                        <div class="col-md-4 mb-3">
                            <label class="form-label">ISO Drawing No. *</label>
                            <input type="text" class="form-control" name="ISODrawingNo" value="{_val('ISODrawingNo')}" required>
                        </div>
                        <div class="col-md-2 mb-3">
                            <label class="form-label">Sheet No.</label>
                            <input type="text" class="form-control" name="SheetNo" value="{_val('SheetNo')}">
                        </div>
                        <div class="col-md-2 mb-3">
                            <label class="form-label">Rev No.</label>
                            <input type="text" class="form-control" name="RevNo" value="{_val('RevNo')}">
                        </div>
                    </div>
                    <div class="mb-3">
                        <label class="form-label">Description *</label>
                        <textarea class="form-control" name="Description" rows="3" required>{_val('Description')}</textarea>
                    </div>
                    <div class="row">
                        <div class="col-md-4 mb-3">
                            <label class="form-label">Category</label>
                            <input type="text" class="form-control" name="Category" value="{_val('Category')}">
                        </div>
                        <div class="col-md-4 mb-3">
                            <label class="form-label">Cause</label>
                            <input type="text" class="form-control" name="Cause" value="{_val('Cause')}">
                        </div>
                        <div class="col-md-4 mb-3">
                            <label class="form-label">Issued By</label>
                            <input type="text" class="form-control" name="IssuedBy" value="{_val('IssuedBy')}">
                        </div>
                    </div>
                    <div class="row">
                        <div class="col-md-3 mb-3">
                            <label class="form-label">Rectified</label>
                            <select class="form-select" name="Rectified">
                                <option value="N" {"selected" if rectified_flag != 'Y' else ""}>No</option>
                                <option value="Y" {"selected" if rectified_flag == 'Y' else ""}>Yes</option>
                            </select>
                        </div>
                        <div class="col-md-3 mb-3">
                            <label class="form-label">Rectified Date</label>
                            <input type="datetime-local" class="form-control" name="RectifiedDate" value="{_datetime_val('RectifiedDate')}">
                        </div>
                        <div class="col-md-3 mb-3">
                            <label class="form-label">Verified</label>
                            <select class="form-select" name="Verified">
                                <option value="N" {"selected" if verified_flag != 'Y' else ""}>No</option>
                                <option value="Y" {"selected" if verified_flag == 'Y' else ""}>Yes</option>
                            </select>
                        </div>
                        <div class="col-md-3 mb-3">
                            <label class="form-label">Verified Date</label>
                            <input type="datetime-local" class="form-control" name="VerifiedDate" value="{_datetime_val('VerifiedDate')}">
                        </div>
                    </div>
                    <div class="d-flex justify-content-end gap-2">
                        <a class="btn btn-outline-secondary" href="{back_url}">取消</a>
                        <button type="submit" class="btn btn-primary">保存</button>
                    </div>
                </form>
            </div>
        </body>
    </html>
    """


def _validate_punch_form(punch):
    errors = []
    if not (punch.get('ISODrawingNo') or '').strip():
        errors.append('ISO Drawing Number 为必填项。')
    if not (punch.get('Description') or '').strip():
        errors.append('请填写 Punch 描述。')
    rect_flag = parse_flag(punch.get('Rectified'))
    ver_flag = parse_flag(punch.get('Verified'))
    if rect_flag is None and punch.get('Rectified'):
        errors.append('整改状态只能为 Y/N。')
    if ver_flag is None and punch.get('Verified'):
        errors.append('验收状态只能为 Y/N。')
    return errors


def _collect_punch_form_data(form):
    return {
        'PunchNo': (form.get('PunchNo') or '').strip(),
        'ISODrawingNo': (form.get('ISODrawingNo') or '').strip(),
        'SheetNo': (form.get('SheetNo') or '').strip(),
        'RevNo': (form.get('RevNo') or '').strip(),
        'Description': (form.get('Description') or '').strip(),
        'Category': (form.get('Category') or '').strip(),
        'Cause': (form.get('Cause') or '').strip(),
        'IssuedBy': (form.get('IssuedBy') or '').strip(),
        'Rectified': (form.get('Rectified') or 'N').upper(),
        'RectifiedDate': (form.get('RectifiedDate') or '').strip(),
        'Verified': (form.get('Verified') or 'N').upper(),
        'VerifiedDate': (form.get('VerifiedDate') or '').strip(),
    }


def _punch_form_to_record(punch):
    def clean(key):
        value = punch.get(key)
        if value is None:
            return None
        value = str(value).strip()
        return value or None

    rect_flag = parse_flag(punch.get('Rectified')) or 'N'
    ver_flag = parse_flag(punch.get('Verified')) or 'N'
    record = {
        'PunchNo': clean('PunchNo'),
        'ISODrawingNo': clean('ISODrawingNo'),
        'SheetNo': clean('SheetNo'),
        'RevNo': clean('RevNo'),
        'Description': (punch.get('Description') or '').strip(),
        'Category': clean('Category'),
        'Cause': clean('Cause'),
        'IssuedBy': clean('IssuedBy'),
        'Rectified': rect_flag,
        'Verified': ver_flag,
        'RectifiedDate': parse_datetime(punch.get('RectifiedDate')),
        'VerifiedDate': parse_datetime(punch.get('VerifiedDate')),
    }
    if record['Rectified'] == 'Y' and record['RectifiedDate'] is None:
        record['RectifiedDate'] = datetime.now()
    if record['Verified'] == 'Y' and record['VerifiedDate'] is None:
        record['VerifiedDate'] = datetime.now()
    return record


@test_package_bp.route('/test_packages/export')
def export_test_packages():
    query = (request.args.get('q') or '').strip()
    filter_system = (request.args.get('system_code') or '').strip()
    filter_subsystem = (request.args.get('subsystem_code') or '').strip()
    filter_status = (request.args.get('status') or '').strip()
    fac_filters = {
        'subproject_code': (request.args.get('subproject_code') or '').strip(),
        'train': (request.args.get('train') or '').strip(),
        'unit': (request.args.get('unit') or '').strip(),
        'simpleblk': (request.args.get('simpleblk') or '').strip(),
        'mainblock': (request.args.get('mainblock') or '').strip(),
        'block': (request.args.get('block') or '').strip(),
        'bccquarter': (request.args.get('bccquarter') or '').strip()
    }
    sort_order = (request.args.get('sort') or '').strip()

    matched_drawing_numbers = match_faclist_drawings(fac_filters)
    packages, *_ = TestPackageModel.list_test_packages(
        search=query or None,
        system_code=filter_system or None,
        subsystem_code=filter_subsystem or None,
        status=filter_status or None,
        allowed_drawing_numbers=list(matched_drawing_numbers) if matched_drawing_numbers is not None else None,
        page=1,
        per_page=5000,
        sort_order=sort_order or None
    )
    selected_columns = request.args.getlist('columns') or None
    return export_test_packages_to_excel(packages, selected_columns)


@test_package_bp.route('/test_packages/delete/<test_package_id>', methods=['POST'])
def delete_test_package(test_package_id):
    import logging
    logger = logging.getLogger('routes.test_package_routes')
    client_ip = request.remote_addr
    logger.info(f'[API] 删除试压包请求: test_package_id={test_package_id}, 客户端IP: {client_ip}')
    
    TestPackageModel.delete_test_package(test_package_id)
    logger.info(f'[API] 试压包删除成功: test_package_id={test_package_id}, 客户端IP: {client_ip}')
    
    redirect_response = redirect('/test_packages')
    return redirect_response


@test_package_bp.route('/api/subsystems/<system_code>')
def get_subsystems_by_system(system_code):
    return jsonify(SubsystemModel.get_subsystems_by_system(system_code))


@test_package_bp.route('/api/test_packages/<test_package_id>/pid_list', methods=['GET'])
def get_pid_list(test_package_id):
    conn = create_connection()
    if not conn:
        return jsonify([])
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT ID, TestPackageID, PIDNo, RevNo, created_at, updated_at
            FROM PIDList
            WHERE TestPackageID = %s
            ORDER BY ID
            """,
            (test_package_id,)
        )
        rows = cur.fetchall()
        for row in rows:
            for key in ('created_at', 'updated_at'):
                if row.get(key):
                    row[key] = row[key].strftime('%Y-%m-%d %H:%M:%S')
        return jsonify(rows)
    finally:
        conn.close()


@test_package_bp.route('/api/test_packages/<test_package_id>/pid_list', methods=['POST'])
def add_pid(test_package_id):
    data = request.json or {}
    conn = create_connection()
    if not conn:
        return jsonify({'error': '数据库连接失败'}), 500
    try:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO PIDList (TestPackageID, PIDNo, RevNo) VALUES (%s, %s, %s)",
            (test_package_id, data.get('PIDNo'), data.get('RevNo'))
        )
        conn.commit()
        return jsonify({'success': True, 'id': cur.lastrowid})
    except Exception as exc:
        conn.rollback()
        current_app.logger.error(f"添加 PID 失败: {exc}", exc_info=True)
        return jsonify({'error': '服务器内部错误，请稍后重试'}), 500
    finally:
        conn.close()


@test_package_bp.route('/api/test_packages/<test_package_id>/pid_list/<int:pid_id>', methods=['DELETE'])
def delete_pid(test_package_id, pid_id):
    import logging
    logger = logging.getLogger('routes.test_package_routes')
    client_ip = request.remote_addr
    logger.info(f'[API] 删除PID请求: test_package_id={test_package_id}, pid_id={pid_id}, 客户端IP: {client_ip}')
    
    conn = create_connection()
    if not conn:
        logger.error(f'[API] 数据库连接失败，客户端IP: {client_ip}')
        response = jsonify({'error': '数据库连接失败'})
        response.headers['Content-Length'] = str(len(response.get_data()))
        return response, 500
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM PIDList WHERE ID = %s AND TestPackageID = %s", (pid_id, test_package_id))
        conn.commit()
        logger.info(f'[API] PID删除成功: pid_id={pid_id}, 客户端IP: {client_ip}')
        
        response = jsonify({'success': True})
        response_data = response.get_data()
        response.headers['Content-Length'] = str(len(response_data))
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        return response
    finally:
        if conn:
            conn.close()


@test_package_bp.route('/api/test_packages/<test_package_id>/iso_list', methods=['GET'])
def get_iso_list(test_package_id):
    conn = create_connection()
    if not conn:
        return jsonify([])
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT DISTINCT DrawingNumber AS ISODrawingNo
            FROM WeldingList
            WHERE TestPackageID = %s AND DrawingNumber LIKE '%ISO%'
            ORDER BY DrawingNumber
            """,
            (test_package_id,)
        )
        rows = cur.fetchall()
        for idx, row in enumerate(rows, start=1):
            row['ID'] = idx
            row['RevNo'] = ''
        return jsonify(rows)
    finally:
        conn.close()


@test_package_bp.route('/api/test_packages/<test_package_id>/iso_list', methods=['POST'])
def add_iso(test_package_id):
    data = request.json or {}
    conn = create_connection()
    if not conn:
        return jsonify({'error': '数据库连接失败'}), 500
    try:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO ISODrawingList (TestPackageID, ISODrawingNo, RevNo) VALUES (%s, %s, %s)",
            (test_package_id, data.get('ISODrawingNo'), data.get('RevNo'))
        )
        conn.commit()
        return jsonify({'success': True, 'id': cur.lastrowid})
    except Exception as exc:
        conn.rollback()
        current_app.logger.error(f"添加 ISO 失败: {exc}", exc_info=True)
        return jsonify({'error': '服务器内部错误，请稍后重试'}), 500
    finally:
        conn.close()


@test_package_bp.route('/api/test_packages/<test_package_id>/iso_list/<int:iso_id>', methods=['DELETE'])
def delete_iso(test_package_id, iso_id):
    import logging
    logger = logging.getLogger('routes.test_package_routes')
    client_ip = request.remote_addr
    logger.info(f'[API] 删除ISO请求: test_package_id={test_package_id}, iso_id={iso_id}, 客户端IP: {client_ip}')
    
    conn = create_connection()
    if not conn:
        logger.error(f'[API] 数据库连接失败，客户端IP: {client_ip}')
        response = jsonify({'error': '数据库连接失败'})
        response.headers['Content-Length'] = str(len(response.get_data()))
        return response, 500
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM ISODrawingList WHERE ID = %s AND TestPackageID = %s", (iso_id, test_package_id))
        conn.commit()
        logger.info(f'[API] ISO删除成功: iso_id={iso_id}, 客户端IP: {client_ip}')
        
        response = jsonify({'success': True})
        response_data = response.get_data()
        response.headers['Content-Length'] = str(len(response_data))
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        return response
    finally:
        if conn:
            conn.close()


@test_package_bp.route('/api/test_packages/<test_package_id>/punch_list', methods=['GET'])
def get_punch_list(test_package_id):
    conn = create_connection()
    if not conn:
        return jsonify([])
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT ID, PunchNo, ISODrawingNo, SheetNo, RevNo, Description, Category, Cause,
                   IssuedBy, Rectified, RectifiedDate, Verified, VerifiedDate, Deleted, created_at, updated_at
            FROM PunchList
            WHERE TestPackageID = %s AND (Deleted IS NULL OR Deleted != 'Y')
            ORDER BY ID
            """,
            (test_package_id,)
        )
        rows = cur.fetchall()
        for row in rows:
            for key in ('RectifiedDate', 'VerifiedDate', 'created_at', 'updated_at'):
                if row.get(key):
                    row[key] = row[key].strftime('%Y-%m-%d %H:%M:%S')
        return jsonify(rows)
    finally:
        conn.close()


@test_package_bp.route('/api/test_packages/<test_package_id>/punch_list', methods=['POST'])
def add_punch(test_package_id):
    data = request.json or {}
    conn = create_connection()
    if not conn:
        return jsonify({'error': '数据库连接失败'}), 500
    try:
        cur = conn.cursor()
        
        # 自动生成唯一的PunchNo
        # 格式：{TestPackageID}-P{序号}，例如 "TP001-P001"
        normalized_id = unquote(test_package_id)
        cur.execute(
            """
            SELECT COALESCE(MAX(CAST(SUBSTRING_INDEX(PunchNo, '-P', -1) AS UNSIGNED)), 0) as max_num
            FROM PunchList
            WHERE TestPackageID = %s AND PunchNo LIKE %s
            """,
            (normalized_id, normalized_id + "-P%")
        )
        result = cur.fetchone()
        max_num = result[0] if result else 0
        # 确保为整数再进行格式化，避免数据库返回的类型异常导致格式错误
        try:
            next_num_int = int(max_num) + 1
        except (TypeError, ValueError):
            next_num_int = 1
        # 使用字符串拼接生成 PunchNo（形如 TPID-P001）
        punch_no = normalized_id + "-P" + "{:03d}".format(next_num_int)
        
        cur.execute(
            """
            INSERT INTO PunchList (
                PunchNo, TestPackageID, ISODrawingNo, SheetNo, RevNo, Description,
                Category, Cause, IssuedBy, Rectified, RectifiedDate, Verified, VerifiedDate, Deleted
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                punch_no,
                normalized_id,
                data.get('ISODrawingNo'),
                data.get('SheetNo'),
                data.get('RevNo'),
                data.get('Description'),
                data.get('Category'),
                data.get('Cause'),
                data.get('IssuedBy'),
                data.get('Rectified', 'N'),
                parse_datetime(data.get('RectifiedDate')),
                data.get('Verified', 'N'),
                parse_datetime(data.get('VerifiedDate')),
                data.get('Deleted', 'N')
            )
        )
        conn.commit()
        return jsonify({'success': True, 'id': cur.lastrowid, 'punch_no': punch_no})
    except Exception as exc:
        conn.rollback()
        current_app.logger.error(f"添加Punch记录失败: {exc}", exc_info=True)
        return jsonify({'error': '服务器内部错误，请稍后重试'}), 500
    finally:
        conn.close()


@test_package_bp.route('/api/test_packages/<test_package_id>/punch_list/<int:punch_id>', methods=['DELETE'])
def delete_punch(test_package_id, punch_id):
    import logging
    logger = logging.getLogger('routes.test_package_routes')
    client_ip = request.remote_addr
    logger.info(f'[API] 删除Punch请求: test_package_id={test_package_id}, punch_id={punch_id}, 客户端IP: {client_ip}')
    
    conn = create_connection()
    if not conn:
        logger.error(f'[API] 数据库连接失败，客户端IP: {client_ip}')
        response = jsonify({'error': '数据库连接失败'})
        response.headers['Content-Length'] = str(len(response.get_data()))
        return response, 500
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM PunchList WHERE ID = %s AND TestPackageID = %s", (punch_id, test_package_id))
        conn.commit()
        logger.info(f'[API] Punch删除成功: punch_id={punch_id}, 客户端IP: {client_ip}')
        
        response = jsonify({'success': True})
        response_data = response.get_data()
        response.headers['Content-Length'] = str(len(response_data))
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        return response
    finally:
        if conn:
            conn.close()


@test_package_bp.route('/test_packages/<test_package_id>/punch/import/template')
def download_punch_template(test_package_id):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
    from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
    
    ensure_punch_list_schema()
    normalized_id = unquote(test_package_id)
    system_code = ''
    subsystem_code = ''
    punch_data = []
    conn = create_connection()
    if conn:
        try:
            cur = conn.cursor(dictionary=True)
            cur.execute(
                """
                SELECT SystemCode, SubSystemCode
                FROM HydroTestPackageList
                WHERE TestPackageID = %s
                """,
                (normalized_id,)
            )
            row = cur.fetchone()
            if row:
                system_code = row.get('SystemCode') or ''
                subsystem_code = row.get('SubSystemCode') or ''
            
            # 查询已有的punch记录
            cur.execute(
                """
                SELECT PunchNo, ISODrawingNo, SheetNo, RevNo, Description,
                       Category, Cause, IssuedBy, created_at as IssuedDate,
                       Rectified, RectifiedDate, Verified, VerifiedDate, Deleted
                FROM PunchList
                WHERE TestPackageID = %s
                ORDER BY ID
                """,
                (normalized_id,)
            )
            punch_data = cur.fetchall()
            # 格式化日期字段
            for punch in punch_data:
                if punch.get('IssuedDate'):
                    punch['IssuedDate'] = punch['IssuedDate'].strftime('%Y-%m-%d %H:%M:%S')
                if punch.get('RectifiedDate'):
                    punch['RectifiedDate'] = punch['RectifiedDate'].strftime('%Y-%m-%d %H:%M:%S')
                if punch.get('VerifiedDate'):
                    punch['VerifiedDate'] = punch['VerifiedDate'].strftime('%Y-%m-%d %H:%M:%S')
        finally:
            conn.close()

    info_df = pd.DataFrame([{
        'TestPackageID': normalized_id,
        'SystemCode': system_code,
        'SubSystemCode': subsystem_code,
        'GeneratedAt': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }])
    # 包含PunchNo列（用于更新记录），但用户不能自定义，只能填写已有记录的PunchNo
    template_columns = [
        'PunchNo', 'ISODrawingNo', 'SheetNo', 'RevNo', 'Description',
        'Category', 'Cause', 'IssuedBy', 'IssuedDate',
        'Rectified', 'RectifiedDate', 'Verified', 'VerifiedDate', 'Deleted'
    ]
    # 如果有已有数据，使用已有数据；否则创建空模板
    if punch_data:
        template_df = pd.DataFrame(punch_data, columns=template_columns)
    else:
        template_df = pd.DataFrame(columns=template_columns)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        info_df.to_excel(writer, sheet_name='Info', index=False)
        template_df.to_excel(writer, sheet_name='PunchList', index=False)
        
        # 获取工作表对象以应用样式
        workbook = writer.book
        ws_punch = workbook['PunchList']
        
        # 定义样式
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 蓝色底纹
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # 设置列宽
        column_widths = {
            'A': 15,   # PunchNo（只读，用于更新）
            'B': 18,   # ISODrawingNo
            'C': 10,   # SheetNo
            'D': 10,   # RevNo
            'E': 40,   # Description
            'F': 15,   # Category
            'G': 15,   # Cause
            'H': 12,   # IssuedBy
            'I': 18,   # IssuedDate
            'J': 10,   # Rectified
            'K': 15,   # RectifiedDate
            'L': 10,   # Verified
            'M': 15,   # VerifiedDate
            'N': 10    # Deleted
        }
        for col, width in column_widths.items():
            ws_punch.column_dimensions[col].width = width
        
        # 设置行高
        ws_punch.row_dimensions[1].height = 25  # 表头行高
        
        # 应用表头样式（第1行）
        for col_idx, col_name in enumerate(template_columns, start=1):
            cell = ws_punch.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 为数据区域添加边框和数据验证（第2-100行，预留足够空间）
        # 找到相关列的索引
        punch_no_col = template_columns.index('PunchNo') + 1  # A列
        category_col = template_columns.index('Category') + 1  # F列
        cause_col = template_columns.index('Cause') + 1  # G列
        rectified_col = template_columns.index('Rectified') + 1  # I列
        verified_col = template_columns.index('Verified') + 1     # K列
        deleted_col = template_columns.index('Deleted') + 1       # M列
        
        # 创建数据验证：Category只能选择A/B/C/D
        category_validation = DataValidation(
            type="list",
            formula1='"A,B,C,D"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="输入错误",
            error="只能选择 A, B, C 或 D"
        )
        
        # 创建数据验证：Cause只能选择N/F/E
        cause_validation = DataValidation(
            type="list",
            formula1='"N,F,E"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="输入错误",
            error="只能选择 N(Non-Conformity), F(Field Revion) 或 E(Re-Engineering)"
        )
        
        # 创建数据验证：Rectified、Verified和Deleted只能选择Y或N
        rectified_validation = DataValidation(
            type="list",
            formula1='"Y,N"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="输入错误",
            error="只能输入 Y 或 N"
        )
        verified_validation = DataValidation(
            type="list",
            formula1='"Y,N"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="输入错误",
            error="只能输入 Y 或 N"
        )
        deleted_validation = DataValidation(
            type="list",
            formula1='"Y,N"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="输入错误",
            error="只能输入 Y 或 N（Y表示已删除，不计算在计数中）"
        )
        
        # 为PunchNo列添加注释说明
        from openpyxl.comments import Comment  # type: ignore
        punch_no_comment = Comment(
            "PunchNo由系统自动生成，新增记录时留空。\n如需更新或删除已有记录，请填写对应的PunchNo。",
            "系统提示"
        )
        
        for row in range(2, 101):
            for col in range(1, len(template_columns) + 1):
                cell = ws_punch.cell(row=row, column=col)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                # 为PunchNo列添加注释（仅第一行数据作为示例）
                if col == punch_no_col and row == 2:
                    cell.comment = punch_no_comment
                
                # 为Category列添加数据验证
                if col == category_col:
                    category_validation.add(cell)
                # 为Cause列添加数据验证
                if col == cause_col:
                    cause_validation.add(cell)
                # 为Rectified列添加数据验证
                if col == rectified_col:
                    rectified_validation.add(cell)
                # 为Verified列添加数据验证
                if col == verified_col:
                    verified_validation.add(cell)
                # 为Deleted列添加数据验证
                if col == deleted_col:
                    deleted_validation.add(cell)
        
        # 将数据验证添加到工作表
        ws_punch.add_data_validation(category_validation)
        ws_punch.add_data_validation(cause_validation)
        ws_punch.add_data_validation(rectified_validation)
        ws_punch.add_data_validation(verified_validation)
        ws_punch.add_data_validation(deleted_validation)
        
        # 冻结首行
        ws_punch.freeze_panes = 'A2'
        
    output.seek(0)
    filename = f"PunchListTemplate_{normalized_id}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@test_package_bp.route('/api/test_packages/<test_package_id>/punch_list/import', methods=['POST'])
def import_punch_list_api(test_package_id):
    ensure_punch_list_schema()
    upload = request.files.get('file')
    if not upload or upload.filename == '':
        return jsonify({'success': False, 'message': '请上传需要导入的 Excel 文件'}), 400
    file_bytes = upload.read()
    try:
        excel = pd.ExcelFile(BytesIO(file_bytes))
    except Exception as exc:
        return jsonify({'success': False, 'message': '读取文件失败: {}'.format(str(exc))}), 400

    sheet_name = next((name for name in excel.sheet_names if 'punch' in name.lower()), excel.sheet_names[0])
    df = excel.parse(sheet_name)
    if df.empty:
        return jsonify({'success': False, 'message': 'Excel 内容为空'}), 400

    rename_map = {}
    for column in df.columns:
        mapped = PUNCH_HEADER_MAP.get(normalize_header(column))
        if mapped:
            rename_map[column] = mapped
    df = df.rename(columns=rename_map)

    required_columns = ['ISODrawingNo', 'Description']
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        return jsonify({'success': False, 'message': "缺少必要字段: {}".format(', '.join(missing))}), 400

    records = []
    errors = []

    for idx, row in df.iterrows():
        line_no = idx + 2

        def cell(name):
            value = row.get(name)
            if pd.isna(value):
                return ''
            return str(value).strip()

        punch_no = cell('PunchNo')
        iso_no = cell('ISODrawingNo')
        description = cell('Description')
        
        # 如果有PunchNo，说明是更新已有记录；如果没有，说明是新增记录
        if punch_no:
            # 更新记录：只需要PunchNo即可，其他字段可选
            record = {
                'PunchNo': punch_no,
                'ISODrawingNo': iso_no or None,
                'SheetNo': cell('SheetNo') or None,
                'RevNo': cell('RevNo') or None,
                'Description': description or None,
                'Category': cell('Category') or None,
                'Cause': cell('Cause') or None,
                'IssuedBy': cell('IssuedBy') or None,
                'Rectified': parse_flag(cell('Rectified')) or None,
                'Verified': parse_flag(cell('Verified')) or None,
                'RectifiedDate': parse_datetime(cell('RectifiedDate')),
                'VerifiedDate': parse_datetime(cell('VerifiedDate')),
                'Deleted': parse_flag(cell('Deleted')) or 'N',
                'is_update': True
            }
        else:
            # 新增记录：必须要有ISO和Description
            if not iso_no:
                errors.append("第 {} 行：新增记录时，ISO 图号不能为空。".format(line_no))
                continue
            if not description:
                errors.append("第 {} 行：新增记录时，请填写 Description。".format(line_no))
                continue
            record = {
                'PunchNo': None,  # 由系统自动生成
                'ISODrawingNo': iso_no,
                'SheetNo': cell('SheetNo') or None,
                'RevNo': cell('RevNo') or None,
                'Description': description,
                'Category': cell('Category') or None,
                'Cause': cell('Cause') or None,
                'IssuedBy': cell('IssuedBy') or None,
                'Rectified': parse_flag(cell('Rectified')) or 'N',
                'Verified': parse_flag(cell('Verified')) or 'N',
                'RectifiedDate': parse_datetime(cell('RectifiedDate')),
                'VerifiedDate': parse_datetime(cell('VerifiedDate')),
                'Deleted': parse_flag(cell('Deleted')) or 'N',
                'is_update': False
            }
        records.append(record)

    if errors:
        return jsonify({'success': False, 'message': '数据校验失败', 'errors': errors}), 400
    if not records:
        return jsonify({'success': False, 'message': '没有可以导入的数据'}), 400

    conn = create_connection()
    if not conn:
        return jsonify({'success': False, 'message': '数据库连接失败'}), 500

    inserted = 0
    updated = 0
    normalized_id = unquote(test_package_id)
    try:
        cur = conn.cursor(dictionary=True)
        
        # 分离新增和更新的记录
        update_records = [r for r in records if r.get('is_update')]
        insert_records = [r for r in records if not r.get('is_update')]
        
        # 处理更新记录：通过PunchNo匹配
        if update_records:
            punch_numbers = [r['PunchNo'] for r in update_records]
            placeholders = ','.join(['%s'] * len(punch_numbers))
            # 直接使用字符串拼接构建SQL，避免format可能的问题
            query = """
                SELECT ID, PunchNo, ISODrawingNo, SheetNo, RevNo, Description,
                       Category, Cause, IssuedBy, Rectified, RectifiedDate, Verified, VerifiedDate, Deleted
                FROM PunchList
                WHERE TestPackageID = %s AND PunchNo IN (""" + placeholders + """)"""
            cur.execute(
                query,
                tuple([normalized_id] + punch_numbers)
            )
            existing_map = {row['PunchNo']: row for row in cur.fetchall()}
            
            for record in update_records:
                punch_no = record['PunchNo']
                if punch_no not in existing_map:
                    errors.append("PunchNo {} 不存在，无法更新。".format(punch_no))
                    continue
                
                existing = existing_map[punch_no]
                # 只更新提供的字段，未提供的字段保持原值
                update_fields = []
                update_values = []
                
                if record.get('ISODrawingNo'):
                    update_fields.append('ISODrawingNo=%s')
                    update_values.append(record['ISODrawingNo'])
                if record.get('SheetNo') is not None:
                    update_fields.append('SheetNo=%s')
                    update_values.append(record['SheetNo'])
                if record.get('RevNo') is not None:
                    update_fields.append('RevNo=%s')
                    update_values.append(record['RevNo'])
                if record.get('Description'):
                    update_fields.append('Description=%s')
                    update_values.append(record['Description'])
                if record.get('Category') is not None:
                    update_fields.append('Category=%s')
                    update_values.append(record['Category'])
                if record.get('Cause') is not None:
                    update_fields.append('Cause=%s')
                    update_values.append(record['Cause'])
                if record.get('IssuedBy') is not None:
                    update_fields.append('IssuedBy=%s')
                    update_values.append(record['IssuedBy'])
                if record.get('Rectified') is not None:
                    update_fields.append('Rectified=%s')
                    update_values.append(record['Rectified'])
                    if record['Rectified'] == 'Y' and not record.get('RectifiedDate'):
                        update_fields.append('RectifiedDate=%s')
                        update_values.append(datetime.now())
                if record.get('Verified') is not None:
                    update_fields.append('Verified=%s')
                    update_values.append(record['Verified'])
                    if record['Verified'] == 'Y' and not record.get('VerifiedDate'):
                        update_fields.append('VerifiedDate=%s')
                        update_values.append(datetime.now())
                if record.get('RectifiedDate') is not None:
                    update_fields.append('RectifiedDate=%s')
                    update_values.append(record['RectifiedDate'])
                if record.get('VerifiedDate') is not None:
                    update_fields.append('VerifiedDate=%s')
                    update_values.append(record['VerifiedDate'])
                if record.get('Deleted') is not None:
                    update_fields.append('Deleted=%s')
                    update_values.append(record['Deleted'])
                
                if update_fields:
                    update_values.extend([normalized_id, existing['ID']])
                    # 使用字符串拼接而不是format，避免用户输入中的花括号导致格式错误
                    fields_str = ', '.join(update_fields)
                    # 直接使用字符串拼接构建SQL，避免format可能的问题
                    query = "UPDATE PunchList SET " + fields_str + " WHERE TestPackageID=%s AND ID=%s"
                    cur.execute(
                        query,
                        tuple(update_values)
                    )
                    updated += 1
        
        # 处理新增记录：自动生成PunchNo
        if insert_records:
            # 获取当前最大序号，用于生成新的PunchNo
            cur.execute(
                """
                SELECT COALESCE(MAX(CAST(SUBSTRING_INDEX(PunchNo, '-P', -1) AS UNSIGNED)), 0) as max_num
                FROM PunchList
                WHERE TestPackageID = %s AND PunchNo LIKE %s
                """,
                (normalized_id, normalized_id + "-P%")
            )
            result = cur.fetchone()
            max_num = result['max_num'] if result else 0
            # 同样确保为整数，避免类型异常
            try:
                start_num = int(max_num)
            except (TypeError, ValueError):
                start_num = 0

            for record in insert_records:
                # 自动生成唯一的PunchNo
                start_num += 1
                punch_no = normalized_id + "-P" + "{:03d}".format(start_num)
                
                rectified_date = record['RectifiedDate'] or (datetime.now() if record['Rectified'] == 'Y' else None)
                verified_date = record['VerifiedDate'] or (datetime.now() if record['Verified'] == 'Y' else None)
                cur.execute(
                    """
                    INSERT INTO PunchList (
                        PunchNo, TestPackageID, ISODrawingNo, SheetNo, RevNo, Description,
                        Category, Cause, IssuedBy, Rectified, RectifiedDate, Verified, VerifiedDate, Deleted
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        punch_no, normalized_id, record['ISODrawingNo'], record['SheetNo'], record['RevNo'],
                        record['Description'], record['Category'], record['Cause'], record['IssuedBy'],
                        record['Rectified'], rectified_date, record['Verified'], verified_date, record['Deleted']
                    )
                )
                inserted += 1
        
        if errors:
            return jsonify({'success': False, 'message': '部分记录处理失败', 'errors': errors}), 400

        cur.execute(
            """
            INSERT INTO PunchListImportLog
                (TestPackageID, FileName, TotalCount, InsertedCount, UpdatedCount, ErrorCount, Message)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (
                normalized_id,
                upload.filename,
                len(records),
                inserted,
                updated,
                len(errors),
                '导入完成' if not errors else '导入完成，但有 {} 个错误'.format(len(errors))
            )
        )
        conn.commit()
        return jsonify({'success': True, 'inserted': inserted, 'updated': updated})
    except Exception as exc:
        conn.rollback()
        import traceback
        error_detail = traceback.format_exc()
        current_app.logger.error(f"导入Punch List失败: {exc}\n{error_detail}")
        # 使用字符串拼接而不是format，避免异常信息中的花括号导致格式错误
        return jsonify({'success': False, 'message': '导入失败: ' + str(exc)}), 500
    finally:
        conn.close()


@test_package_bp.route('/test_packages/<test_package_id>/punch/add', methods=['GET', 'POST'])
def add_punch_form(test_package_id):
    normalized_id = unquote(test_package_id)
    action_url = f"/test_packages/{normalized_id}/punch/add"
    punch_data = {'Rectified': 'N', 'Verified': 'N'}
    if request.method == 'POST':
        punch_data = _collect_punch_form_data(request.form)
        errors = _validate_punch_form(punch_data)
        if errors:
            return _render_punch_form("新增 Punch", action_url, normalized_id, punch=punch_data, errors=errors)
        record = _punch_form_to_record(punch_data)
        conn = create_connection()
        if not conn:
            return _render_punch_form("新增 Punch", action_url, normalized_id, punch=punch_data, errors=['数据库连接失败'])
        try:
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO PunchList (
                    PunchNo, TestPackageID, ISODrawingNo, SheetNo, RevNo, Description,
                    Category, Cause, IssuedBy, Rectified, RectifiedDate, Verified, VerifiedDate
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    record['PunchNo'], normalized_id, record['ISODrawingNo'], record['SheetNo'], record['RevNo'],
                    record['Description'], record['Category'], record['Cause'], record['IssuedBy'],
                    record['Rectified'], record['RectifiedDate'], record['Verified'], record['VerifiedDate']
                )
            )
            conn.commit()
            return redirect(f"/test_packages/edit/{normalized_id}#punch")
        except Exception as exc:
            conn.rollback()
            return _render_punch_form("新增 Punch", action_url, normalized_id, punch=punch_data, errors=[f'保存失败: {exc}'])
        finally:
            conn.close()
    return _render_punch_form("新增 Punch", action_url, normalized_id, punch=punch_data)


@test_package_bp.route('/test_packages/<test_package_id>/punch/edit/<int:punch_id>', methods=['GET', 'POST'])
def edit_punch_form(test_package_id, punch_id):
    normalized_id = unquote(test_package_id)
    conn = create_connection()
    if not conn:
        return "数据库连接失败", 500
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT ID, PunchNo, ISODrawingNo, SheetNo, RevNo, Description,
                   Category, Cause, IssuedBy, Rectified, RectifiedDate, Verified, VerifiedDate
            FROM PunchList
            WHERE TestPackageID = %s AND ID = %s
            """,
            (normalized_id, punch_id)
        )
        punch = cur.fetchone()
    finally:
        conn.close()

    if not punch:
        return "未找到 Punch 记录", 404

    punch_data = {
        'PunchNo': punch.get('PunchNo'),
        'ISODrawingNo': punch.get('ISODrawingNo'),
        'SheetNo': punch.get('SheetNo'),
        'RevNo': punch.get('RevNo'),
        'Description': punch.get('Description'),
        'Category': punch.get('Category'),
        'Cause': punch.get('Cause'),
        'IssuedBy': punch.get('IssuedBy'),
        'Rectified': punch.get('Rectified') or 'N',
        'RectifiedDate': punch.get('RectifiedDate'),
        'Verified': punch.get('Verified') or 'N',
        'VerifiedDate': punch.get('VerifiedDate'),
    }

    action_url = f"/test_packages/{normalized_id}/punch/edit/{punch_id}"
    if request.method == 'POST':
        punch_data = _collect_punch_form_data(request.form)
        errors = _validate_punch_form(punch_data)
        if errors:
            return _render_punch_form("编辑 Punch", action_url, normalized_id, punch=punch_data, errors=errors)
        record = _punch_form_to_record(punch_data)
        conn = create_connection()
        if not conn:
            return _render_punch_form("编辑 Punch", action_url, normalized_id, punch=punch_data, errors=['数据库连接失败'])
        try:
            cur = conn.cursor()
            cur.execute(
                """
                UPDATE PunchList
                SET PunchNo=%s, ISODrawingNo=%s, SheetNo=%s, RevNo=%s,
                    Description=%s, Category=%s, Cause=%s, IssuedBy=%s,
                    Rectified=%s, RectifiedDate=%s, Verified=%s, VerifiedDate=%s
                WHERE TestPackageID=%s AND ID=%s
                """,
                (
                    record['PunchNo'], record['ISODrawingNo'], record['SheetNo'], record['RevNo'],
                    record['Description'], record['Category'], record['Cause'], record['IssuedBy'],
                    record['Rectified'], record['RectifiedDate'], record['Verified'], record['VerifiedDate'],
                    normalized_id, punch_id
                )
            )
            conn.commit()
            return redirect(f"/test_packages/edit/{normalized_id}#punch")
        except Exception as exc:
            conn.rollback()
            return _render_punch_form("编辑 Punch", action_url, normalized_id, punch=punch_data, errors=[f'保存失败: {exc}'])
        finally:
            conn.close()
    return _render_punch_form("编辑 Punch", action_url, normalized_id, punch=punch_data)


@test_package_bp.route('/api/test_packages/<test_package_id>/attachments/<module_name>', methods=['GET'])
def get_attachments(test_package_id, module_name):
    conn = create_connection()
    if not conn:
        return jsonify([])
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT ID, FileName, FilePath, FileSize, UploadedBy, UploadedAt
            FROM TestPackageAttachments
            WHERE TestPackageID = %s AND ModuleName = %s
            ORDER BY UploadedAt DESC
            """,
            (test_package_id, module_name)
        )
        rows = cur.fetchall()
        for row in rows:
            if row.get('UploadedAt'):
                row['UploadedAt'] = row['UploadedAt'].strftime('%Y-%m-%d %H:%M:%S')
        return jsonify(rows)
    finally:
        conn.close()


@test_package_bp.route('/api/test_packages/<test_package_id>/attachments/<module_name>', methods=['POST'])
def upload_attachment(test_package_id, module_name):
    import logging
    logger = logging.getLogger('routes.test_package_routes')
    client_ip = request.remote_addr
    logger.info(f'[API] 上传附件请求: test_package_id={test_package_id}, module={module_name}, 客户端IP: {client_ip}')
    
    if 'files' not in request.files:
        response = jsonify({'error': '未选择文件'})
        return response, 400
    files = request.files.getlist('files')
    if not files:
        response = jsonify({'error': '未选择文件'})
        return response, 400

    conn = create_connection()
    if not conn:
        logger.error(f'[API] 数据库连接失败，客户端IP: {client_ip}')
        response = jsonify({'error': '数据库连接失败'})
        response.headers['Content-Length'] = str(len(response.get_data()))
        return response, 500
    try:
        folder = ensure_upload_folder(test_package_id, module_name)
        cur = conn.cursor()
        uploaded = []
        for file in files:
            ok, err = _validate_upload_file(file)
            if not ok:
                # 对于单个文件出错，跳过并记录一条错误信息
                current_app.logger.warning(f"试压包附件校验失败: {err} (filename={file.filename})")
                continue
            original = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{original}"
            file_path = os.path.join(folder, filename)
            file.save(file_path)
            file_size = os.path.getsize(file_path)
            cur.execute(
                """
                INSERT INTO TestPackageAttachments (TestPackageID, ModuleName, FileName, FilePath, FileSize, UploadedBy)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (test_package_id, module_name, original, file_path, file_size, 'web')
            )
            uploaded.append({'id': cur.lastrowid, 'filename': original, 'size': file_size})
        conn.commit()
        logger.info(f'[API] 附件上传成功: test_package_id={test_package_id}, module={module_name}, 文件数={len(uploaded)}, 客户端IP: {client_ip}')
        
        response = jsonify({'success': True, 'files': uploaded})
        response_data = response.get_data()
        response.headers['Content-Length'] = str(len(response_data))
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        return response
    except Exception as exc:
        import traceback
        logger.error(f'[API] 附件上传失败: {exc}, 客户端IP: {client_ip}')
        logger.error(f'[API] 错误堆栈: {traceback.format_exc()}')
        conn.rollback()
        response = jsonify({'error': str(exc)})
        response.headers['Content-Length'] = str(len(response.get_data()))
        return response, 500
    finally:
        if conn:
            conn.close()


@test_package_bp.route('/api/test_packages/<test_package_id>/attachments/<int:attachment_id>/download')
def download_attachment(test_package_id, attachment_id):
    import logging
    logger = logging.getLogger('routes.test_package_routes')
    client_ip = request.remote_addr
    
    conn = create_connection()
    if not conn:
        logger.error(f'[下载附件] 数据库连接失败，客户端IP: {client_ip}')
        return "数据库连接失败", 500
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT FilePath, FileName FROM TestPackageAttachments WHERE ID = %s AND TestPackageID = %s",
            (attachment_id, test_package_id)
        )
        row = cur.fetchone()
        
        if not row:
            logger.warning(f'[下载附件] 数据库未找到记录: attachment_id={attachment_id}, test_package_id={test_package_id}, 客户端IP: {client_ip}')
            return "文件不存在", 404
        
        file_path = row['FilePath']
        file_name = row['FileName']
        
        # 如果路径是相对路径，转换为绝对路径（基于应用根目录）
        if not os.path.isabs(file_path):
            # 获取应用根目录（Flask 应用实例的根路径）
            if current_app:
                app_root = current_app.root_path
            else:
                # 如果 current_app 不可用，使用当前文件所在目录的父目录作为项目根目录
                app_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            file_path = os.path.join(app_root, file_path)
        
        # 规范化路径（统一路径分隔符，Windows 上会将正斜杠转换为反斜杠）
        file_path = os.path.normpath(file_path)
        
        logger.info(f'[下载附件] 尝试下载: attachment_id={attachment_id}, test_package_id={test_package_id}, file_path={file_path}, 客户端IP: {client_ip}')
        
        # 检查文件是否存在
        if not os.path.exists(file_path):
            logger.error(f'[下载附件] 文件不存在: file_path={file_path}, 客户端IP: {client_ip}')
            return "文件不存在", 404
        
        # 检查是否是文件（而不是目录）
        if not os.path.isfile(file_path):
            logger.error(f'[下载附件] 路径不是文件: file_path={file_path}, 客户端IP: {client_ip}')
            return "文件不存在", 404
        
        logger.info(f'[下载附件] 文件下载成功: attachment_id={attachment_id}, file_name={file_name}, 客户端IP: {client_ip}')
        return send_file(file_path, as_attachment=True, download_name=file_name)
    except Exception as exc:
        logger.error(f'[下载附件] 下载异常: attachment_id={attachment_id}, test_package_id={test_package_id}, 错误={exc}, 客户端IP: {client_ip}', exc_info=True)
        return f"下载失败: {str(exc)}", 500
    finally:
        conn.close()


@test_package_bp.route('/api/test_packages/<test_package_id>/attachments/<int:attachment_id>', methods=['DELETE'])
def delete_attachment(test_package_id, attachment_id):
    import logging
    logger = logging.getLogger('routes.test_package_routes')
    client_ip = request.remote_addr
    logger.info(f'[API] 删除附件请求: test_package_id={test_package_id}, attachment_id={attachment_id}, 客户端IP: {client_ip}')
    
    conn = create_connection()
    if not conn:
        logger.error(f'[API] 数据库连接失败，客户端IP: {client_ip}')
        response = jsonify({'error': '数据库连接失败'})
        response.headers['Content-Length'] = str(len(response.get_data()))
        return response, 500
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT FilePath FROM TestPackageAttachments WHERE ID = %s AND TestPackageID = %s",
            (attachment_id, test_package_id)
        )
        row = cur.fetchone()
        if row and row.get('FilePath'):
            file_path = row['FilePath']
            # 如果路径是相对路径，转换为绝对路径（基于应用根目录）
            if not os.path.isabs(file_path):
                if current_app:
                    app_root = current_app.root_path
                else:
                    app_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                file_path = os.path.join(app_root, file_path)
            
            # 规范化路径（统一路径分隔符，Windows 上会将正斜杠转换为反斜杠）
            file_path = os.path.normpath(file_path)
            
            if os.path.exists(file_path) and os.path.isfile(file_path):
                os.remove(file_path)
                logger.info(f'[API] 已删除文件: {file_path}, 客户端IP: {client_ip}')
        cur.execute("DELETE FROM TestPackageAttachments WHERE ID = %s AND TestPackageID = %s", (attachment_id, test_package_id))
        conn.commit()
        logger.info(f'[API] 附件删除成功: attachment_id={attachment_id}, 客户端IP: {client_ip}')
        
        # 确保响应正确发送，避免连接重置
        response = jsonify({'success': True})
        response_data = response.get_data()
        response.headers['Content-Length'] = str(len(response_data))
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        logger.info(f'[API] 准备返回响应，大小: {len(response_data)} 字节，客户端IP: {client_ip}')
        return response
    except Exception as e:
        import traceback
        logger.error(f'[API] 删除附件失败: {e}, 客户端IP: {client_ip}')
        logger.error(f'[API] 错误堆栈: {traceback.format_exc()}')
        response = jsonify({'error': str(e)})
        response.headers['Content-Length'] = str(len(response.get_data()))
        return response, 500
    finally:
        if conn:
            conn.close()


@test_package_bp.route('/api/test_packages/<test_package_id>/joint_summary')
def get_joint_summary(test_package_id):
    conn = create_connection()
    if not conn:
        return jsonify({'error': '数据库连接失败'}), 500

    def _to_int(value):
        try:
            return int(value or 0)
        except (TypeError, ValueError):
            return 0

    def _to_float(value):
        try:
            return float(value or 0)
        except (TypeError, ValueError):
            return 0.0

    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT COUNT(*) AS total_joints,
                   SUM(CASE WHEN WeldDate IS NOT NULL OR Status IN ('已完成', 'Completed') THEN 1 ELSE 0 END) AS completed_joints,
                   COALESCE(SUM(Size), 0) AS total_din,
                   COALESCE(SUM(CASE WHEN WeldDate IS NOT NULL OR Status IN ('已完成', 'Completed') THEN Size ELSE 0 END), 0) AS completed_din
            FROM WeldingList
            WHERE TestPackageID = %s
            """,
            (test_package_id,)
        )
        overall_row = cur.fetchone() or {}
        overall = {
            'total_joints': _to_int(overall_row.get('total_joints')),
            'completed_joints': _to_int(overall_row.get('completed_joints')),
            'total_din': _to_float(overall_row.get('total_din')),
            'completed_din': _to_float(overall_row.get('completed_din'))
        }

        cur.execute(
            """
            SELECT wl.PipelineNumber AS pipeline_no,
                   COUNT(*) AS total_joints,
                   SUM(CASE WHEN wl.WeldDate IS NOT NULL OR wl.Status IN ('已完成', 'Completed') THEN 1 ELSE 0 END) AS completed_joints,
                   COALESCE(SUM(wl.Size), 0) AS total_din,
                   COALESCE(SUM(CASE WHEN wl.WeldDate IS NOT NULL OR wl.Status IN ('已完成', 'Completed') THEN wl.Size ELSE 0 END), 0) AS completed_din,
                   (
                       COUNT(DISTINCT NULLIF(TRIM(wl.WelderRoot), '')) +
                       COUNT(DISTINCT NULLIF(TRIM(wl.WelderFill), ''))
                   ) AS welder_count,
                   MAX(ll.NDEGrade) AS nde_grade,
                   MAX(CASE WHEN ll.LineID IS NOT NULL THEN 1 ELSE 0 END) AS matched_flag
            FROM WeldingList wl
            LEFT JOIN LineList ll ON wl.PipelineNumber = ll.LineID
            WHERE wl.TestPackageID = %s AND wl.PipelineNumber IS NOT NULL AND wl.PipelineNumber <> ''
            GROUP BY wl.PipelineNumber
            ORDER BY wl.PipelineNumber
            """,
            (test_package_id,)
        )
        rows = cur.fetchall()
        pipelines = []
        unmatched = 0
        for row in rows:
            matched = bool(row.get('matched_flag'))
            if not matched:
                unmatched += 1
            pipelines.append({
                'pipeline_no': row.get('pipeline_no'),
                'total_joints': _to_int(row.get('total_joints')),
                'completed_joints': _to_int(row.get('completed_joints')),
                'total_din': _to_float(row.get('total_din')),
                'completed_din': _to_float(row.get('completed_din')),
                'welder_count': _to_int(row.get('welder_count')),
                'nde_grade': row.get('nde_grade'),
                'matched': matched
            })

        return jsonify({
            'overall': overall,
            'pipelines': pipelines,
            'unmatched_count': unmatched
        })
    finally:
        conn.close()


@test_package_bp.route('/api/test_packages/<test_package_id>/nde_pwht_status')
def get_nde_pwht_status(test_package_id):
    refresh_nde_pwht_status(test_package_id)
    conn = create_connection()
    if not conn:
        return jsonify({'error': '数据库连接失败'}), 500

    def _to_int_or_none(value):
        if value is None:
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT * FROM NDEPWHTStatus WHERE TestPackageID = %s",
            (test_package_id,)
        )
        row = cur.fetchone() or {}
        order = ['VT', 'RT', 'PT', 'UT', 'MT', 'PMI', 'FT', 'HT', 'PWHT']
        payload = {}
        for key in order:
            total_val = _to_int_or_none(row.get(f'{key}_Total'))
            completed_val = _to_int_or_none(row.get(f'{key}_Completed'))
            remaining = None
            if total_val is not None and completed_val is not None:
                remaining = max(total_val - completed_val, 0)
            payload[key] = {
                'total': total_val,
                'completed': completed_val,
                'remaining': remaining
            }
        return jsonify(payload)
    finally:
        conn.close()


@test_package_bp.route('/test_packages/<test_package_id>/export_package')
def export_single_test_package(test_package_id):
    normalized = unquote(test_package_id)
    conn = create_connection()
    if not conn:
        return "数据库连接失败", 500
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM HydroTestPackageList WHERE TestPackageID = %s", (normalized,))
        test_package = cur.fetchone()
        if not test_package:
            return "未找到试压包", 404
        cur.execute("SELECT * FROM SystemList WHERE SystemCode = %s", (test_package.get('SystemCode'),))
        system = cur.fetchone() or {}
        cur.execute("SELECT * FROM SubsystemList WHERE SubSystemCode = %s", (test_package.get('SubSystemCode'),))
        subsystem = cur.fetchone() or {}
        include_attachments = request.args.get('include_attachments', 'false').lower() == 'true'
        return export_test_package_from_template(test_package, system, subsystem, include_attachments)
    finally:
        conn.close()


@test_package_bp.route('/admin/refresh_master')
def refresh_master():
    stats = TestPackageModel.sync_all_from_welding()
    return f"刷新完成: {stats}"

