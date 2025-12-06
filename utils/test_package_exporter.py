"""试压包资料包导出功能 - 生成完整的试压包Excel文档"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from io import BytesIO
from flask import Response
import os

# 获取项目根目录的绝对路径
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 尺寸缓存（避免每次都启动Excel COM）
# 使用threading.Lock保证线程安全
import threading
_DIMENSIONS_CACHE = {}
_CACHE_LOCK = threading.Lock()

# Excel COM导出的并发限制（最多5个并发）
from threading import Semaphore
_EXCEL_COM_SEMAPHORE = Semaphore(5)

def copy_dimensions_from_template(target_ws, template_path, template_sheet_name, 
                                  max_col=None, max_row=None, verbose=False):
    """
    从模板工作表复制列宽和行高到目标工作表（使用Excel COM接口 + 缓存）
    
    Args:
        target_ws: 目标工作表对象（openpyxl）
        template_path: 模板文件路径（绝对路径或相对于BASE_DIR的路径）
        template_sheet_name: 模板工作表名称
        max_col: 最大列数（默认200）
        max_row: 最大行数（默认1000）
        verbose: 是否显示详细日志
    
    Returns:
        bool: 成功返回True，失败返回False
    """
    from openpyxl.utils import get_column_letter
    
    # 处理相对路径
    if not os.path.isabs(template_path):
        template_path = os.path.join(BASE_DIR, template_path)
    
    # 检查模板文件是否存在
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")
    
    template_path = os.path.abspath(template_path)
    
    # 生成缓存键
    cache_key = f"{template_path}::{template_sheet_name}::{max_col}::{max_row}"
    
    # 检查缓存（线程安全）
    with _CACHE_LOCK:
        if cache_key in _DIMENSIONS_CACHE:
            cached_data = _DIMENSIONS_CACHE[cache_key]
            
            # 应用缓存的列宽
            for col_letter, width in cached_data['col_widths'].items():
                target_ws.column_dimensions[col_letter].width = width
            
            # 应用缓存的行高
            for row_idx, height in cached_data['row_heights'].items():
                target_ws.row_dimensions[row_idx].height = height
            
            return True
    
    # 缓存未命中，使用COM读取
    import win32com.client
    import pythoncom
    
    excel = None
    wb = None
    
    try:
        # 初始化COM（多线程环境需要）
        pythoncom.CoInitialize()
        
        # 创建Excel应用实例（设置超时保护）
        import time
        start_time = time.time()
        
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
        except Exception as e:
            import pywintypes
            # 捕获COM错误并提供详细的错误信息
            if isinstance(e, pywintypes.com_error):
                error_code = e.args[0] if e.args else None
                if error_code == -2147221005:
                    error_msg = (
                        "Excel COM 组件初始化失败 (错误代码: -2147221005)。\n"
                        "可能的原因：\n"
                        "1. Excel 未安装在服务器上\n"
                        "2. COM 组件注册损坏（运行: excel.exe /regserver）\n"
                        "3. 权限问题或 DCOM 配置问题"
                    )
                    print(f"[ERROR] {error_msg}")
                else:
                    print(f"[ERROR] Excel COM 错误：{str(e)}")
            else:
                print(f"[ERROR] 无法启动 Excel 应用程序：{str(e)}")
            raise
        
        # 打开模板文件（只读模式，提高速度）
        wb = excel.Workbooks.Open(template_path, ReadOnly=True, UpdateLinks=False)
        
        elapsed = time.time() - start_time
        if elapsed > 5:  # 超过5秒警告
            print(f"[WARNING] Excel COM启动耗时: {elapsed:.1f}秒")
        
        # 查找工作表
        ws = None
        for sheet in wb.Sheets:
            if sheet.Name == template_sheet_name:
                ws = sheet
                break
        
        if ws is None:
            raise ValueError(f"模板中不存在工作表: {template_sheet_name}")
        
        # 设置默认值
        if max_col is None:
            max_col = 200
        if max_row is None:
            max_row = 1000
        
        # 读取列宽（使用ColumnWidth属性，与VBA一致）
        col_widths = {}
        for col_idx in range(1, max_col + 1):
            try:
                col_width = ws.Cells(1, col_idx).EntireColumn.ColumnWidth
                if col_width and col_width > 0:
                    col_letter = get_column_letter(col_idx)
                    col_widths[col_letter] = col_width
                    if verbose:
                        print(f"  列{col_letter}(#{col_idx}): ColumnWidth={col_width:.2f}")
            except:
                break  # 超出范围
        
        # 读取行高（使用RowHeight属性，与VBA一致）
        row_heights = {}
        for row_idx in range(1, max_row + 1):
            try:
                row_height = ws.Cells(row_idx, 1).EntireRow.RowHeight
                if row_height and row_height > 0:
                    row_heights[row_idx] = row_height
                    if verbose:
                        print(f"  行{row_idx}: RowHeight={row_height:.2f}")
            except:
                break  # 超出范围
        
        # 应用到目标工作表
        for col_letter, width in col_widths.items():
            target_ws.column_dimensions[col_letter].width = width
        for row_idx, height in row_heights.items():
            target_ws.row_dimensions[row_idx].height = height
        
        # 存入缓存（线程安全）
        with _CACHE_LOCK:
            _DIMENSIONS_CACHE[cache_key] = {
                'col_widths': col_widths,
                'row_heights': row_heights
            }
        
        return True
        
    except Exception as e:
        print(f"[ERROR] 从模板复制尺寸失败: {e}")
        import traceback
        traceback.print_exc()
        raise  # 重新抛出异常，不使用降级方案
        
    finally:
        # 清理
        if wb:
            wb.Close(SaveChanges=False)
        if excel:
            excel.Quit()
        # 释放COM
        pythoncom.CoUninitialize()

# 管道材质中英俄三语映射
PIPE_MATERIAL_MAPPING = {
    'PE': {
        'en': 'PE',
        'ru': 'Полиэтилен',
        'cn': 'PE'
    },
    'CS': {
        'en': 'CS',
        'ru': 'Углеродистая сталь',
        'cn': '碳钢'
    },
    'SS': {
        'en': 'SS',
        'ru': 'Нержавеющая сталь',
        'cn': '不锈钢'
    },
    'DUSS': {
        'en': 'DUSS',
        'ru': 'Дуплекс из нержавеющей стали',
        'cn': '双相不锈钢'
    },
    'LACS': {
        'en': 'LACS',
        'ru': 'Легированная сталь',
        'cn': '合金钢'
    },
    'LTCS': {
        'en': 'LTCS',
        'ru': 'Низкотемпературная углеродистая сталь',
        'cn': '低温碳钢'
    },
    'NIAS': {
        'en': 'NIAS',
        'ru': 'Никелевый сплав',
        'cn': '镍合金'
    }
}

# 测试类型映射
TEST_TYPE_MAPPING = {
    '气压': {
        'en': 'Pneumatic test',
        'ru': 'Пневматическое испытание'
    },
    '水压': {
        'en': 'Hydraulic test',
        'ru': 'Гидравлическое испытание'
    },
    '观察包': {
        'en': 'Observation package',
        'ru': 'Пакет наблюдения'
    }
}

# 测试介质映射
TEST_MEDIUM_MAPPING = {
    '水': {'en': 'Water', 'ru': 'Вода'},
    '空气': {'en': 'Air', 'ru': 'Воздух'},
    'Air': {'en': 'Air', 'ru': 'Воздух'},
    'Water': {'en': 'Water', 'ru': 'Вода'},
}


def create_cover_page(wb, test_package_data, system_data, subsystem_data):
    """创建Cover Page工作表 - 严格按照模板布局（A-R列，1-32行）"""
    ws = wb.create_sheet("1.Cover", 0)
    
    # 从模板复制列宽和行高（使用Excel COM接口，与VBA完全一致）
    template_path = os.path.join('nordinfo', '试压包模板(202511).xlsx')  # 相对于BASE_DIR
    try:
        copy_dimensions_from_template(
            ws, 
            template_path, 
            '1.Cover',
            max_col=18,    # A-R列
            max_row=32,    # 1-32行
            verbose=False  # 不显示详细日志
        )
    except Exception as e:
        # 如果COM失败，使用备用尺寸（避免整个导出失败）
        print(f"[WARNING] 模板读取失败，使用备用尺寸: {e}")
        # 使用VBA提供的实际值作为备用
        column_widths = {
            'A': 5.05, 'B': 5.05, 'C': 5.05, 'D': 5.05, 'E': 5.05, 'F': 5.05,
            'G': 3.53, 'H': 8.16, 'I': 5.05, 'J': 5.05, 'K': 5.05, 'L': 5.05,
            'M': 5.05, 'N': 5.05, 'O': 9.58, 'P': 5.05, 'Q': 5.05, 'R': 8.16
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        row_heights = {
            1: 28.0, 2: 23.5, 3: 41.1, 4: 28.0, 5: 28.0, 6: 28.0, 7: 28.0,
            8: 30.0, 9: 28.0, 10: 28.0, 11: 28.0, 12: 28.0, 13: 28.0, 14: 28.0,
            15: 28.0, 16: 28.0, 17: 28.0, 18: 30.0, 19: 30.0, 20: 30.0, 21: 30.0,
            22: 30.0, 23: 30.0, 24: 30.0, 25: 30.0, 26: 30.0, 27: 28.0, 28: 28.0,
            29: 28.0, 30: 28.0, 31: 28.0, 32: 28.0
        }
        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height
    
    # 样式定义（完全按照模板）
    title_font = Font(name='Times New Roman', size=18, bold=True)
    project_font = Font(name='Times New Roman', size=10, bold=True)
    package_label_font = Font(name='Times New Roman', size=24, bold=True)
    package_label_ru_font = Font(name='Times New Roman', size=20, bold=True, color='FF0000FF')  # 蓝色
    package_id_font = Font(name='Times New Roman', size=24, bold=True)
    
    data_label_font = Font(name='Times New Roman', size=12, bold=True)
    data_value_font = Font(name='Times New Roman', size=12, bold=True)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 边框样式
    medium_border = Side(style='medium', color='000000')
    thin_border = Side(style='thin', color='000000')
    
    # Logo区域（A1:C3 左侧, P1:R3 右侧）
    ws.merge_cells('A1:C3')
    ws['A1'].alignment = center_align
    
    ws.merge_cells('P1:R3')
    ws['P1'].alignment = center_align
    
    # 加载logo图片（GCC在左侧A1，BCC在右侧P1）
    # 使用绝对路径确保在Flask应用中也能找到图片
    bcc_logo_path = os.path.join(BASE_DIR, 'static', 'images', 'bcc_logo.png')
    gcc_logo_path = os.path.join(BASE_DIR, 'static', 'images', 'gcc_logo.png')
    
    
    # 添加GCC Logo（左侧）
    try:
        if os.path.exists(gcc_logo_path):
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU, EMU_to_pixels
            
            img_gcc = XLImage(gcc_logo_path)
            # 调整大小
            img_gcc.width = 100   # pixels
            img_gcc.height = 70   # pixels
            
            # 创建自定义锚点：A1单元格（col=0, row=0），往下偏移10像素
            marker = AnchorMarker(col=0, colOff=0, row=0, rowOff=pixels_to_EMU(10))
            size = XDRPositiveSize2D(cx=pixels_to_EMU(100), cy=pixels_to_EMU(70))
            img_gcc.anchor = OneCellAnchor(_from=marker, ext=size)
            
            ws._images.append(img_gcc)
    except Exception as e:
        print(f"[ERROR] 加载GCC Logo失败: {e}")
        import traceback
        traceback.print_exc()
    
    # 添加BCC Logo（右侧）
    try:
        if os.path.exists(bcc_logo_path):
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            
            img_bcc = XLImage(bcc_logo_path)
            # 调整大小
            img_bcc.width = 100   # pixels
            img_bcc.height = 70   # pixels
            
            # 创建自定义锚点：P1单元格（col=15, row=0），往右偏移5像素，往下偏移10像素
            marker = AnchorMarker(col=15, colOff=pixels_to_EMU(5), row=0, rowOff=pixels_to_EMU(10))
            size = XDRPositiveSize2D(cx=pixels_to_EMU(100), cy=pixels_to_EMU(70))
            img_bcc.anchor = OneCellAnchor(_from=marker, ext=size)
            
            ws._images.append(img_bcc)
    except Exception as e:
        print(f"[ERROR] 加载BCC Logo失败: {e}")
        import traceback
        traceback.print_exc()
    
    # 标题区域（D1:O2，合并）- Pressure Test Package + 俄文（俄文用蓝色）
    ws.merge_cells('D1:O2')
    # 使用RichText设置不同颜色
    title_black = InlineFont(rFont='Times New Roman', sz=18.0, b=True, color='000000')
    title_blue = InlineFont(rFont='Times New Roman', sz=18.0, b=True, color='FF0000FF')
    ws['D1'].value = CellRichText(
        TextBlock(title_black, 'Pressure Test Package\n'),
        TextBlock(title_blue, 'Пакет документации по испытаниям')
    )
    ws['D1'].alignment = center_wrap
    
    # 项目名称（D3:O3）- 英文黑色，俄文蓝色
    ws.merge_cells('D3:O3')
    project_black = InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000')
    project_blue = InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF')
    ws['D3'].value = CellRichText(
        TextBlock(project_black, 'Gas Chemical Complex within Ethane-rich Gas Processing Complex\n'),
        TextBlock(project_blue, 'Газохимический комплекс в составе Комплекса переработки этансодержащего газа')
    )
    ws['D3'].alignment = center_wrap
    
    # 空行（4-7）
    for row in [4, 5, 6, 7]:
        ws.merge_cells(f'A{row}:R{row}')
    
    # Test Package No. 标签（A8:R8）
    ws.merge_cells('A8:R8')
    ws['A8'].value = 'Test Package No.'
    ws['A8'].font = package_label_font
    ws['A8'].alignment = center_align
    
    # 俄文标签（A9:R9）
    ws.merge_cells('A9:R9')
    ws['A9'].value = 'Пакет документации по испытаниям №'
    ws['A9'].font = package_label_ru_font
    ws['A9'].alignment = center_align
    
    # 空行（A10:R10）
    ws.merge_cells('A10:R10')
    
    # 试压包编号（A11:R11）
    ws.merge_cells('A11:R11')
    ws['A11'].value = test_package_data.get('TestPackageID', '')
    ws['A11'].font = package_id_font
    ws['A11'].alignment = center_align
    
    # 空行（12-17）
    for row in [12, 13, 14, 15, 16, 17]:
        ws.merge_cells(f'A{row}:R{row}')
    
    # 获取数据映射
    pipe_material = test_package_data.get('PipeMaterial', '')
    material_info = PIPE_MATERIAL_MAPPING.get(pipe_material, {'en': pipe_material or '', 'ru': '', 'cn': pipe_material or ''})
    
    test_type = test_package_data.get('TestType', '')
    test_type_info = TEST_TYPE_MAPPING.get(test_type, {'en': test_type or '', 'ru': ''})
    
    test_medium = test_package_data.get('TestMedium', '')
    medium_info = TEST_MEDIUM_MAPPING.get(test_medium, {'en': test_medium or '', 'ru': ''})
    
    # 详细信息区域（从行18开始，按照模板格式）
    # 每行：标签（A-G列），值（H-R列）
    
    # 行18: System No. - 英文黑色，俄文蓝色
    ws.merge_cells('A18:G18')
    label_black = InlineFont(rFont='Times New Roman', sz=12.0, b=True, color='000000')
    label_blue = InlineFont(rFont='Times New Roman', sz=12.0, b=True, color='FF0000FF')
    ws['A18'].value = CellRichText(
        TextBlock(label_black, 'System No.:\n'),
        TextBlock(label_blue, 'Система №.:')
    )
    ws['A18'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    
    ws.merge_cells('H18:R18')
    ws['H18'].value = system_data.get('SystemCode', 'N/A')
    ws['H18'].font = data_value_font
    ws['H18'].alignment = left_align
    
    # 行19: Sub-System No. - 英文黑色，俄文蓝色
    ws.merge_cells('A19:G19')
    ws['A19'].value = CellRichText(
        TextBlock(label_black, 'Sub-System No.:\n'),
        TextBlock(label_blue, 'Подсистема №.:')
    )
    ws['A19'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    
    ws.merge_cells('H19:R19')
    ws['H19'].value = subsystem_data.get('SubSystemCode', 'N/A')
    ws['H19'].font = data_value_font
    ws['H19'].alignment = left_align
    
    # 行20: System / Sub-system Description - 英文黑色，俄文蓝色
    ws.merge_cells('A20:G20')
    ws['A20'].value = CellRichText(
        TextBlock(label_black, 'System / Sub-system Description:\n'),
        TextBlock(label_blue, 'Система / описание подсистемы:')
    )
    ws['A20'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    
    ws.merge_cells('H20:R20')
    desc = subsystem_data.get('SubSystemDescriptionENG', 'N/A')
    desc_ru = subsystem_data.get('SubSystemDescriptionRUS', '')
    value_black = InlineFont(rFont='Times New Roman', sz=12.0, b=True, color='000000')
    value_blue = InlineFont(rFont='Times New Roman', sz=12.0, b=True, color='FF0000FF')
    if desc_ru:
        ws['H20'].value = CellRichText(
            TextBlock(value_black, f"{desc}\n"),
            TextBlock(value_blue, desc_ru)
        )
    else:
        ws['H20'].value = desc
        ws['H20'].font = data_value_font
    ws['H20'].alignment = left_align
    
    # 行21: Piping Material - 英文黑色，俄文蓝色
    ws.merge_cells('A21:G21')
    ws['A21'].value = CellRichText(
        TextBlock(label_black, 'Piping Material:\n'),
        TextBlock(label_blue, 'Материал трубы:')
    )
    ws['A21'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    
    ws.merge_cells('H21:R21')
    if material_info['ru']:
        ws['H21'].value = CellRichText(
            TextBlock(value_black, f"{material_info['en']}\n"),
            TextBlock(value_blue, material_info['ru'])
        )
    else:
        ws['H21'].value = material_info['en'] or 'N/A'
        ws['H21'].font = data_value_font
    ws['H21'].alignment = left_align
    
    # 行22: Piping class - 英文黑色，俄文蓝色
    ws.merge_cells('A22:G22')
    ws['A22'].value = CellRichText(
        TextBlock(label_black, 'Piping class /\n'),
        TextBlock(label_blue, 'Класс трубопровода')
    )
    ws['A22'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    
    ws.merge_cells('H22:R22')
    ws['H22'].value = test_package_data.get('PipingClass', '')  # 这个字段可能需要添加
    ws['H22'].font = data_value_font
    ws['H22'].alignment = left_align
    
    # 行23: Test Type - 英文黑色，俄文蓝色
    ws.merge_cells('A23:G23')
    ws['A23'].value = CellRichText(
        TextBlock(label_black, 'Test Type:\n'),
        TextBlock(label_blue, 'Тип Испытания:')
    )
    ws['A23'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    
    ws.merge_cells('H23:R23')
    if test_type_info.get('ru'):
        ws['H23'].value = CellRichText(
            TextBlock(value_black, f"{test_type_info['en']}\n"),
            TextBlock(value_blue, test_type_info['ru'])
        )
    else:
        ws['H23'].value = test_type_info['en'] or 'N/A'
        ws['H23'].font = data_value_font
    ws['H23'].alignment = left_align
    
    # 行24: Test Medium - 英文黑色，俄文蓝色
    ws.merge_cells('A24:G24')
    ws['A24'].value = CellRichText(
        TextBlock(label_black, 'Test Medium:\n'),
        TextBlock(label_blue, 'Среда Испытания:')
    )
    ws['A24'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    
    ws.merge_cells('H24:R24')
    if medium_info.get('ru'):
        ws['H24'].value = CellRichText(
            TextBlock(value_black, f"{medium_info['en']}\n"),
            TextBlock(value_blue, medium_info['ru'])
        )
    else:
        ws['H24'].value = medium_info['en'] or 'N/A'
        ws['H24'].font = data_value_font
    ws['H24'].alignment = left_align
    
    # 行25: Design Pressure - 英文黑色，俄文蓝色
    ws.merge_cells('A25:G25')
    ws['A25'].value = CellRichText(
        TextBlock(label_black, 'Design Pressure:\n'),
        TextBlock(label_blue, 'Расчетное давление:')
    )
    ws['A25'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    
    # H25是压力值
    ws['H25'].value = test_package_data.get('DesignPressure', '')
    ws['H25'].font = data_value_font
    ws['H25'].alignment = Alignment(horizontal='center', vertical='center')
    
    # I25:R25是单位
    ws.merge_cells('I25:R25')
    ws['I25'].value = 'Mpa'
    ws['I25'].font = data_value_font
    ws['I25'].alignment = left_align
    
    # 行26: Test Pressure - 英文黑色，俄文蓝色
    ws.merge_cells('A26:G26')
    ws['A26'].value = CellRichText(
        TextBlock(label_black, 'Test Pressure:\n'),
        TextBlock(label_blue, 'Испытательное давление:')
    )
    ws['A26'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    
    # H26是压力值
    ws['H26'].value = test_package_data.get('TestPressure', '')
    ws['H26'].font = data_value_font
    ws['H26'].alignment = Alignment(horizontal='center', vertical='center')
    
    # I26:R26是单位
    ws.merge_cells('I26:R26')
    ws['I26'].value = 'Mpa'
    ws['I26'].font = data_value_font
    ws['I26'].alignment = left_align
    
    # 剩余空行（27-32）
    for row in range(27, 33):
        ws.merge_cells(f'A{row}:I{row}')
        ws.merge_cells(f'J{row}:R{row}')
    
    # 添加边框（根据模板详细扫描结果，完全复制边框模式）
    # 清空现有边框
    for row in range(1, 33):
        for col_idx in range(1, 19):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border()
    
    # === 第1行边框 ===
    # A1: L,R,T,B (medium)
    ws['A1'].border = Border(left=medium_border, right=medium_border, top=medium_border, bottom=medium_border)
    # B1-N1: T (medium)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws[f'{col}1'].border = Border(top=medium_border)
    # C1: 额外加 R (medium)
    ws['C1'].border = Border(right=medium_border, top=medium_border)
    # D1: L,R,T (medium)
    ws['D1'].border = Border(left=medium_border, right=medium_border, top=medium_border)
    # O1: R,T (medium)
    ws['O1'].border = Border(right=medium_border, top=medium_border)
    # P1: L,R,T,B (medium)
    ws['P1'].border = Border(left=medium_border, right=medium_border, top=medium_border, bottom=medium_border)
    # Q1: T (medium)
    ws['Q1'].border = Border(top=medium_border)
    # R1: R,T (medium)
    ws['R1'].border = Border(right=medium_border, top=medium_border)
    
    # === 第2行边框 ===
    ws['A2'].border = Border(left=medium_border)
    ws['C2'].border = Border(right=medium_border)
    ws['D2'].border = Border(left=medium_border)
    ws['O2'].border = Border(right=medium_border)
    ws['P2'].border = Border(left=medium_border)
    ws['R2'].border = Border(right=medium_border)
    
    # === 第3行边框 ===
    # A3-R3全部有底部边框
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        ws[f'{col}3'].border = Border(bottom=medium_border)
    # 额外的左右边框
    ws['A3'].border = Border(left=medium_border, bottom=medium_border)
    ws['C3'].border = Border(right=medium_border, bottom=medium_border)
    ws['D3'].border = Border(left=medium_border, right=medium_border, bottom=medium_border)
    ws['O3'].border = Border(right=medium_border, bottom=medium_border)
    ws['P3'].border = Border(left=medium_border, bottom=medium_border)
    ws['R3'].border = Border(right=medium_border, bottom=medium_border)
    
    # === 第4行边框（顶部边框 + 左右边框）===
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        ws[f'{col}4'].border = Border(top=medium_border)
    # A4特殊：L,R,T
    ws['A4'].border = Border(left=medium_border, right=medium_border, top=medium_border)
    # R4: R,T
    ws['R4'].border = Border(right=medium_border, top=medium_border)
    
    # === 第5-17行边框（仅左右边框）===
    for row in range(5, 18):
        ws[f'A{row}'].border = Border(left=medium_border, right=medium_border)
        ws[f'R{row}'].border = Border(right=medium_border)
    
    # === 第18-24行边框（数据区域，标签列和值列分隔）===
    for row in range(18, 25):
        ws[f'A{row}'].border = Border(left=medium_border)
        ws[f'H{row}'].border = Border(right=medium_border)
        ws[f'R{row}'].border = Border(right=medium_border)
    
    # === 第25-26行边框（压力值列分隔）===
    for row in range(25, 27):
        ws[f'A{row}'].border = Border(left=medium_border)
        ws[f'I{row}'].border = Border(right=medium_border)
        ws[f'R{row}'].border = Border(right=medium_border)
    
    # === 第27-31行边框（空白区域，两列分隔）===
    for row in range(27, 32):
        ws[f'A{row}'].border = Border(left=medium_border)
        ws[f'J{row}'].border = Border(right=medium_border)
        ws[f'R{row}'].border = Border(right=medium_border)
    
    # === 第32行边框（底部边框 + 分隔线）===
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        ws[f'{col}32'].border = Border(bottom=medium_border)
    # 额外的左右边框
    ws['A32'].border = Border(left=medium_border, bottom=medium_border)
    ws['J32'].border = Border(right=medium_border, bottom=medium_border)
    ws['R32'].border = Border(right=medium_border, bottom=medium_border)
    
    # 隐藏网格线
    ws.sheet_view.showGridLines = False
    
    return ws


def check_checklist_status(test_package_id):
    """
    检查试压包Checklist各项的完成状态
    
    Returns:
        dict: 各项的完成状态 {item_no: True/False}
    """
    from database import create_connection
    
    conn = create_connection()
    if not conn:
        return {}
    
    try:
        cur = conn.cursor(dictionary=True)
        status = {}
        
        # 1.0 - Pressure Test Package Content（总是Yes）
        status['1.0'] = True
        
        # 2.0 - P&IDs/ISO Drawing List
        cur.execute("SELECT COUNT(*) as cnt FROM PIDList WHERE TestPackageID = %s", (test_package_id,))
        pid_count = cur.fetchone()['cnt']
        
        # 优先从ISODrawingList读取，如果为空则从WeldingList提取
        cur.execute("SELECT COUNT(*) as cnt FROM ISODrawingList WHERE TestPackageID = %s", (test_package_id,))
        iso_count = cur.fetchone()['cnt']
        if iso_count == 0:
            cur.execute("""
                SELECT COUNT(DISTINCT DrawingNumber) as cnt 
                FROM WeldingList 
                WHERE TestPackageID = %s 
                AND DrawingNumber IS NOT NULL 
                AND DrawingNumber != ''
                AND (DrawingNumber LIKE '%ISO%' OR DrawingNumber LIKE '%IS0%')
            """, (test_package_id,))
            iso_count = cur.fetchone()['cnt']
        
        status['2.0'] = (pid_count > 0 or iso_count > 0)
        
        # 3.0-5.0, 8.0-14.0 - 各种附件
        attachment_modules = {
            '3.0': 'PID_Drawings',
            '4.0': 'ISO_Drawings',
            '5.0': 'Symbols_Legend',
            '8.0': 'Test_Flow_Chart',
            '9.0': 'Test_Check_List',
            '10.0': 'Calibration_Certificates',
            '11.0': 'Test_Certificate',
            '12.0': 'Flushing_Certificate',
            '13.0': 'Reinstatement_Check_List',
            '14.0': 'Others'
        }
        
        for item_no, module_name in attachment_modules.items():
            cur.execute("""
                SELECT COUNT(*) as cnt 
                FROM TestPackageAttachments 
                WHERE TestPackageID = %s AND ModuleName = %s
            """, (test_package_id, module_name))
            count = cur.fetchone()['cnt']
            status[item_no] = (count > 0)
        
        # 6.0 - Extract from NORD PCMS (检查是否有焊口数据)
        cur.execute("""
            SELECT COUNT(*) as cnt 
            FROM WeldingList 
            WHERE TestPackageID = %s
        """, (test_package_id,))
        weld_count = cur.fetchone()['cnt']
        status['6.0'] = (weld_count > 0)
        
        # 7.0 - Punch List
        cur.execute("SELECT COUNT(*) as cnt FROM PunchList WHERE TestPackageID = %s", (test_package_id,))
        punch_count = cur.fetchone()['cnt']
        status['7.0'] = (punch_count > 0)
        
        return status
        
    except Exception as e:
        print(f"[ERROR] 检查Checklist状态失败: {e}")
        return {}
    finally:
        conn.close()


def create_contents_page(wb, test_package_data, system_data, subsystem_data, checklist_status):
    """创建2.Contents工作表 - 试压包内容清单"""
    ws = wb.create_sheet("2.Contents", 1)
    
    # 设置列宽（使用VBA读取的准确值）
    column_widths = {
        'A': 5.05, 'B': 5.05, 'C': 5.05, 'D': 5.05, 'E': 5.05, 'F': 5.05,
        'G': 5.05, 'H': 5.05, 'I': 5.05, 'J': 5.05, 'K': 5.05, 'L': 12.68,
        'M': 0.0, 'N': 12.42, 'O': 5.05, 'P': 5.05, 'Q': 5.05, 'R': 5.05
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 设置行高（使用VBA读取的准确值）
    row_heights = {
        1: 28.0, 2: 28.0, 3: 28.0, 4: 28.0, 5: 58.5, 6: 28.0, 7: 28.0,
        8: 28.0, 9: 28.0, 10: 28.0, 11: 28.0, 12: 28.0, 13: 30.6, 14: 30.6,
        15: 30.6, 16: 30.6, 17: 30.6, 18: 43.5, 19: 30.6, 20: 30.6,
        21: 30.6, 22: 30.6, 23: 30.6, 24: 30.6, 25: 30.6, 26: 30.6,
        27: 28.0, 28: 28.0, 29: 28.0
    }
    for row, height in row_heights.items():
        ws.row_dimensions[row].height = height
    
    # 样式定义
    title_font = Font(name='Times New Roman', size=18, bold=True)
    label_font = Font(name='Times New Roman', size=10, bold=True)
    value_font = Font(name='Times New Roman', size=10, bold=False)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # 浅蓝色
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    medium_border = Side(style='medium', color='000000')
    
    # === 第一部分：标题区域（A1:R3）===
    # Logo区域已合并
    ws.merge_cells('A1:C3')
    ws.merge_cells('P1:R3')
    
    # 添加Logo（与Cover Page相同）
    bcc_logo_path = os.path.join(BASE_DIR, 'static', 'images', 'bcc_logo.png')
    gcc_logo_path = os.path.join(BASE_DIR, 'static', 'images', 'gcc_logo.png')
    
    try:
        if os.path.exists(gcc_logo_path):
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            
            img_gcc = XLImage(gcc_logo_path)
            img_gcc.width = 100
            img_gcc.height = 70
            marker = AnchorMarker(col=0, colOff=0, row=0, rowOff=pixels_to_EMU(10))
            size = XDRPositiveSize2D(cx=pixels_to_EMU(100), cy=pixels_to_EMU(70))
            img_gcc.anchor = OneCellAnchor(_from=marker, ext=size)
            ws._images.append(img_gcc)
    except:
        pass
    
    try:
        if os.path.exists(bcc_logo_path):
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            
            img_bcc = XLImage(bcc_logo_path)
            img_bcc.width = 100
            img_bcc.height = 70
            marker = AnchorMarker(col=15, colOff=pixels_to_EMU(5), row=0, rowOff=pixels_to_EMU(10))
            size = XDRPositiveSize2D(cx=pixels_to_EMU(100), cy=pixels_to_EMU(70))
            img_bcc.anchor = OneCellAnchor(_from=marker, ext=size)
            ws._images.append(img_bcc)
    except:
        pass
    
    # 标题文字（D1:O2）
    ws.merge_cells('D1:O2')
    title_black = InlineFont(rFont='Times New Roman', sz=18.0, b=True, color='000000')
    title_blue = InlineFont(rFont='Times New Roman', sz=18.0, b=True, color='FF0000FF')
    ws['D1'].value = CellRichText(
        TextBlock(title_black, 'Pressure Test Package Content\n'),
        TextBlock(title_blue, 'Содержание тест-пакета')
    )
    ws['D1'].alignment = center_align
    
    # 项目名称（D3:O3）
    ws.merge_cells('D3:O3')
    project_black = InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000')
    project_blue = InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF')
    ws['D3'].value = CellRichText(
        TextBlock(project_black, 'Gas Chemical Complex within Ethane-rich Gas Processing Complex\n'),
        TextBlock(project_blue, 'Газохимический комплекс в составе Комплекса переработки этансодержащего газа')
    )
    ws['D3'].alignment = center_align
    
    # === 第二部分：基础信息（A4:R10）===
    # 第4行：System No. / Sub-System No.
    ws.merge_cells('A4:C4')
    ws['A4'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 'System No.:\n'),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'), 'Система №.:')
    )
    ws['A4'].alignment = left_align
    
    ws.merge_cells('E4:I4')
    ws['E4'].value = system_data.get('SystemCode', 'N/A')
    ws['E4'].font = value_font
    ws['E4'].alignment = left_align
    
    ws.merge_cells('J4:M4')
    ws['J4'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 'Sub-System No.:\n'),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'), 'Подсистема №.:')
    )
    ws['J4'].alignment = left_align
    
    ws.merge_cells('N4:R4')
    ws['N4'].value = subsystem_data.get('SubSystemCode', 'N/A')
    ws['N4'].font = value_font
    ws['N4'].alignment = left_align
    
    # 第5行：Test Package No. / System Description
    ws.merge_cells('A5:C5')
    ws['A5'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 'Test Package No.\n'),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'), 'Тест-пакет №')
    )
    ws['A5'].alignment = left_align
    
    ws.merge_cells('E5:I5')
    ws['E5'].value = test_package_data.get('TestPackageID', 'N/A')
    ws['E5'].font = value_font
    ws['E5'].alignment = left_align
    
    ws.merge_cells('J5:M5')
    ws['J5'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 'System Description:\n'),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'), 'Система описания:')
    )
    ws['J5'].alignment = left_align
    
    ws.merge_cells('N5:R5')
    system_desc_en = system_data.get('SystemDescriptionENG', 'N/A')
    system_desc_ru = system_data.get('SystemDescriptionRUS', '')
    if system_desc_ru:
        ws['N5'].value = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='000000'), f"{system_desc_en}\n"),
            TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='FF0000FF'), system_desc_ru)
        )
    else:
        ws['N5'].value = system_desc_en
        ws['N5'].font = value_font
    ws['N5'].alignment = left_align
    
    # 第6行：Design Pressure / Test Pressure
    ws.merge_cells('A6:C6')
    ws['A6'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 'Design Pressure:\n'),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'), 'Расчетное давление:')
    )
    ws['A6'].alignment = left_align
    
    ws.merge_cells('E6:I6')
    design_pressure = test_package_data.get('DesignPressure', '')
    ws['E6'].value = f"{design_pressure} Mpa" if design_pressure else ''
    ws['E6'].font = value_font
    ws['E6'].alignment = left_align
    
    ws.merge_cells('J6:M6')
    ws['J6'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 'Test Pressure:\n'),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'), 'Испытательное давление:')
    )
    ws['J6'].alignment = left_align
    
    ws.merge_cells('N6:R6')
    test_pressure = test_package_data.get('TestPressure', '')
    ws['N6'].value = f"{test_pressure} Mpa" if test_pressure else ''
    ws['N6'].font = value_font
    ws['N6'].alignment = left_align
    
    # 第7行：Test Type / Test Medium
    ws.merge_cells('A7:C7')
    ws['A7'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 'Test Type:\n'),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'), 'Тип испытания:')
    )
    ws['A7'].alignment = left_align
    
    ws.merge_cells('E7:I7')
    test_type = test_package_data.get('TestType', '')
    test_type_info = TEST_TYPE_MAPPING.get(test_type, {'en': test_type, 'ru': ''})
    if test_type_info['ru']:
        ws['E7'].value = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='000000'), f"{test_type_info['en']}\n"),
            TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='FF0000FF'), test_type_info['ru'])
        )
    else:
        ws['E7'].value = test_type_info['en']
        ws['E7'].font = value_font
    ws['E7'].alignment = left_align
    
    ws.merge_cells('J7:M7')
    ws['J7'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 'Test Medium:\n'),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'), 'Среда испытания:')
    )
    ws['J7'].alignment = left_align
    
    ws.merge_cells('N7:R7')
    test_medium = test_package_data.get('TestMedium', '')
    test_medium_info = TEST_MEDIUM_MAPPING.get(test_medium, {'en': test_medium, 'ru': ''})
    if test_medium_info['ru']:
        ws['N7'].value = CellRichText(
            TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='000000'), f"{test_medium_info['en']}\n"),
            TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='FF0000FF'), test_medium_info['ru'])
        )
    else:
        ws['N7'].value = test_medium_info['en']
        ws['N7'].font = value_font
    ws['N7'].alignment = left_align
    
    # 第8-10行：认证声明（固定文本，合并A8:R10）
    ws.merge_cells('A8:R10')
    ws['A8'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 
                  'Certified that the above piping Test Pack has been completed with respect to fabrication, welding & inspection (Non-Destructive Testing), line checking, isolation of required equipment/instrument and is ready for Pressure Testing.\n'),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'),
                  'Подтверждается, что вышеуказанный тест-пакет трубопровода скомплектован в отношении изготовления, сварки и контроля (неразрушающего контроля), проверки линии, изоляции необходимого оборудования / прибора и готовностью к испытанию под давлением.')
    )
    ws['A8'].alignment = left_align
    
    # === 第三部分：Checklist（A11:R29）===
    # 第11行：Checklist标题
    ws.merge_cells('A11:R11')
    ws['A11'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 
                  'Following Documents are enclosed and verified. / '),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'),
                  'Следующие документы приложены и проверены.')
    )
    ws['A11'].alignment = left_align
    ws['A11'].fill = header_fill
    
    # 第12行：列标题
    ws['A12'].value = 'No.'
    ws['A12'].font = label_font
    ws['A12'].alignment = center_align
    ws['A12'].fill = header_fill
    
    ws.merge_cells('B12:L12')
    ws['B12'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 'Description / '),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'), 'Описание')
    )
    ws['B12'].alignment = center_align
    ws['B12'].fill = header_fill
    
    ws.merge_cells('M12:N12')
    ws['M12'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 'Required\n'),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'), 'Требуется')
    )
    ws['M12'].alignment = center_align
    ws['M12'].fill = header_fill
    
    ws.merge_cells('O12:R12')
    ws['O12'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 'Remarks\n'),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'), 'Примечания')
    )
    ws['O12'].alignment = center_align
    ws['O12'].fill = header_fill
    
    # 第13-26行：14个Checklist项目
    checklist_items = [
        ('1.0', 'Pressure Test Package Content /\nСодержание комплекта для проведения испытания под давлением'),
        ('2.0', 'Piping & Instrument Diagram (P&IDs)/ISO Drawing List\nСхемы трубной обвязки и КИПиА (P&IDs)/Перечень изометрических чертежей'),
        ('3.0', 'Piping & Instrument Diagram (P&IDs) /\nСхемы трубной обвязки и КИПиА (P&IDs)'),
        ('4.0', 'Piping Hydro Test Isometric Drawing\nИзометрические чертежи для гидроиспытания'),
        ('5.0', 'Test Package Symbols Legend\nУсловные обозначения, принятые в тест-пакете'),
        ('6.0', 'Extract from NORD PCMS SOFTWARE (welding log for work scope performed)\nВыгрузка из программы ПО NORD PCMS (журнал сварки на предоставляемый объем работ)'),
        ('7.0', 'Punch List\nПеречная ведомость'),
        ('8.0', 'Test Flow Chart /\nСхема проведения испытания'),
        ('9.0', 'Pressure Test Check List\nЧек-лист для проведения испытания'),
        ('10.0', 'Pressure Test Gauges Calibration Certificates\nСвидетельство о поверке измерительных приборов для испытания'),
        ('11.0', 'Pipeline Test Certificate /\nАкт испытания трубопровода'),
        ('12.0', 'Pipeline Flushing (Purging) Certificate /\nАкт выполнения промывки (продувки) трубопровода'),
        ('13.0', 'Reinstatement Check List\nЧек-лист на обратную сборку'),
        ('14.0', 'Others\nДругие'),
    ]
    
    for idx, (item_no, description) in enumerate(checklist_items):
        row = 13 + idx  # 从第13行开始
        
        # No.列（文本格式，去掉小数点：1.0 -> 1, 10.0 -> 10）
        item_no_display = item_no[:-2] if item_no.endswith('.0') else item_no
        ws[f'A{row}'].value = item_no_display
        ws[f'A{row}'].font = value_font
        ws[f'A{row}'].alignment = center_align
        ws[f'A{row}'].number_format = '@'  # 文本格式
        
        # Description列（B-L合并）- 使用RichText处理英文/俄文
        ws.merge_cells(f'B{row}:L{row}')
        
        # 分离英文和俄文部分
        if '\n' in description:
            parts = description.split('\n', 1)  # 只分割一次
            english_part = parts[0]
            russian_part = parts[1] if len(parts) > 1 else ''
            
            # 使用RichText：英文黑色，俄文蓝色
            ws[f'B{row}'].value = CellRichText(
                TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='000000'), f"{english_part}\n"),
                TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='FF0000FF'), russian_part)
            )
        else:
            ws[f'B{row}'].value = description
            ws[f'B{row}'].font = value_font
        
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
        
        # Required列（M-N合并，显示Yes/No复选框）
        ws.merge_cells(f'M{row}:N{row}')
        is_completed = checklist_status.get(item_no, False)
        if is_completed:
            ws[f'M{row}'].value = '☑ Yes  ☐ No'
        else:
            ws[f'M{row}'].value = '☐ Yes  ☐ No'
        ws[f'M{row}'].font = value_font
        ws[f'M{row}'].alignment = center_align
        
        # Remarks列（O-R合并）
        ws.merge_cells(f'O{row}:R{row}')
        ws[f'O{row}'].value = ''
        ws[f'O{row}'].alignment = left_align
    
    # 第27-29行：Remarks区域
    ws.merge_cells('A27:R27')
    ws['A27'].value = CellRichText(
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000'), 'Remarks:\n'),
        TextBlock(InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF'), 'Примечания:')
    )
    ws['A27'].alignment = left_align
    
    ws.merge_cells('A28:R28')
    ws.merge_cells('A29:R29')
    
    # === 添加边框（基于COM读取的边框模式）===
    # 顶部边框（第1行）
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        ws[f'{col}1'].border = Border(top=medium_border)
    # 左右边框
    ws['A1'].border = Border(left=medium_border, top=medium_border)
    ws['B1'].border = Border(left=medium_border, top=medium_border)
    ws['C1'].border = Border(left=medium_border, right=medium_border, top=medium_border)
    ws['O1'].border = Border(right=medium_border, top=medium_border)
    ws['P1'].border = Border(left=medium_border, right=medium_border, top=medium_border)
    ws['Q1'].border = Border(right=medium_border, top=medium_border)
    ws['R1'].border = Border(right=medium_border, top=medium_border)
    
    # 第2行左右边框
    ws['A2'].border = Border(left=medium_border)
    ws['C2'].border = Border(right=medium_border)
    ws['D2'].border = Border(left=medium_border)
    ws['O2'].border = Border(right=medium_border)
    ws['P2'].border = Border(left=medium_border)
    ws['R2'].border = Border(right=medium_border)
    
    # 第3行底部边框（标题区域分隔）
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        ws[f'{col}3'].border = Border(bottom=medium_border)
    ws['A3'].border = Border(left=medium_border, bottom=medium_border)
    ws['C3'].border = Border(right=medium_border, bottom=medium_border)
    ws['D3'].border = Border(left=medium_border, bottom=medium_border)
    ws['O3'].border = Border(right=medium_border, bottom=medium_border)
    ws['P3'].border = Border(left=medium_border, bottom=medium_border)
    ws['R3'].border = Border(right=medium_border, bottom=medium_border)
    
    # 第4-7行基础信息区域（完整的Thin细框线网格）
    thin_border = Side(style='thin', color='000000')
    for row in range(4, 8):
        for col_idx in range(1, 19):  # A-R
            col = chr(64 + col_idx)
            # 外框用Medium，内部用Thin
            left_border = medium_border if col_idx == 1 else thin_border
            right_border = medium_border if col_idx == 18 else thin_border
            top_border = medium_border if row == 4 else thin_border
            bottom_border = medium_border if row == 7 else thin_border
            
            ws[f'{col}{row}'].border = Border(
                left=left_border,
                right=right_border,
                top=top_border,
                bottom=bottom_border
            )
    
    # 第8-10行（认证声明区域）
    for row in range(8, 11):
        ws[f'A{row}'].border = Border(left=medium_border, top=medium_border if row == 8 else Side())
        ws[f'R{row}'].border = Border(right=medium_border)
    
    # 第10行底部边框（基础信息区域分隔）
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        cell = ws[f'{col}10']
        current_border = cell.border if cell.border else Border()
        ws[f'{col}10'].border = Border(
            left=current_border.left,
            right=current_border.right,
            bottom=medium_border
        )
    
    # 第11行：Checklist标题行（顶部和底部都是Medium）
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        ws[f'{col}11'].border = Border(
            left=medium_border if col == 'A' else Side(),
            right=medium_border if col == 'R' else Side(),
            top=medium_border,
            bottom=medium_border
        )
    
    # 第12-26行：Checklist区域（内部Thin细横线）
    for row in range(12, 27):
        for col_idx in range(1, 19):
            col = chr(64 + col_idx)
            # 外框用Medium，内部横线用Thin
            left_border = medium_border if col_idx == 1 else thin_border
            right_border = medium_border if col_idx == 18 else thin_border
            top_border = medium_border if row == 12 else thin_border
            bottom_border = thin_border
            
            ws[f'{col}{row}'].border = Border(
                left=left_border,
                right=right_border,
                top=top_border,
                bottom=bottom_border
            )
    
    # 第27-29行：Remarks区域
    for row in range(27, 30):
        ws[f'A{row}'].border = Border(
            left=medium_border,
            top=medium_border if row == 27 else Side()
        )
        ws[f'R{row}'].border = Border(right=medium_border)
    
    # 第29行底部边框（整个表格底部）
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        cell = ws[f'{col}29']
        current_border = cell.border if cell.border else Border()
        ws[f'{col}29'].border = Border(
            left=current_border.left,
            right=current_border.right,
            bottom=medium_border
        )
    
    # 隐藏网格线
    ws.sheet_view.showGridLines = False
    
    return ws


def create_pid_iso_page(wb, test_package_data, system_data, subsystem_data, pid_list, iso_list):
    """创建3.P&ID-ISO List工作表 - P&ID和ISO图纸清单"""
    ws = wb.create_sheet("3.P&ID-ISO List", 2)
    
    # 设置列宽（使用COM读取的准确值）
    column_widths = {
        'A': 5.05, 'B': 5.05, 'C': 5.05, 'D': 1.68, 'E': 5.79, 'F': 5.05,
        'G': 5.05, 'H': 5.79, 'I': 7.79, 'J': 5.05, 'K': 5.05, 'L': 5.05,
        'M': 2.68, 'N': 5.05, 'O': 10.16, 'P': 5.05, 'Q': 6.05, 'R': 8.32
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 设置行高（使用COM读取的准确值）
    # 行1-3: 28.0, 行4: 28.0, 行5: 53.3, 行6: 33.0, 行7: 30.8, 行8-31: 28.0
    row_heights = {
        1: 28.0, 2: 28.0, 3: 28.0, 4: 28.0, 5: 53.3, 6: 33.0, 7: 30.8
    }
    for row in range(8, 32):
        row_heights[row] = 28.0
    for row, height in row_heights.items():
        ws.row_dimensions[row].height = height
    
    # 样式定义
    title_font = Font(name='Times New Roman', size=18, bold=True)
    label_font = Font(name='Times New Roman', size=10, bold=True)
    value_font = Font(name='Times New Roman', size=10, bold=False)
    data_font = Font(name='Times New Roman', size=10, bold=False)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    medium_border = Side(style='medium', color='000000')
    thin_border = Side(style='thin', color='000000')
    
    # InlineFont定义（用于RichText）
    title_black = InlineFont(rFont='Times New Roman', sz=18.0, b=True, color='000000')
    title_blue = InlineFont(rFont='Times New Roman', sz=18.0, b=True, color='FF0000FF')
    label_black = InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='000000')
    label_blue = InlineFont(rFont='Times New Roman', sz=10.0, b=True, color='FF0000FF')
    value_black = InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='000000')
    value_blue = InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='FF0000FF')
    
    # === 第一部分：标题区域（A1:R3）===
    # Logo区域（与Cover Page和Contents相同）
    bcc_logo_path = os.path.join(BASE_DIR, 'static', 'images', 'bcc_logo.png')
    gcc_logo_path = os.path.join(BASE_DIR, 'static', 'images', 'gcc_logo.png')
    
    try:
        if os.path.exists(gcc_logo_path):
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            
            img_gcc = XLImage(gcc_logo_path)
            img_gcc.width = 100
            img_gcc.height = 70
            marker = AnchorMarker(col=0, colOff=0, row=0, rowOff=pixels_to_EMU(10))
            size = XDRPositiveSize2D(cx=pixels_to_EMU(100), cy=pixels_to_EMU(70))
            img_gcc.anchor = OneCellAnchor(_from=marker, ext=size)
            ws._images.append(img_gcc)
    except:
        pass
    
    try:
        if os.path.exists(bcc_logo_path):
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            
            img_bcc = XLImage(bcc_logo_path)
            img_bcc.width = 100
            img_bcc.height = 70
            marker = AnchorMarker(col=15, colOff=pixels_to_EMU(5), row=0, rowOff=pixels_to_EMU(10))
            size = XDRPositiveSize2D(cx=pixels_to_EMU(100), cy=pixels_to_EMU(70))
            img_bcc.anchor = OneCellAnchor(_from=marker, ext=size)
            ws._images.append(img_bcc)
    except:
        pass
    
    # 标题文字（D1:O2）
    ws.merge_cells('D1:O2')
    ws['D1'].value = CellRichText(
        TextBlock(title_black, 'P&ID & ISO Drawing List\n'),
        TextBlock(title_blue, 'Перечень Технологических схем и Изометрических чертежей')
    )
    ws['D1'].alignment = center_align
    
    # 项目名称（D3:O3）
    ws.merge_cells('D3:O3')
    ws['D3'].value = CellRichText(
        TextBlock(label_black, 'Gas Chemical Complex within Ethane-rich Gas Processing Complex\n'),
        TextBlock(label_blue, 'Газохимический комплекс в составе Комплекса переработки этансодержащего газа')
    )
    ws['D3'].alignment = center_align
    
    # === 第二部分：基础信息（A4:R7）===
    # 第4行：System No. / Sub-System No.
    ws.merge_cells('A4:D4')
    ws['A4'].value = CellRichText(
        TextBlock(label_black, 'System No.:\n'),
        TextBlock(label_blue, 'Система №.:')
    )
    ws['A4'].alignment = left_align
    
    ws.merge_cells('E4:I4')
    ws['E4'].value = system_data.get('SystemCode', 'N/A')
    ws['E4'].font = value_font
    ws['E4'].alignment = left_align
    
    ws.merge_cells('J4:M4')
    ws['J4'].value = CellRichText(
        TextBlock(label_black, 'Sub-System No.:\n'),
        TextBlock(label_blue, 'Подсистема №.:')
    )
    ws['J4'].alignment = left_align
    
    ws.merge_cells('N4:R4')
    ws['N4'].value = subsystem_data.get('SubSystemCode', 'N/A')
    ws['N4'].font = value_font
    ws['N4'].alignment = left_align
    
    # 第5行：Test Package No. / System Description
    ws.merge_cells('A5:D5')
    ws['A5'].value = CellRichText(
        TextBlock(label_black, 'Test Package No.\n'),
        TextBlock(label_blue, 'Тест-пакет №')
    )
    ws['A5'].alignment = left_align
    
    ws.merge_cells('E5:I5')
    ws['E5'].value = test_package_data.get('TestPackageID', 'N/A')
    ws['E5'].font = value_font
    ws['E5'].alignment = left_align
    
    ws.merge_cells('J5:M5')
    ws['J5'].value = CellRichText(
        TextBlock(label_black, 'System Description:\n'),
        TextBlock(label_blue, 'Система описания:')
    )
    ws['J5'].alignment = left_align
    
    ws.merge_cells('N5:R5')
    system_desc_en = system_data.get('SystemDescriptionENG', 'N/A')
    system_desc_ru = system_data.get('SystemDescriptionRUS', '')
    if system_desc_ru:
        ws['N5'].value = CellRichText(
            TextBlock(value_black, f"{system_desc_en}\n"),
            TextBlock(value_blue, system_desc_ru)
        )
    else:
        ws['N5'].value = system_desc_en
        ws['N5'].font = value_font
    ws['N5'].alignment = left_align
    
    # 第6行：Design Pressure / Test Pressure
    ws.merge_cells('A6:D6')
    ws['A6'].value = CellRichText(
        TextBlock(label_black, 'Design Pressure:\n'),
        TextBlock(label_blue, 'Расчетное давление:')
    )
    ws['A6'].alignment = left_align
    
    ws['E6'].value = test_package_data.get('DesignPressure', '')
    ws['E6'].font = value_font
    ws['E6'].alignment = center_align
    
    ws.merge_cells('F6:I6')
    ws['F6'].value = 'Mpa'
    ws['F6'].font = value_font
    ws['F6'].alignment = left_align
    
    ws.merge_cells('J6:M6')
    ws['J6'].value = CellRichText(
        TextBlock(label_black, 'Test Pressure:\n'),
        TextBlock(label_blue, 'Испытательное давление:')
    )
    ws['J6'].alignment = left_align
    
    ws['N6'].value = test_package_data.get('TestPressure', '')
    ws['N6'].font = value_font
    ws['N6'].alignment = center_align
    
    ws.merge_cells('O6:R6')
    ws['O6'].value = 'Mpa'
    ws['O6'].font = value_font
    ws['O6'].alignment = left_align
    
    # 第7行：Test Type / Test Medium
    ws.merge_cells('A7:D7')
    ws['A7'].value = CellRichText(
        TextBlock(label_black, 'Test Type:\n'),
        TextBlock(label_blue, 'Тип испытания:')
    )
    ws['A7'].alignment = left_align
    
    ws.merge_cells('E7:I7')
    test_type = test_package_data.get('TestType', '')
    test_type_info = TEST_TYPE_MAPPING.get(test_type, {'en': test_type, 'ru': ''})
    if test_type_info['ru']:
        ws['E7'].value = CellRichText(
            TextBlock(value_black, f"{test_type_info['en']}\n"),
            TextBlock(value_blue, test_type_info['ru'])
        )
    else:
        ws['E7'].value = test_type_info['en']
        ws['E7'].font = value_font
    ws['E7'].alignment = left_align
    
    ws.merge_cells('J7:M7')
    ws['J7'].value = CellRichText(
        TextBlock(label_black, 'Test Medium:\n'),
        TextBlock(label_blue, 'Среда испытания:')
    )
    ws['J7'].alignment = left_align
    
    ws.merge_cells('N7:R7')
    test_medium = test_package_data.get('TestMedium', '')
    test_medium_info = TEST_MEDIUM_MAPPING.get(test_medium, {'en': test_medium, 'ru': ''})
    if test_medium_info['ru']:
        ws['N7'].value = CellRichText(
            TextBlock(value_black, f"{test_medium_info['en']}\n"),
            TextBlock(value_blue, test_medium_info['ru'])
        )
    else:
        ws['N7'].value = test_medium_info['en']
        ws['N7'].font = value_font
    ws['N7'].alignment = left_align
    
    # === 第三部分：P&ID清单（A8:R14）===
    # 第8行：P&ID List标题
    ws.merge_cells('A8:R8')
    ws['A8'].value = CellRichText(
        TextBlock(label_black, 'P&ID List / '),
        TextBlock(label_blue, 'Перечень схем Трубопровода и КиПа')
    )
    ws['A8'].alignment = left_align
    
    # 第9行：P&ID表头
    ws['A9'].value = 'No.'
    ws['A9'].font = label_font
    ws['A9'].alignment = center_align
    
    ws.merge_cells('B9:H9')
    ws['B9'].value = CellRichText(
        TextBlock(label_black, 'P & ID No.\n'),
        TextBlock(label_blue, 'Номер схемы Трубопровод и КиП')
    )
    ws['B9'].alignment = center_align
    
    ws['I9'].value = 'Rev. No.'
    ws['I9'].font = label_font
    ws['I9'].alignment = center_align
    
    ws['J9'].value = 'No.'
    ws['J9'].font = label_font
    ws['J9'].alignment = center_align
    
    ws.merge_cells('K9:Q9')
    ws['K9'].value = CellRichText(
        TextBlock(label_black, 'P & ID No.\n'),
        TextBlock(label_blue, 'Номер схемы Трубопровод и КиП')
    )
    ws['K9'].alignment = center_align
    
    ws['R9'].value = 'Rev. No.'
    ws['R9'].font = label_font
    ws['R9'].alignment = center_align
    
    # 第10-14行：P&ID数据（左侧1-5，右侧6-10）
    for i in range(5):
        row = 10 + i
        # 左侧编号
        ws[f'A{row}'].value = i + 1
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].alignment = center_align
        
        # 左侧数据
        if i < len(pid_list):
            ws.merge_cells(f'B{row}:H{row}')
            ws[f'B{row}'].value = pid_list[i].get('DrawingNumber', '')
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].alignment = left_align
            
            ws[f'I{row}'].value = pid_list[i].get('RevisionNumber', '')
            ws[f'I{row}'].font = data_font
            ws[f'I{row}'].alignment = center_align
        else:
            ws.merge_cells(f'B{row}:H{row}')
        
        # 右侧编号
        ws[f'J{row}'].value = i + 6
        ws[f'J{row}'].font = label_font
        ws[f'J{row}'].alignment = center_align
        
        # 右侧数据
        if (i + 5) < len(pid_list):
            ws.merge_cells(f'K{row}:Q{row}')
            ws[f'K{row}'].value = pid_list[i + 5].get('DrawingNumber', '')
            ws[f'K{row}'].font = data_font
            ws[f'K{row}'].alignment = left_align
            
            ws[f'R{row}'].value = pid_list[i + 5].get('RevisionNumber', '')
            ws[f'R{row}'].font = data_font
            ws[f'R{row}'].alignment = center_align
        else:
            ws.merge_cells(f'K{row}:Q{row}')
    
    # === 第四部分：ISO清单（A15:R31）===
    # 第15行：ISO List标题
    ws.merge_cells('A15:R15')
    ws['A15'].value = CellRichText(
        TextBlock(label_black, 'ISO Drawing List / '),
        TextBlock(label_blue, 'Перечень Изометрических чертежей')
    )
    ws['A15'].alignment = left_align
    
    # 第16行：ISO表头
    ws['A16'].value = 'No.'
    ws['A16'].font = label_font
    ws['A16'].alignment = center_align
    
    ws.merge_cells('B16:H16')
    ws['B16'].value = CellRichText(
        TextBlock(label_black, 'ISO Drawing No.\n'),
        TextBlock(label_blue, 'Номер изометрического чертежа')
    )
    ws['B16'].alignment = center_align
    
    ws['I16'].value = 'Rev. No.'
    ws['I16'].font = label_font
    ws['I16'].alignment = center_align
    
    ws['J16'].value = 'No.'
    ws['J16'].font = label_font
    ws['J16'].alignment = center_align
    
    ws.merge_cells('K16:Q16')
    ws['K16'].value = CellRichText(
        TextBlock(label_black, 'ISO Drawing No.\n'),
        TextBlock(label_blue, 'Номер изометрического чертежа')
    )
    ws['K16'].alignment = center_align
    
    ws['R16'].value = 'Rev. No.'
    ws['R16'].font = label_font
    ws['R16'].alignment = center_align
    
    # 第17-31行：ISO数据（左侧1-15，右侧16-30）
    for i in range(15):
        row = 17 + i
        # 左侧编号
        ws[f'A{row}'].value = i + 1
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].alignment = center_align
        
        # 左侧数据
        if i < len(iso_list):
            ws.merge_cells(f'B{row}:H{row}')
            ws[f'B{row}'].value = iso_list[i].get('DrawingNumber', '')
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].alignment = left_align
            
            ws[f'I{row}'].value = iso_list[i].get('RevisionNumber', '')
            ws[f'I{row}'].font = data_font
            ws[f'I{row}'].alignment = center_align
        else:
            ws.merge_cells(f'B{row}:H{row}')
        
        # 右侧编号
        ws[f'J{row}'].value = i + 16
        ws[f'J{row}'].font = label_font
        ws[f'J{row}'].alignment = center_align
        
        # 右侧数据
        if (i + 15) < len(iso_list):
            ws.merge_cells(f'K{row}:Q{row}')
            ws[f'K{row}'].value = iso_list[i + 15].get('DrawingNumber', '')
            ws[f'K{row}'].font = data_font
            ws[f'K{row}'].alignment = left_align
            
            ws[f'R{row}'].value = iso_list[i + 15].get('RevisionNumber', '')
            ws[f'R{row}'].font = data_font
            ws[f'R{row}'].alignment = center_align
        else:
            ws.merge_cells(f'K{row}:Q{row}')
    
    # === 添加边框 ===
    # 标题区域（A1:R3）
    for col_idx in range(1, 19):
        for row in range(1, 4):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=thin_border if col_idx in [1, 4, 16] else Side(),
                right=thin_border if col_idx in [3, 15, 18] else Side(),
                top=thin_border if row == 1 else Side(),
                bottom=thin_border if row == 3 else Side()
            )
    
    # 基础信息区域（A4:R7）- 每个单元格都有完整的Thin框线
    for col_idx in range(1, 19):
        for row in range(4, 8):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=thin_border,
                right=thin_border,
                top=thin_border,
                bottom=thin_border
            )
    
    # P&ID清单区域（A8:R14）
    # 第8行：标题行
    for col_idx in range(1, 19):
        ws.cell(row=8, column=col_idx).border = Border(
            left=thin_border if col_idx == 1 else Side(),
            right=thin_border if col_idx == 18 else Side(),
            top=thin_border,
            bottom=thin_border
        )
    
    # 第9-14行：表头和数据行
    for row in range(9, 15):
        for col_idx in range(1, 19):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=thin_border if col_idx in [1, 2, 10, 11] else Side(),
                right=thin_border if col_idx in [1, 9, 10, 18] else Side(),
                top=thin_border if row == 9 else Side(),
                bottom=thin_border  # 每行都有底部边框
            )
    
    # ISO清单区域（A15:R31）
    # 第15行：标题行
    for col_idx in range(1, 19):
        ws.cell(row=15, column=col_idx).border = Border(
            left=thin_border if col_idx == 1 else Side(),
            right=thin_border if col_idx == 18 else Side(),
            top=thin_border,
            bottom=thin_border
        )
    
    # 第16-31行：表头和数据行
    for row in range(16, 32):
        for col_idx in range(1, 19):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=thin_border if col_idx in [1, 2, 10, 11] else Side(),
                right=thin_border if col_idx in [1, 9, 10, 18] else Side(),
                top=thin_border if row == 16 else Side(),
                bottom=thin_border  # 每行都有底部边框
            )
    
    # 隐藏网格线
    ws.sheet_view.showGridLines = False
    
    return ws


def export_test_package_to_excel(test_package_data, system_data, subsystem_data, include_attachments=False):
    """
    导出试压包资料包
    
    Args:
        test_package_data: 试压包数据
        system_data: 系统数据
        subsystem_data: 子系统数据
        include_attachments: 是否包含附件（True=ZIP压缩包，False=仅Excel）
    
    Returns:
        Flask Response对象（Excel或ZIP文件）
    """
    # 生成Excel文件
    wb = Workbook()
    # 移除默认的Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 创建Cover Page
    create_cover_page(wb, test_package_data, system_data, subsystem_data)
    
    # 创建Contents Page
    test_package_id = test_package_data.get('TestPackageID')
    checklist_status = check_checklist_status(test_package_id)
    create_contents_page(wb, test_package_data, system_data, subsystem_data, checklist_status)
    
    # 查询P&ID和ISO数据
    from database import create_connection
    pid_list = []
    iso_list = []
    
    conn = create_connection()
    if conn:
        try:
            cur = conn.cursor(dictionary=True)
            
            # 查询P&ID清单
            cur.execute("""
                SELECT PIDNo as DrawingNumber, RevNo as RevisionNumber 
                FROM PIDList 
                WHERE TestPackageID = %s 
                ORDER BY PIDNo
            """, (test_package_id,))
            pid_list = cur.fetchall()
            
            # 查询ISO清单（从ISODrawingList表或WeldingList表提取）
            # 优先从ISODrawingList读取
            cur.execute("""
                SELECT ISODrawingNo as DrawingNumber, RevNo as RevisionNumber
                FROM ISODrawingList 
                WHERE TestPackageID = %s
                ORDER BY ISODrawingNo
            """, (test_package_id,))
            iso_list = cur.fetchall()
            
            # 如果ISODrawingList为空，则从WeldingList提取
            if not iso_list:
                cur.execute("""
                    SELECT DISTINCT DrawingNumber, '' as RevisionNumber
                    FROM WeldingList 
                    WHERE TestPackageID = %s 
                    AND DrawingNumber IS NOT NULL 
                    AND DrawingNumber != ''
                    AND (DrawingNumber LIKE '%ISO%' OR DrawingNumber LIKE '%IS0%')
                    ORDER BY DrawingNumber
                """, (test_package_id,))
                iso_list = cur.fetchall()
            
        finally:
            conn.close()
    
    # 创建P&ID-ISO List Page
    create_pid_iso_page(wb, test_package_data, system_data, subsystem_data, pid_list, iso_list)
    
    # 生成文件名
    package_id = test_package_data.get('TestPackageID', 'unknown').replace('/', '-')
    
    if not include_attachments:
        # 仅导出Excel
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Test_Package_{package_id}.xlsx"
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    else:
        # 导出Excel + 附件（ZIP压缩包）
        return export_test_package_with_attachments(wb, test_package_id, package_id)


def export_test_package_with_attachments(wb, test_package_id, package_id):
    """导出试压包Excel和附件的ZIP压缩包"""
    import zipfile
    import tempfile
    import shutil
    from database import create_connection
    
    # 创建临时目录
    temp_dir = tempfile.mkdtemp()
    
    try:
        # 保存Excel到临时目录
        excel_filename = f"Test_Package_{package_id}.xlsx"
        excel_path = os.path.join(temp_dir, excel_filename)
        wb.save(excel_path)
        
        # 创建Attachments目录
        attachments_dir = os.path.join(temp_dir, 'Attachments')
        os.makedirs(attachments_dir, exist_ok=True)
        
        # 查询并复制所有附件
        conn = create_connection()
        if conn:
            try:
                cur = conn.cursor(dictionary=True)
                cur.execute("""
                    SELECT ModuleName, FileName, FilePath 
                    FROM TestPackageAttachments 
                    WHERE TestPackageID = %s
                    ORDER BY ModuleName, UploadedAt
                """, (test_package_id,))
                
                attachments = cur.fetchall()
                
                # 按模块分类复制附件
                module_mapping = {
                    'PID_Drawings': '3.PID_Drawings',
                    'ISO_Drawings': '4.ISO_Drawings',
                    'Symbols_Legend': '5.Symbols_Legend',
                    'Test_Flow_Chart': '8.Test_Flow_Chart',
                    'Test_Check_List': '9.Test_Check_List',
                    'Calibration_Certificates': '10.Calibration_Certificates',
                    'Test_Certificate': '11.Test_Certificate',
                    'Flushing_Certificate': '12.Flushing_Certificate',
                    'Reinstatement_Check_List': '13.Reinstatement_Check_List',
                    'Others': '14.Others'
                }
                
                for att in attachments:
                    module_name = att['ModuleName']
                    file_path = att['FilePath']
                    file_name = att['FileName']
                    
                    if module_name in module_mapping and os.path.exists(file_path):
                        # 创建模块目录
                        module_dir = os.path.join(attachments_dir, module_mapping[module_name])
                        os.makedirs(module_dir, exist_ok=True)
                        
                        # 复制文件
                        dest_path = os.path.join(module_dir, file_name)
                        shutil.copy2(file_path, dest_path)
                
            finally:
                conn.close()
        
        # 创建ZIP压缩包
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 添加Excel文件
            zip_file.write(excel_path, excel_filename)
            
            # 添加所有附件
            for root, dirs, files in os.walk(attachments_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zip_file.write(file_path, arcname)
        
        zip_buffer.seek(0)
        
        # 返回ZIP文件
        zip_filename = f"Test_Package_{package_id}_with_Attachments.zip"
        
        return Response(
            zip_buffer.getvalue(),
            mimetype='application/zip',
            headers={'Content-Disposition': f'attachment; filename={zip_filename}'}
        )
        
    finally:
        # 清理临时目录
        try:
            shutil.rmtree(temp_dir)
        except:
            pass


# ============================================================================
# 基于模板的新导出方案（推荐）
# ============================================================================

def has_russian(text):
    """检查文本是否包含俄语字符"""
    if not text:
        return False
    # 俄语字符范围：\u0400-\u04FF
    return any('\u0400' <= char <= '\u04FF' for char in str(text))


def fill_cover_page_com(ws, test_package_data, system_data, subsystem_data):
    """
    使用COM填充Cover Page数据
    只修改值单元格，不修改标签单元格（保留模板的蓝色俄语）
    
    Args:
        ws: Excel COM工作表对象
    """
    # 填充试压包编号（A11）
    ws.Cells(11, 1).Value = test_package_data.get('TestPackageID', '')
    
    # 填充详细信息值（H列，不修改A列标签！）
    # H18: System Code
    ws.Cells(18, 8).Value = system_data.get('SystemCode', 'N/A')
    
    # H19: Sub-System Code
    ws.Cells(19, 8).Value = subsystem_data.get('SubSystemCode', 'N/A')
    
    # H20: System / Sub-system Description
    desc_en = subsystem_data.get('SubSystemDescriptionENG', 'N/A')
    desc_ru = subsystem_data.get('SubSystemDescriptionRUS', '')
    if desc_ru:
        ws.Cells(20, 8).Value = f"{desc_en}\n{desc_ru}"
    else:
        ws.Cells(20, 8).Value = desc_en
    
    # H21: Piping Material
    pipe_material = test_package_data.get('PipeMaterial', '')
    material_info = PIPE_MATERIAL_MAPPING.get(pipe_material, {'en': pipe_material or '', 'ru': ''})
    if material_info['ru']:
        ws.Cells(21, 8).Value = f"{material_info['en']}\n{material_info['ru']}"
    else:
        ws.Cells(21, 8).Value = material_info['en'] or 'N/A'
    
    # H22: Piping class
    ws.Cells(22, 8).Value = test_package_data.get('PipingClass', '')
    
    # H23: Test Type
    test_type = test_package_data.get('TestType', '')
    test_type_info = TEST_TYPE_MAPPING.get(test_type, {'en': test_type or '', 'ru': ''})
    if test_type_info['ru']:
        ws.Cells(23, 8).Value = f"{test_type_info['en']}\n{test_type_info['ru']}"
    else:
        ws.Cells(23, 8).Value = test_type_info['en'] or 'N/A'
    
    # H24: Test Medium
    test_medium = test_package_data.get('TestMedium', '')
    medium_info = TEST_MEDIUM_MAPPING.get(test_medium, {'en': test_medium or '', 'ru': ''})
    if medium_info['ru']:
        ws.Cells(24, 8).Value = f"{medium_info['en']}\n{medium_info['ru']}"
    else:
        ws.Cells(24, 8).Value = medium_info['en'] or 'N/A'
    
    # H25: Design Pressure（数字格式，两位小数，不带单位）
    design_pressure = test_package_data.get('DesignPressure', '')
    if design_pressure:
        try:
            ws.Cells(25, 8).Value = float(design_pressure)
            ws.Cells(25, 8).NumberFormat = "0.00"
        except:
            ws.Cells(25, 8).Value = design_pressure
    else:
        ws.Cells(25, 8).Value = ''
    
    # H26: Test Pressure（数字格式，两位小数，不带单位）
    test_pressure = test_package_data.get('TestPressure', '')
    if test_pressure:
        try:
            ws.Cells(26, 8).Value = float(test_pressure)
            ws.Cells(26, 8).NumberFormat = "0.00"
        except:
            ws.Cells(26, 8).Value = test_pressure
    else:
        ws.Cells(26, 8).Value = ''


def fill_contents_page_com(ws, test_package_data, system_data, subsystem_data, checklist_status):
    """
    使用COM填充Contents Page数据
    只修改值单元格，不修改标签单元格
    """
    # 基础信息值（不修改标签A4, J4, A5, J5等！）
    # E4: System Code
    ws.Cells(4, 5).Value = system_data.get('SystemCode', 'N/A')
    
    # N4: Sub-System Code
    ws.Cells(4, 14).Value = subsystem_data.get('SubSystemCode', 'N/A')
    
    # E5: Test Package No
    ws.Cells(5, 5).Value = test_package_data.get('TestPackageID', 'N/A')
    
    # N5: System Description
    system_desc_en = system_data.get('SystemDescriptionENG', 'N/A')
    system_desc_ru = system_data.get('SystemDescriptionRUS', '')
    if system_desc_ru:
        ws.Cells(5, 14).Value = f"{system_desc_en}\n{system_desc_ru}"
    else:
        ws.Cells(5, 14).Value = system_desc_en
    
    # E6: Design Pressure（数字格式，两位小数，不带单位）
    design_pressure = test_package_data.get('DesignPressure', '')
    if design_pressure:
        try:
            ws.Cells(6, 5).Value = float(design_pressure)
            ws.Cells(6, 5).NumberFormat = "0.00"
        except:
            ws.Cells(6, 5).Value = design_pressure
    else:
        ws.Cells(6, 5).Value = ''
    
    # N6: Test Pressure（数字格式，两位小数，不带单位）
    test_pressure = test_package_data.get('TestPressure', '')
    if test_pressure:
        try:
            ws.Cells(6, 14).Value = float(test_pressure)
            ws.Cells(6, 14).NumberFormat = "0.00"
        except:
            ws.Cells(6, 14).Value = test_pressure
    else:
        ws.Cells(6, 14).Value = ''
    
    # E7: Test Type
    test_type = test_package_data.get('TestType', '')
    test_type_info = TEST_TYPE_MAPPING.get(test_type, {'en': test_type, 'ru': ''})
    if test_type_info['ru']:
        ws.Cells(7, 5).Value = f"{test_type_info['en']}\n{test_type_info['ru']}"
    else:
        ws.Cells(7, 5).Value = test_type_info['en']
    
    # N7: Test Medium
    test_medium = test_package_data.get('TestMedium', '')
    test_medium_info = TEST_MEDIUM_MAPPING.get(test_medium, {'en': test_medium, 'ru': ''})
    if test_medium_info['ru']:
        ws.Cells(7, 14).Value = f"{test_medium_info['en']}\n{test_medium_info['ru']}"
    else:
        ws.Cells(7, 14).Value = test_medium_info['en']
    
    # Checklist区域（第13-26行）
    # 只修改：A列编号（去掉.0）和M列复选框
    checklist_items = [
        ('1', '1.0'), ('2', '2.0'), ('3', '3.0'), ('4', '4.0'),
        ('5', '5.0'), ('6', '6.0'), ('7', '7.0'), ('8', '8.0'),
        ('9', '9.0'), ('10', '10.0'), ('11', '11.0'), ('12', '12.0'),
        ('13', '13.0'), ('14', '14.0'),
    ]
    
    for idx, (display_no, item_key) in enumerate(checklist_items):
        row = 13 + idx
        
        # 修改编号（A列）- 去掉.0
        ws.Cells(row, 1).Value = display_no
        
        # 修改复选框（M列）- 不修改B列描述！
        is_completed = checklist_status.get(item_key, False)
        if is_completed:
            ws.Cells(row, 13).Value = '☑ Yes  ☐ No'
        else:
            ws.Cells(row, 13).Value = '☐ Yes  ☐ No'


def fill_pid_iso_page_com(ws, test_package_data, system_data, subsystem_data, pid_list, iso_list):
    """
    使用COM填充P&ID-ISO List Page数据
    只修改值单元格，不修改标签单元格
    """
    # 基础信息值（不修改标签！）
    # E4: System Code
    ws.Cells(4, 5).Value = system_data.get('SystemCode', 'N/A')
    
    # N4: Sub-System Code
    ws.Cells(4, 14).Value = subsystem_data.get('SubSystemCode', 'N/A')
    
    # E5: Test Package No
    ws.Cells(5, 5).Value = test_package_data.get('TestPackageID', 'N/A')
    
    # N5: System Description
    system_desc_en = system_data.get('SystemDescriptionENG', 'N/A')
    system_desc_ru = system_data.get('SystemDescriptionRUS', '')
    if system_desc_ru:
        ws.Cells(5, 14).Value = f"{system_desc_en}\n{system_desc_ru}"
    else:
        ws.Cells(5, 14).Value = system_desc_en
    
    # E6: Design Pressure（数字格式，两位小数，不带单位）
    design_pressure = test_package_data.get('DesignPressure', '')
    if design_pressure:
        try:
            ws.Cells(6, 5).Value = float(design_pressure)
            ws.Cells(6, 5).NumberFormat = "0.00"
        except:
            ws.Cells(6, 5).Value = design_pressure
    else:
        ws.Cells(6, 5).Value = ''
    
    # N6: Test Pressure（数字格式，两位小数，不带单位）
    test_pressure = test_package_data.get('TestPressure', '')
    if test_pressure:
        try:
            ws.Cells(6, 14).Value = float(test_pressure)
            ws.Cells(6, 14).NumberFormat = "0.00"
        except:
            ws.Cells(6, 14).Value = test_pressure
    else:
        ws.Cells(6, 14).Value = ''
    
    # E7: Test Type
    test_type = test_package_data.get('TestType', '')
    test_type_info = TEST_TYPE_MAPPING.get(test_type, {'en': test_type, 'ru': ''})
    if test_type_info['ru']:
        ws.Cells(7, 5).Value = f"{test_type_info['en']}\n{test_type_info['ru']}"
    else:
        ws.Cells(7, 5).Value = test_type_info['en']
    
    # N7: Test Medium
    test_medium = test_package_data.get('TestMedium', '')
    test_medium_info = TEST_MEDIUM_MAPPING.get(test_medium, {'en': test_medium, 'ru': ''})
    if test_medium_info['ru']:
        ws.Cells(7, 14).Value = f"{test_medium_info['en']}\n{test_medium_info['ru']}"
    else:
        ws.Cells(7, 14).Value = test_medium_info['en']
    
    # P&ID清单（第10-14行）
    # 清空示例数据
    for i in range(5):
        row = 10 + i
        ws.Cells(row, 1).Value = i + 1  # 编号1-5
        ws.Cells(row, 2).Value = ''     # B列清空
        ws.Cells(row, 9).Value = ''     # I列清空
        ws.Cells(row, 10).Value = i + 6 # 编号6-10
        ws.Cells(row, 11).Value = ''    # K列清空
        ws.Cells(row, 18).Value = ''    # R列清空
    
    # 填充P&ID数据
    for i, pid in enumerate(pid_list[:10]):
        row = 10 + (i % 5)
        if i < 5:  # 左侧
            ws.Cells(row, 2).Value = pid.get('DrawingNumber', '')  # B列
            ws.Cells(row, 9).Value = pid.get('RevisionNumber', '') # I列
        else:  # 右侧
            ws.Cells(row, 11).Value = pid.get('DrawingNumber', '')  # K列
            ws.Cells(row, 18).Value = pid.get('RevisionNumber', '') # R列
    
    # ISO清单（第17-31行）
    # 清空示例数据
    for i in range(15):
        row = 17 + i
        ws.Cells(row, 1).Value = i + 1   # 编号1-15
        ws.Cells(row, 2).Value = ''      # B列清空
        ws.Cells(row, 9).Value = ''      # I列清空
        ws.Cells(row, 10).Value = i + 16 # 编号16-30
        ws.Cells(row, 11).Value = ''     # K列清空
        ws.Cells(row, 18).Value = ''     # R列清空
    
    # 填充ISO数据
    for i, iso in enumerate(iso_list[:30]):
        row = 17 + (i % 15)
        if i < 15:  # 左侧
            ws.Cells(row, 2).Value = iso.get('DrawingNumber', '')   # B列
            ws.Cells(row, 9).Value = iso.get('RevisionNumber', '')  # I列
        else:  # 右侧
            ws.Cells(row, 11).Value = iso.get('DrawingNumber', '')  # K列
            ws.Cells(row, 18).Value = iso.get('RevisionNumber', '') # R列


def fill_cover_page_from_template(ws, test_package_data, system_data, subsystem_data):
    """
    基于模板填充Cover Page数据
    模板已包含所有格式，只需填充动态数据
    """
    # 添加Logo（模板图片被清除，需要重新添加）
    bcc_logo_path = os.path.join(BASE_DIR, 'static', 'images', 'bcc_logo.png')
    gcc_logo_path = os.path.join(BASE_DIR, 'static', 'images', 'gcc_logo.png')
    
    try:
        if os.path.exists(gcc_logo_path):
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            
            img_gcc = XLImage(gcc_logo_path)
            img_gcc.width = 100
            img_gcc.height = 70
            marker = AnchorMarker(col=0, colOff=0, row=0, rowOff=pixels_to_EMU(10))
            size = XDRPositiveSize2D(cx=pixels_to_EMU(100), cy=pixels_to_EMU(70))
            img_gcc.anchor = OneCellAnchor(_from=marker, ext=size)
            ws._images.append(img_gcc)
    except:
        pass
    
    try:
        if os.path.exists(bcc_logo_path):
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            
            img_bcc = XLImage(bcc_logo_path)
            img_bcc.width = 100
            img_bcc.height = 70
            marker = AnchorMarker(col=15, colOff=pixels_to_EMU(5), row=0, rowOff=pixels_to_EMU(10))
            size = XDRPositiveSize2D(cx=pixels_to_EMU(100), cy=pixels_to_EMU(70))
            img_bcc.anchor = OneCellAnchor(_from=marker, ext=size)
            ws._images.append(img_bcc)
    except:
        pass
    
    # 定义RichText字体（用于双语数据）
    value_black = InlineFont(rFont='Times New Roman', sz=12.0, b=True, color='000000')
    value_blue = InlineFont(rFont='Times New Roman', sz=12.0, b=True, color='FF0000FF')
    
    # 填充试压包编号（A11）
    ws['A11'].value = test_package_data.get('TestPackageID', '')
    
    # 填充详细信息（H18-R26）
    # H18: System Code
    ws['H18'].value = system_data.get('SystemCode', 'N/A')
    
    # H19: Sub-System Code
    ws['H19'].value = subsystem_data.get('SubSystemCode', 'N/A')
    
    # H20: System / Sub-system Description（可能包含俄语）
    desc_en = subsystem_data.get('SubSystemDescriptionENG', 'N/A')
    desc_ru = subsystem_data.get('SubSystemDescriptionRUS', '')
    if desc_ru:
        ws['H20'].value = CellRichText(
            TextBlock(value_black, f"{desc_en}\n"),
            TextBlock(value_blue, desc_ru)
        )
    else:
        ws['H20'].value = desc_en
    
    # H21: Piping Material（可能包含俄语）
    pipe_material = test_package_data.get('PipeMaterial', '')
    material_info = PIPE_MATERIAL_MAPPING.get(pipe_material, {'en': pipe_material or '', 'ru': ''})
    if material_info['ru']:
        ws['H21'].value = CellRichText(
            TextBlock(value_black, f"{material_info['en']}\n"),
            TextBlock(value_blue, material_info['ru'])
        )
    else:
        ws['H21'].value = material_info['en'] or 'N/A'
    
    # H22: Piping class
    ws['H22'].value = test_package_data.get('PipingClass', '')
    
    # H23: Test Type（可能包含俄语）
    test_type = test_package_data.get('TestType', '')
    test_type_info = TEST_TYPE_MAPPING.get(test_type, {'en': test_type or '', 'ru': ''})
    if test_type_info['ru']:
        ws['H23'].value = CellRichText(
            TextBlock(value_black, f"{test_type_info['en']}\n"),
            TextBlock(value_blue, test_type_info['ru'])
        )
    else:
        ws['H23'].value = test_type_info['en'] or 'N/A'
    
    # H24: Test Medium（可能包含俄语）
    test_medium = test_package_data.get('TestMedium', '')
    medium_info = TEST_MEDIUM_MAPPING.get(test_medium, {'en': test_medium or '', 'ru': ''})
    if medium_info['ru']:
        ws['H24'].value = CellRichText(
            TextBlock(value_black, f"{medium_info['en']}\n"),
            TextBlock(value_blue, medium_info['ru'])
        )
    else:
        ws['H24'].value = medium_info['en'] or 'N/A'
    
    # H25: Design Pressure, I25: Mpa
    ws['H25'].value = test_package_data.get('DesignPressure', '')
    
    # H26: Test Pressure, I26: Mpa
    ws['H26'].value = test_package_data.get('TestPressure', '')


def fill_contents_page_from_template(ws, test_package_data, system_data, subsystem_data, checklist_status):
    """
    基于模板填充Contents Page数据
    """
    # 添加Logo
    bcc_logo_path = os.path.join(BASE_DIR, 'static', 'images', 'bcc_logo.png')
    gcc_logo_path = os.path.join(BASE_DIR, 'static', 'images', 'gcc_logo.png')
    
    try:
        if os.path.exists(gcc_logo_path):
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            
            img_gcc = XLImage(gcc_logo_path)
            img_gcc.width = 100
            img_gcc.height = 70
            marker = AnchorMarker(col=0, colOff=0, row=0, rowOff=pixels_to_EMU(10))
            size = XDRPositiveSize2D(cx=pixels_to_EMU(100), cy=pixels_to_EMU(70))
            img_gcc.anchor = OneCellAnchor(_from=marker, ext=size)
            ws._images.append(img_gcc)
    except:
        pass
    
    try:
        if os.path.exists(bcc_logo_path):
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            
            img_bcc = XLImage(bcc_logo_path)
            img_bcc.width = 100
            img_bcc.height = 70
            marker = AnchorMarker(col=15, colOff=pixels_to_EMU(5), row=0, rowOff=pixels_to_EMU(10))
            size = XDRPositiveSize2D(cx=pixels_to_EMU(100), cy=pixels_to_EMU(70))
            img_bcc.anchor = OneCellAnchor(_from=marker, ext=size)
            ws._images.append(img_bcc)
    except:
        pass
    
    # 定义RichText字体（用于双语数据）
    value_black = InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='000000')
    value_blue = InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='FF0000FF')
    
    # === 基础信息区域（第4-7行）===
    # 第4行：System No. / Sub-System No.
    ws['E4'].value = system_data.get('SystemCode', 'N/A')
    ws['N4'].value = subsystem_data.get('SubSystemCode', 'N/A')
    
    # 第5行：Test Package No. / System Description
    ws['E5'].value = test_package_data.get('TestPackageID', 'N/A')
    
    system_desc_en = system_data.get('SystemDescriptionENG', 'N/A')
    system_desc_ru = system_data.get('SystemDescriptionRUS', '')
    if system_desc_ru:
        ws['N5'].value = CellRichText(
            TextBlock(value_black, f"{system_desc_en}\n"),
            TextBlock(value_blue, system_desc_ru)
        )
    else:
        ws['N5'].value = system_desc_en
    
    # 第6行：Design Pressure / Test Pressure
    design_pressure = test_package_data.get('DesignPressure', '')
    ws['E6'].value = f"{design_pressure} Mpa" if design_pressure else ''
    
    test_pressure = test_package_data.get('TestPressure', '')
    ws['N6'].value = f"{test_pressure} Mpa" if test_pressure else ''
    
    # 第7行：Test Type / Test Medium
    test_type = test_package_data.get('TestType', '')
    test_type_info = TEST_TYPE_MAPPING.get(test_type, {'en': test_type, 'ru': ''})
    if test_type_info['ru']:
        ws['E7'].value = CellRichText(
            TextBlock(value_black, f"{test_type_info['en']}\n"),
            TextBlock(value_blue, test_type_info['ru'])
        )
    else:
        ws['E7'].value = test_type_info['en']
    
    test_medium = test_package_data.get('TestMedium', '')
    test_medium_info = TEST_MEDIUM_MAPPING.get(test_medium, {'en': test_medium, 'ru': ''})
    if test_medium_info['ru']:
        ws['N7'].value = CellRichText(
            TextBlock(value_black, f"{test_medium_info['en']}\n"),
            TextBlock(value_blue, test_medium_info['ru'])
        )
    else:
        ws['N7'].value = test_medium_info['en']
    
    # === Checklist区域（第13-26行）===
    # Checklist项目定义（编号去掉.0）
    checklist_items = [
        ('1', '1.0'),
        ('2', '2.0'),
        ('3', '3.0'),
        ('4', '4.0'),
        ('5', '5.0'),
        ('6', '6.0'),
        ('7', '7.0'),
        ('8', '8.0'),
        ('9', '9.0'),
        ('10', '10.0'),
        ('11', '11.0'),
        ('12', '12.0'),
        ('13', '13.0'),
        ('14', '14.0'),
    ]
    
    for idx, (display_no, item_key) in enumerate(checklist_items):
        row = 13 + idx
        
        # 填充编号（A列）
        ws[f'A{row}'].value = display_no
        
        # 填充Yes/No复选框（M列，模板已合并M-N）
        is_completed = checklist_status.get(item_key, False)
        if is_completed:
            ws[f'M{row}'].value = '☑ Yes  ☐ No'
        else:
            ws[f'M{row}'].value = '☐ Yes  ☐ No'


def fill_pid_iso_page_from_template(ws, test_package_data, system_data, subsystem_data, pid_list, iso_list):
    """
    基于模板填充P&ID-ISO List Page数据
    """
    # 添加Logo
    bcc_logo_path = os.path.join(BASE_DIR, 'static', 'images', 'bcc_logo.png')
    gcc_logo_path = os.path.join(BASE_DIR, 'static', 'images', 'gcc_logo.png')
    
    try:
        if os.path.exists(gcc_logo_path):
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            
            img_gcc = XLImage(gcc_logo_path)
            img_gcc.width = 100
            img_gcc.height = 70
            marker = AnchorMarker(col=0, colOff=0, row=0, rowOff=pixels_to_EMU(10))
            size = XDRPositiveSize2D(cx=pixels_to_EMU(100), cy=pixels_to_EMU(70))
            img_gcc.anchor = OneCellAnchor(_from=marker, ext=size)
            ws._images.append(img_gcc)
    except:
        pass
    
    try:
        if os.path.exists(bcc_logo_path):
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            
            img_bcc = XLImage(bcc_logo_path)
            img_bcc.width = 100
            img_bcc.height = 70
            marker = AnchorMarker(col=15, colOff=pixels_to_EMU(5), row=0, rowOff=pixels_to_EMU(10))
            size = XDRPositiveSize2D(cx=pixels_to_EMU(100), cy=pixels_to_EMU(70))
            img_bcc.anchor = OneCellAnchor(_from=marker, ext=size)
            ws._images.append(img_bcc)
    except:
        pass
    
    # 定义RichText字体（用于双语数据）
    value_black = InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='000000')
    value_blue = InlineFont(rFont='Times New Roman', sz=10.0, b=False, color='FF0000FF')
    
    # === 基础信息区域（第4-7行）===
    # 第4行：System No. / Sub-System No.
    ws['E4'].value = system_data.get('SystemCode', 'N/A')
    ws['N4'].value = subsystem_data.get('SubSystemCode', 'N/A')
    
    # 第5行：Test Package No. / System Description
    ws['E5'].value = test_package_data.get('TestPackageID', 'N/A')
    
    system_desc_en = system_data.get('SystemDescriptionENG', 'N/A')
    system_desc_ru = system_data.get('SystemDescriptionRUS', '')
    if system_desc_ru:
        ws['N5'].value = CellRichText(
            TextBlock(value_black, f"{system_desc_en}\n"),
            TextBlock(value_blue, system_desc_ru)
        )
    else:
        ws['N5'].value = system_desc_en
    
    # 第6行：Design Pressure / Test Pressure
    ws['E6'].value = test_package_data.get('DesignPressure', '')
    ws['N6'].value = test_package_data.get('TestPressure', '')
    
    # 第7行：Test Type / Test Medium
    test_type = test_package_data.get('TestType', '')
    test_type_info = TEST_TYPE_MAPPING.get(test_type, {'en': test_type, 'ru': ''})
    if test_type_info['ru']:
        ws['E7'].value = CellRichText(
            TextBlock(value_black, f"{test_type_info['en']}\n"),
            TextBlock(value_blue, test_type_info['ru'])
        )
    else:
        ws['E7'].value = test_type_info['en']
    
    test_medium = test_package_data.get('TestMedium', '')
    test_medium_info = TEST_MEDIUM_MAPPING.get(test_medium, {'en': test_medium, 'ru': ''})
    if test_medium_info['ru']:
        ws['N7'].value = CellRichText(
            TextBlock(value_black, f"{test_medium_info['en']}\n"),
            TextBlock(value_blue, test_medium_info['ru'])
        )
    else:
        ws['N7'].value = test_medium_info['en']
    
    # === P&ID清单（第10-14行）===
    # 清空模板示例数据（模板已合并单元格，不再重复merge_cells）
    for i in range(5):
        row = 10 + i
        ws[f'A{row}'].value = i + 1
        ws[f'B{row}'].value = ''
        ws[f'I{row}'].value = ''
        ws[f'J{row}'].value = i + 6
        ws[f'K{row}'].value = ''
        ws[f'R{row}'].value = ''
    
    # 填充P&ID数据
    for i, pid in enumerate(pid_list[:10]):  # 最多10条
        row = 10 + (i % 5)  # 左侧0-4，右侧0-4
        
        if i < 5:  # 左侧（1-5）
            ws[f'B{row}'].value = pid.get('DrawingNumber', '')
            ws[f'I{row}'].value = pid.get('RevisionNumber', '')
        else:  # 右侧（6-10）
            ws[f'K{row}'].value = pid.get('DrawingNumber', '')
            ws[f'R{row}'].value = pid.get('RevisionNumber', '')
    
    # === ISO清单（第17-31行）===
    # 清空模板示例数据（模板已合并单元格，不再重复merge_cells）
    for i in range(15):
        row = 17 + i
        ws[f'A{row}'].value = i + 1
        ws[f'B{row}'].value = ''
        ws[f'I{row}'].value = ''
        ws[f'J{row}'].value = i + 16
        ws[f'K{row}'].value = ''
        ws[f'R{row}'].value = ''
    
    # 填充ISO数据
    for i, iso in enumerate(iso_list[:30]):  # 最多30条
        row = 17 + (i % 15)  # 左侧0-14，右侧0-14
        
        if i < 15:  # 左侧（1-15）
            ws[f'B{row}'].value = iso.get('DrawingNumber', '')
            ws[f'I{row}'].value = iso.get('RevisionNumber', '')
        else:  # 右侧（16-30）
            ws[f'K{row}'].value = iso.get('DrawingNumber', '')
            ws[f'R{row}'].value = iso.get('RevisionNumber', '')


def fill_welding_log_com(ws, test_package_data, system_data, subsystem_data, test_package_id):
    """
    使用COM填充Welding log数据
    填充基础信息（E4-R7）、焊口汇总（第11行）和NDE/PWHT状态（第15-21行）
    """
    from database import create_connection
    
    # 填充基础信息（E4-R7）- 不使用公式，使用实际数据
    ws.Cells(4, 5).Value = system_data.get('SystemCode', 'N/A')  # E4: System Code
    ws.Cells(4, 14).Value = subsystem_data.get('SubSystemCode', 'N/A')  # N4: Sub-System Code
    ws.Cells(5, 5).Value = test_package_data.get('TestPackageID', 'N/A')  # E5: Test Package No
    
    # N5: System Description
    system_desc_en = system_data.get('SystemDescriptionENG', 'N/A')
    system_desc_ru = system_data.get('SystemDescriptionRUS', '')
    if system_desc_ru:
        ws.Cells(5, 14).Value = f"{system_desc_en}\n{system_desc_ru}"
    else:
        ws.Cells(5, 14).Value = system_desc_en
    
    # E6: Design Pressure（数字格式，两位小数）
    design_pressure = test_package_data.get('DesignPressure', '')
    if design_pressure:
        try:
            ws.Cells(6, 5).Value = float(design_pressure)
            ws.Cells(6, 5).NumberFormat = "0.00"
        except:
            ws.Cells(6, 5).Value = design_pressure
    
    # M6: Test Pressure（数字格式，两位小数，不是N6！）
    test_pressure = test_package_data.get('TestPressure', '')
    if test_pressure:
        try:
            ws.Cells(6, 13).Value = float(test_pressure)
            ws.Cells(6, 13).NumberFormat = "0.00"
        except:
            ws.Cells(6, 13).Value = test_pressure
    
    # E7: Test Type
    test_type = test_package_data.get('TestType', '')
    test_type_info = TEST_TYPE_MAPPING.get(test_type, {'en': test_type, 'ru': ''})
    if test_type_info['ru']:
        ws.Cells(7, 5).Value = f"{test_type_info['en']}\n{test_type_info['ru']}"
    else:
        ws.Cells(7, 5).Value = test_type_info['en']
    
    # N7: Test Medium
    test_medium = test_package_data.get('TestMedium', '')
    test_medium_info = TEST_MEDIUM_MAPPING.get(test_medium, {'en': test_medium, 'ru': ''})
    if test_medium_info['ru']:
        ws.Cells(7, 14).Value = f"{test_medium_info['en']}\n{test_medium_info['ru']}"
    else:
        ws.Cells(7, 14).Value = test_medium_info['en']
    
    conn = create_connection()
    if not conn:
        return
    
    try:
        cur = conn.cursor(dictionary=True)
        
        # 查询焊口汇总数据（JointSummary）
        cur.execute("""
            SELECT TotalJoints, CompletedJoints, RemainingJoints,
                   TotalDIN, CompletedDIN, RemainingDIN
            FROM JointSummary 
            WHERE TestPackageID = %s
        """, (test_package_id,))
        joint_summary = cur.fetchone()
        
        if joint_summary:
            # 填充第11行：焊口数量和DIN汇总（使用数字格式，两位小数）
            ws.Cells(11, 1).Value = joint_summary['TotalJoints'] or 0  # A11: Total Joints
            ws.Cells(11, 3).Value = joint_summary['CompletedJoints'] or 0  # C11: Done Joints
            ws.Cells(11, 5).Value = joint_summary['RemainingJoints'] or 0  # E11: Remain Joints
            
            # Total DIN - 设置为数字格式（两位小数）
            ws.Cells(11, 7).Value = float(joint_summary['TotalDIN'] or 0.0)
            ws.Cells(11, 7).NumberFormat = "0.00"
            ws.Cells(11, 9).Value = float(joint_summary['CompletedDIN'] or 0.0)
            ws.Cells(11, 9).NumberFormat = "0.00"
            ws.Cells(11, 11).Value = float(joint_summary['RemainingDIN'] or 0.0)
            ws.Cells(11, 11).NumberFormat = "0.00"
        
        # 查询NDE/PWHT状态数据（NDEPWHTStatus）
        cur.execute("""
            SELECT VT_Total, VT_Completed, VT_Remaining,
                   RT_Total, RT_Completed, RT_Remaining,
                   PT_Total, PT_Completed, PT_Remaining,
                   HT_Total, HT_Completed, HT_Remaining,
                   PWHT_Total, PWHT_Completed, PWHT_Remaining,
                   PMI_Total, PMI_Completed, PMI_Remaining,
                   UT_Total, UT_Completed, UT_Remaining,
                   MT_Total, MT_Completed, MT_Remaining,
                   FT_Total, FT_Completed, FT_Remaining
            FROM NDEPWHTStatus 
            WHERE TestPackageID = %s
        """, (test_package_id,))
        nde_pwht = cur.fetchone()
        
        if nde_pwht:
            # 根据模板实际结构填充NDE/PWHT数据
            # 第15行：VT（左A/C/E）、RT（中G/I/K）、PT（右M/O/Q）
            ws.Cells(15, 1).Value = nde_pwht['VT_Total'] or 0
            ws.Cells(15, 3).Value = nde_pwht['VT_Completed'] or 0
            ws.Cells(15, 5).Value = nde_pwht['VT_Remaining'] or 0
            ws.Cells(15, 7).Value = nde_pwht['RT_Total'] or 0
            ws.Cells(15, 9).Value = nde_pwht['RT_Completed'] or 0
            ws.Cells(15, 11).Value = nde_pwht['RT_Remaining'] or 0
            ws.Cells(15, 13).Value = nde_pwht['PT_Total'] or 0
            ws.Cells(15, 15).Value = nde_pwht['PT_Completed'] or 0
            ws.Cells(15, 17).Value = nde_pwht['PT_Remaining'] or 0
            
            # 第18行：HT（左A/C/E）、PWHT（中G/I/K）、PMI（右M/O/Q）
            ws.Cells(18, 1).Value = nde_pwht['HT_Total'] or 0
            ws.Cells(18, 3).Value = nde_pwht['HT_Completed'] or 0
            ws.Cells(18, 5).Value = nde_pwht['HT_Remaining'] or 0
            ws.Cells(18, 7).Value = nde_pwht['PWHT_Total'] or 0
            ws.Cells(18, 9).Value = nde_pwht['PWHT_Completed'] or 0
            ws.Cells(18, 11).Value = nde_pwht['PWHT_Remaining'] or 0
            ws.Cells(18, 13).Value = nde_pwht['PMI_Total'] or 0
            ws.Cells(18, 15).Value = nde_pwht['PMI_Completed'] or 0
            ws.Cells(18, 17).Value = nde_pwht['PMI_Remaining'] or 0
            
            # 第21行：UT（左A/C/E）、MT（中G/I/K）、FT（右M/O/Q）
            ws.Cells(21, 1).Value = nde_pwht['UT_Total'] or 0
            ws.Cells(21, 3).Value = nde_pwht['UT_Completed'] or 0
            ws.Cells(21, 5).Value = nde_pwht['UT_Remaining'] or 0
            ws.Cells(21, 7).Value = nde_pwht['MT_Total'] or 0
            ws.Cells(21, 9).Value = nde_pwht['MT_Completed'] or 0
            ws.Cells(21, 11).Value = nde_pwht['MT_Remaining'] or 0
            ws.Cells(21, 13).Value = nde_pwht['FT_Total'] or 0
            ws.Cells(21, 15).Value = nde_pwht['FT_Completed'] or 0
            ws.Cells(21, 17).Value = nde_pwht['FT_Remaining'] or 0
    
    finally:
        conn.close()


def fill_punch_list_com(ws, test_package_data, system_data, subsystem_data, test_package_id):
    """
    使用COM填充Punch List数据
    填充基础信息（E4-R7）和Punch数据
    Punch List有5个表格区域（每个13行数据），分隔行不要动
    表头行：9, 24, 39, 54, 69
    数据行：10-22, 25-37, 40-52, 55-67, 70-82
    分隔行（不动）：23, 38, 53, 68, 83
    """
    from database import create_connection
    
    # 填充基础信息（E4-R7）- 不使用公式，使用实际数据
    ws.Cells(4, 5).Value = system_data.get('SystemCode', 'N/A')  # E4
    ws.Cells(4, 14).Value = subsystem_data.get('SubSystemCode', 'N/A')  # N4
    ws.Cells(5, 5).Value = test_package_data.get('TestPackageID', 'N/A')  # E5
    
    # N5: System Description
    system_desc_en = system_data.get('SystemDescriptionENG', 'N/A')
    system_desc_ru = system_data.get('SystemDescriptionRUS', '')
    if system_desc_ru:
        ws.Cells(5, 14).Value = f"{system_desc_en}\n{system_desc_ru}"
    else:
        ws.Cells(5, 14).Value = system_desc_en
    
    # E6, N6: Pressures
    design_pressure = test_package_data.get('DesignPressure', '')
    if design_pressure:
        try:
            ws.Cells(6, 5).Value = float(design_pressure)
            ws.Cells(6, 5).NumberFormat = "0.00"
        except:
            ws.Cells(6, 5).Value = design_pressure
    
    test_pressure = test_package_data.get('TestPressure', '')
    if test_pressure:
        try:
            ws.Cells(6, 14).Value = float(test_pressure)
            ws.Cells(6, 14).NumberFormat = "0.00"
        except:
            ws.Cells(6, 14).Value = test_pressure
    
    # E7, N7: Test Type and Medium
    test_type = test_package_data.get('TestType', '')
    test_type_info = TEST_TYPE_MAPPING.get(test_type, {'en': test_type, 'ru': ''})
    if test_type_info['ru']:
        ws.Cells(7, 5).Value = f"{test_type_info['en']}\n{test_type_info['ru']}"
    else:
        ws.Cells(7, 5).Value = test_type_info['en']
    
    test_medium = test_package_data.get('TestMedium', '')
    test_medium_info = TEST_MEDIUM_MAPPING.get(test_medium, {'en': test_medium, 'ru': ''})
    if test_medium_info['ru']:
        ws.Cells(7, 14).Value = f"{test_medium_info['en']}\n{test_medium_info['ru']}"
    else:
        ws.Cells(7, 14).Value = test_medium_info['en']
    
    conn = create_connection()
    if not conn:
        return
    
    try:
        cur = conn.cursor(dictionary=True)
        
        # 查询Punch List数据
        cur.execute("""
            SELECT PunchNo, ISODrawingNo, SheetNo, RevNo, Description,
                   Category, Cause, IssuedBy, Rectified, Verified
            FROM PunchList 
            WHERE TestPackageID = %s
            ORDER BY ISODrawingNo, SheetNo
        """, (test_package_id,))
        punch_list = cur.fetchall()
        
        # Punch List的5个表格区域（每个13行数据）
        # 数据行范围：10-22, 25-37, 40-52, 55-67, 70-82
        data_ranges = [
            range(10, 23),   # 第1组：13行
            range(25, 38),   # 第2组：13行
            range(40, 53),   # 第3组：13行
            range(55, 68),   # 第4组：13行
            range(70, 83),   # 第5组：13行
        ]
        
        # 将数据分组填充（每组13行）
        punch_idx = 0
        for group_idx, row_range in enumerate(data_ranges):
            for row in row_range:
                if punch_idx >= len(punch_list):
                    # 清空剩余行（如果数据不足）
                    ws.Cells(row, 1).Value = ''
                    ws.Cells(row, 2).Value = ''
                    ws.Cells(row, 3).Value = ''
                    ws.Cells(row, 5).Value = ''
                    ws.Cells(row, 6).Value = ''
                    ws.Cells(row, 7).Value = ''
                    ws.Cells(row, 14).Value = ''
                    ws.Cells(row, 15).Value = ''
                    ws.Cells(row, 16).Value = ''
                    ws.Cells(row, 17).Value = ''
                    ws.Cells(row, 18).Value = ''
                else:
                    punch = punch_list[punch_idx]
                    # 填充数据
                    ws.Cells(row, 1).Value = punch_idx + 1  # No.
                    ws.Cells(row, 3).Value = punch.get('PunchNo', '')  # Punch No.
                    ws.Cells(row, 2).Value = punch.get('ISODrawingNo', '')  # ISO Drawing No.
                    ws.Cells(row, 5).Value = punch.get('SheetNo', '')  # Sheet No.
                    ws.Cells(row, 6).Value = punch.get('RevNo', '')  # Rev. No.
                    ws.Cells(row, 7).Value = punch.get('Description', '')  # Description
                    ws.Cells(row, 14).Value = punch.get('Category', '')  # Category (N列)
                    ws.Cells(row, 15).Value = punch.get('Cause', '')  # Cause (O列)
                    ws.Cells(row, 16).Value = punch.get('IssuedBy', '')  # Issued by (P列)
                    ws.Cells(row, 17).Value = punch.get('Rectified', '')  # Rectified (Q列)
                    ws.Cells(row, 18).Value = punch.get('Verified', '')  # Verified (R列)
                    punch_idx += 1
    
    finally:
        conn.close()


def fill_simple_sheet_basic_info(ws, test_package_data, system_data, subsystem_data):
    """
    填充简单工作表的基础信息（通用函数）
    适用于：4.Legend, 7.Test flow chart, 8.Test Check List, 
           9.Calibration certificates, 10.Pipeline test certificate,
           11.Pipeline Flushing (Purging), 12.Reinstatement Check List
    
    这些工作表主要只需要填充基础信息，其他内容保持模板原样
    """
    sheet_name = ws.Name
    
    # 根据不同的工作表确定填充位置
    # 大部分工作表的基础信息位置相似，但有些略有不同
    
    if sheet_name == '4.Legend':
        # 4.Legend: System No.(E4), SubSystem No.(N4), Test Package No.(E5), etc.
        ws.Cells(4, 5).Value = system_data.get('SystemCode', 'N/A')  # E4
        ws.Cells(4, 14).Value = subsystem_data.get('SubSystemCode', 'N/A')  # N4
        ws.Cells(5, 5).Value = test_package_data.get('TestPackageID', 'N/A')  # E5
        
        # System Description (N5)
        system_desc_en = system_data.get('SystemDescriptionENG', 'N/A')
        system_desc_ru = system_data.get('SystemDescriptionRUS', '')
        if system_desc_ru:
            ws.Cells(5, 14).Value = f"{system_desc_en}\n{system_desc_ru}"
        else:
            ws.Cells(5, 14).Value = system_desc_en
        
        # Design Pressure (E6)
        design_pressure = test_package_data.get('DesignPressure', '')
        if design_pressure:
            try:
                ws.Cells(6, 5).Value = float(design_pressure)
                ws.Cells(6, 5).NumberFormat = "0.00"
            except:
                ws.Cells(6, 5).Value = design_pressure
        
        # Test Type (E7), Test Medium (N7)
        test_type = test_package_data.get('TestType', '')
        test_type_info = TEST_TYPE_MAPPING.get(test_type, {'en': test_type, 'ru': ''})
        if test_type_info['ru']:
            ws.Cells(7, 5).Value = f"{test_type_info['en']}\n{test_type_info['ru']}"
        else:
            ws.Cells(7, 5).Value = test_type_info['en']
        
        test_medium = test_package_data.get('TestMedium', '')
        test_medium_info = TEST_MEDIUM_MAPPING.get(test_medium, {'en': test_medium, 'ru': ''})
        if test_medium_info['ru']:
            ws.Cells(7, 14).Value = f"{test_medium_info['en']}\n{test_medium_info['ru']}"
        else:
            ws.Cells(7, 14).Value = test_medium_info['en']
    
    elif sheet_name == '7.Test flow chart':
        # 7.Test flow chart: System No.(O1), Subsystem No.(O2), Test Package No.(O3)
        ws.Cells(1, 15).Value = system_data.get('SystemCode', 'N/A')  # O1
        ws.Cells(2, 15).Value = subsystem_data.get('SubSystemCode', 'N/A')  # O2
        ws.Cells(3, 15).Value = test_package_data.get('TestPackageID', 'N/A')  # O3
    
    elif sheet_name == '8.Test Check List':
        # 8.Test Check List: System No.(D4), SubSystem No.(K4), Test Package No.(D5)
        ws.Cells(4, 4).Value = system_data.get('SystemCode', 'N/A')  # D4
        ws.Cells(4, 11).Value = subsystem_data.get('SubSystemCode', 'N/A')  # K4
        ws.Cells(5, 4).Value = test_package_data.get('TestPackageID', 'N/A')  # D5
        
        # Sub-System Description (K5)
        system_desc_en = system_data.get('SystemDescriptionENG', 'N/A')
        system_desc_ru = system_data.get('SystemDescriptionRUS', '')
        if system_desc_ru:
            ws.Cells(5, 11).Value = f"{system_desc_en}\n{system_desc_ru}"
        else:
            ws.Cells(5, 11).Value = system_desc_en
        
        # Design Pressure (D6)
        design_pressure = test_package_data.get('DesignPressure', '')
        if design_pressure:
            try:
                ws.Cells(6, 4).Value = float(design_pressure)
                ws.Cells(6, 4).NumberFormat = "0.00"
            except:
                ws.Cells(6, 4).Value = design_pressure
        
        # Test Type (D7), Test Medium (K7)
        test_type = test_package_data.get('TestType', '')
        test_type_info = TEST_TYPE_MAPPING.get(test_type, {'en': test_type, 'ru': ''})
        if test_type_info['ru']:
            ws.Cells(7, 4).Value = f"{test_type_info['en']}\n{test_type_info['ru']}"
        else:
            ws.Cells(7, 4).Value = test_type_info['en']
        
        test_medium = test_package_data.get('TestMedium', '')
        test_medium_info = TEST_MEDIUM_MAPPING.get(test_medium, {'en': test_medium, 'ru': ''})
        if test_medium_info['ru']:
            ws.Cells(7, 11).Value = f"{test_medium_info['en']}\n{test_medium_info['ru']}"
        else:
            ws.Cells(7, 11).Value = test_medium_info['en']
    
    elif sheet_name in ['10.Pipeline test certificate', '11.Pipeline Flushing (Purging) ']:
        # 10, 11: System No.(K1), Subsystem No.(K2), Test Package No.(K3)
        ws.Cells(1, 11).Value = system_data.get('SystemCode', 'N/A')  # K1
        ws.Cells(2, 11).Value = subsystem_data.get('SubSystemCode', 'N/A')  # K2
        ws.Cells(3, 11).Value = test_package_data.get('TestPackageID', 'N/A')  # K3
    
    elif sheet_name == '12.Reinstatement Check List':
        # 12.Reinstatement Check List: System No.(E4), SubSystem No.(N4), Test Package No.(E5)
        ws.Cells(4, 5).Value = system_data.get('SystemCode', 'N/A')  # E4
        ws.Cells(4, 14).Value = subsystem_data.get('SubSystemCode', 'N/A')  # N4
        ws.Cells(5, 5).Value = test_package_data.get('TestPackageID', 'N/A')  # E5
        
        # System Description (N5)
        system_desc_en = system_data.get('SystemDescriptionENG', 'N/A')
        system_desc_ru = system_data.get('SystemDescriptionRUS', '')
        if system_desc_ru:
            ws.Cells(5, 14).Value = f"{system_desc_en}\n{system_desc_ru}"
        else:
            ws.Cells(5, 14).Value = system_desc_en
    
    # 9.Calibration certificates 暂时不填充（可能需要具体数据）


def export_test_package_from_template(test_package_data, system_data, subsystem_data, include_attachments=False):
    """
    基于模板导出试压包（使用Excel COM方案）
    使用COM操作模板，完美保留所有格式包括RichText蓝色俄语
    
    Args:
        test_package_data: 试压包数据
        system_data: 系统数据
        subsystem_data: 子系统数据
        include_attachments: 是否包含附件
    
    Returns:
        Flask Response对象（Excel或ZIP文件）
    """
    import shutil
    import tempfile
    import win32com.client
    import pythoncom
    from database import create_connection
    from utils.refresh_aggregated_data import refresh_all_aggregated_data
    
    # 在导出前刷新聚合数据（JointSummary、NDEPWHTStatus、ISODrawingList）
    test_package_id = test_package_data.get('TestPackageID')
    if test_package_id:
        refresh_all_aggregated_data(test_package_id)
    
    # 使用信号量限制并发（最多5个Excel COM实例）
    with _EXCEL_COM_SEMAPHORE:
        # 初始化COM
        pythoncom.CoInitialize()
        
        try:
            # 模板文件路径
            template_path = os.path.join(BASE_DIR, 'nordinfo', '试压包模板(202511).xlsx')
            
            # 创建临时文件
            temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            temp_file.close()
            
            try:
                # 复制模板到临时文件
                shutil.copy2(template_path, temp_file.name)
                
                # 打开Excel COM
                try:
                    excel = win32com.client.Dispatch('Excel.Application')
                    excel.Visible = False
                    excel.DisplayAlerts = False
                except Exception as e:
                    import logging
                    import pywintypes
                    logger = logging.getLogger(__name__)
                    
                    # 捕获COM错误并提供详细的错误信息
                    if isinstance(e, pywintypes.com_error):
                        error_code = e.args[0] if e.args else None
                        if error_code == -2147221005:
                            error_msg = (
                                "Excel COM 组件初始化失败 (错误代码: -2147221005 - 无效的类字符串)。\n"
                                "可能的原因和解决方案：\n"
                                "1. Excel 未安装：请在服务器上安装 Microsoft Excel\n"
                                "2. COM 组件注册损坏：请以管理员身份运行以下命令修复：\n"
                                "   - excel.exe /regserver\n"
                                "   或重新安装 Office\n"
                                "3. 权限问题：确保应用程序有权限访问 Excel COM 组件\n"
                                "4. 如果是在 Web 服务器上运行，请检查 IIS/服务账户的 DCOM 配置"
                            )
                        else:
                            error_msg = f"Excel COM 组件初始化失败：{str(e)}"
                    else:
                        error_msg = f"无法启动 Excel 应用程序：{str(e)}"
                    
                    logger.error(error_msg)
                    
                    # 返回友好的错误响应
                    from flask import jsonify
                    return jsonify({
                        'error': 'Excel COM 组件初始化失败',
                        'message': error_msg,
                        'suggestion': '请联系系统管理员检查服务器上的 Excel 安装和配置'
                    }), 500
                
                # 打开模板文件
                wb = excel.Workbooks.Open(temp_file.name, ReadOnly=False, UpdateLinks=False)
                
                try:
                    # 查询数据
                    test_package_id = test_package_data.get('TestPackageID')
                    checklist_status = check_checklist_status(test_package_id)
                    
                    pid_list = []
                    iso_list = []
                    
                    conn = create_connection()
                    if conn:
                        try:
                            cur = conn.cursor(dictionary=True)
                            
                            # 查询P&ID清单
                            cur.execute("""
                                SELECT PIDNo as DrawingNumber, RevNo as RevisionNumber 
                                FROM PIDList 
                                WHERE TestPackageID = %s 
                                ORDER BY PIDNo
                            """, (test_package_id,))
                            pid_list = cur.fetchall()
                            
                            # 查询ISO清单
                            cur.execute("""
                                SELECT ISODrawingNo as DrawingNumber, RevNo as RevisionNumber
                                FROM ISODrawingList 
                                WHERE TestPackageID = %s
                                ORDER BY ISODrawingNo
                            """, (test_package_id,))
                            iso_list = cur.fetchall()
                            
                            # 如果ISODrawingList为空，则从WeldingList提取
                            if not iso_list:
                                cur.execute("""
                                    SELECT DISTINCT DrawingNumber, '' as RevisionNumber
                                    FROM WeldingList 
                                    WHERE TestPackageID = %s 
                                    AND DrawingNumber IS NOT NULL 
                                    AND DrawingNumber != ''
                                    AND (DrawingNumber LIKE '%ISO%' OR DrawingNumber LIKE '%IS0%')
                                    ORDER BY DrawingNumber
                                """, (test_package_id,))
                                iso_list = cur.fetchall()
                            
                        finally:
                            conn.close()
                    
                    # 填充Cover Page（使用COM）
                    ws_cover = wb.Sheets('1.Cover')
                    fill_cover_page_com(ws_cover, test_package_data, system_data, subsystem_data)
                    
                    # 填充Contents Page（使用COM）
                    ws_contents = wb.Sheets('2.Contents')
                    fill_contents_page_com(ws_contents, test_package_data, system_data, subsystem_data, checklist_status)
                    
                    # 填充P&ID-ISO List Page（使用COM）
                    ws_pid_iso = wb.Sheets('3.P&ID-ISO List')
                    fill_pid_iso_page_com(ws_pid_iso, test_package_data, system_data, subsystem_data, pid_list, iso_list)
                    
                    # 填充Welding log Page（使用COM）
                    ws_welding = wb.Sheets('5.Welding log')
                    fill_welding_log_com(ws_welding, test_package_data, system_data, subsystem_data, test_package_id)
                    
                    # 填充Punch List Page（使用COM）
                    ws_punch = wb.Sheets('6.Punch List')
                    fill_punch_list_com(ws_punch, test_package_data, system_data, subsystem_data, test_package_id)
                    
                    # 填充剩余7个工作表的基础信息
                    fill_simple_sheet_basic_info(wb.Sheets('4.Legend'), test_package_data, system_data, subsystem_data)
                    fill_simple_sheet_basic_info(wb.Sheets('7.Test flow chart'), test_package_data, system_data, subsystem_data)
                    fill_simple_sheet_basic_info(wb.Sheets('8.Test Check List'), test_package_data, system_data, subsystem_data)
                    fill_simple_sheet_basic_info(wb.Sheets('9. Calibration certificates'), test_package_data, system_data, subsystem_data)
                    fill_simple_sheet_basic_info(wb.Sheets('10.Pipeline test certificate'), test_package_data, system_data, subsystem_data)
                    fill_simple_sheet_basic_info(wb.Sheets('11.Pipeline Flushing (Purging) '), test_package_data, system_data, subsystem_data)
                    fill_simple_sheet_basic_info(wb.Sheets('12.Reinstatement Check List'), test_package_data, system_data, subsystem_data)
                    
                    # 保存文件
                    wb.Save()
                    
                finally:
                    # 关闭Excel
                    wb.Close(SaveChanges=False)
                    excel.Quit()
                
                # 生成文件名
                package_id = test_package_data.get('TestPackageID', 'unknown').replace('/', '-')
                
                if not include_attachments:
                    # 仅导出Excel
                    with open(temp_file.name, 'rb') as f:
                        excel_data = f.read()
                    
                    filename = f"Test_Package_{package_id}.xlsx"
                    
                    return Response(
                        excel_data,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': f'attachment; filename={filename}'}
                    )
                else:
                    # 导出Excel + 附件（ZIP压缩包）
                    return export_test_package_with_attachments_from_template(temp_file.name, test_package_id, package_id)
            
            finally:
                # 清理临时文件
                try:
                    os.unlink(temp_file.name)
                except:
                    pass
        
        finally:
            # 清理COM
            pythoncom.CoUninitialize()


def export_test_package_with_attachments_from_template(excel_path, test_package_id, package_id):
    """导出试压包Excel和附件的ZIP压缩包（基于模板方案）"""
    import zipfile
    import tempfile
    import shutil
    from database import create_connection
    
    # 创建临时目录
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Excel文件名
        excel_filename = f"Test_Package_{package_id}.xlsx"
        excel_dest = os.path.join(temp_dir, excel_filename)
        shutil.copy2(excel_path, excel_dest)
        
        # 创建附件目录
        attachments_dir = os.path.join(temp_dir, 'Attachments')
        os.makedirs(attachments_dir, exist_ok=True)
        
        # 附件模块映射（与Checklist对应）
        module_mapping = {
            'PID_Drawings': '3.0_Piping_Instrument_Diagram',
            'ISO_Drawings': '4.0_Piping_Hydro_Test_Isometric_Drawing',
            'Symbols_Legend': '5.0_Test_Package_Symbols_Legend',
            'Test_Flow_Chart': '8.0_Test_Flow_Chart',
            'Test_Check_List': '9.0_Pressure_Test_Check_List',
            'Calibration_Certificates': '10.0_Pressure_Test_Gauges_Calibration_Certificates',
            'Test_Certificate': '11.0_Pipeline_Test_Certificate',
            'Flushing_Certificate': '12.0_Pipeline_Flushing_Purging_Certificate',
            'Reinstatement_Check_List': '13.0_Reinstatement_Check_List',
            'Others': '14.0_Others'
        }
        
        # 查询并复制附件
        conn = create_connection()
        if conn:
            try:
                cur = conn.cursor(dictionary=True)
                cur.execute("""
                    SELECT ModuleName, FileName, FilePath 
                    FROM TestPackageAttachments 
                    WHERE TestPackageID = %s
                    ORDER BY ModuleName, UploadedAt
                """, (test_package_id,))
                attachments = cur.fetchall()
                
                for att in attachments:
                    module_name = att['ModuleName']
                    file_path = att['FilePath']
                    file_name = att['FileName']
                    
                    if module_name in module_mapping and os.path.exists(file_path):
                        # 创建模块目录
                        module_dir = os.path.join(attachments_dir, module_mapping[module_name])
                        os.makedirs(module_dir, exist_ok=True)
                        
                        # 复制文件
                        dest_path = os.path.join(module_dir, file_name)
                        shutil.copy2(file_path, dest_path)
                
            finally:
                conn.close()
        
        # 创建ZIP压缩包
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 添加Excel文件
            zip_file.write(excel_dest, excel_filename)
            
            # 添加所有附件
            for root, dirs, files in os.walk(attachments_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zip_file.write(file_path, arcname)
        
        zip_buffer.seek(0)
        
        # 返回ZIP文件
        zip_filename = f"Test_Package_{package_id}_with_Attachments.zip"
        
        return Response(
            zip_buffer.getvalue(),
            mimetype='application/zip',
            headers={'Content-Disposition': f'attachment; filename={zip_filename}'}
        )
        
    finally:
        # 清理临时目录
        try:
            shutil.rmtree(temp_dir)
        except:
            pass
